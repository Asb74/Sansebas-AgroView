import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import DB_PEDIDOS
from db.connection import get_runtime_database_path


def format_forfait_label(row: dict[str, Any] | None) -> str:
    if not row:
        return ""
    desc = re.sub(r"\s+", " ", str(row.get("DescripcionForfait") or "")).strip(" |")
    material = re.sub(
        r"\s+",
        " ",
        str(row.get("MaterialEnvase") or row.get("Material") or ""),
    ).strip(" |")
    if material and "|" in desc:
        parts = [part.strip() for part in desc.split("|") if part.strip()]
        if len(parts) > 1:
            suffix = parts[-1].upper()
            material_norm = material.upper()
            if suffix == material_norm or suffix in material_norm or material_norm in suffix:
                desc = " | ".join(parts[:-1]).strip(" |")
    if not desc:
        return material
    if not material:
        return desc
    if material.upper() in desc.upper():
        return desc
    return f"{desc} | {material}"


class ForfaitRepository:
    TABLE_FORFAIT = "ForfaitConfeccion"
    TABLE_EQUIV = "EquivalenciaForfaitConfeccion"
    TABLE_RELATED = "ForfaitConfeccionRelacionada"
    RELATED_REQUIRED_COLUMNS = [
        "Campaña",
        "Cultivo",
        "IdConfeccion",
        "GRUPO",
        "Eur/kg Material",
        "Eur/kg Recoleción y Transporte",
        "Eur/kg Gastos Generales",
        "Eur/kg Mano obra",
        "Eur/kg total",
    ]
    EDITABLE_FORFAIT_FIELDS = {
        "GrupoForfait",
        "KgForfait",
        "UnidadesForfait",
        "KgUnidad",
        "Medidas",
        "TipoEnvase",
        "MaterialEnvase",
        "DescripcionForfait",
        "DescripcionNormalizada",
    }
    FORFAIT_KEY_FIELDS = {
        "GrupoForfait",
        "KgForfait",
        "UnidadesForfait",
        "KgUnidad",
        "Medidas",
        "TipoEnvase",
        "MaterialEnvase",
    }

    FORFAIT_EXTRA_COLUMNS = {
        "GrupoForfait": "TEXT",
        "KgForfait": "REAL",
        "UnidadesForfait": "INTEGER",
        "KgUnidad": "REAL",
        "Medidas": "TEXT",
        "TipoEnvase": "TEXT",
        "MaterialEnvase": "TEXT",
        "ColorEnvase": "TEXT",
        "DescripcionNormalizada": "TEXT",
        "ClaveForfait": "TEXT",
    }
    EQUIV_EXTRA_COLUMNS = {
        "GrupoForfait": "TEXT",
        "KgForfait": "REAL",
        "UnidadesForfait": "INTEGER",
        "KgUnidad": "REAL",
        "Medidas": "TEXT",
        "TipoEnvase": "TEXT",
        "MaterialEnvase": "TEXT",
        "ClaveForfait": "TEXT",
        "ConfianzaSugerencia": "REAL",
    }
    RELATED_EXTRA_COLUMNS = {
        "NombreConfeccion": "TEXT",
        "GrupoConfeccion": "TEXT",
        "Marca": "TEXT",
    }

    def __init__(self, db_calc: str | None = None, db_pedidos: str | None = None) -> None:
        self.db_calc = Path(db_calc) if db_calc else get_runtime_database_path("DBAgroViewCalc.sqlite")
        self.db_pedidos = Path(db_pedidos) if db_pedidos else get_runtime_database_path(DB_PEDIDOS)
        self.initialize()

    def initialize(self) -> None:
        self.db_calc.parent.mkdir(parents=True, exist_ok=True)
        with self._connect_calc() as conn:
            conn.execute(
                f"""
                CREATE TABLE IF NOT EXISTS "{self.TABLE_FORFAIT}" (
                    Id INTEGER PRIMARY KEY AUTOINCREMENT,
                    Cultivo TEXT NOT NULL,
                    "Campaña" TEXT NOT NULL,
                    NombreForfait TEXT,
                    DescripcionForfait TEXT,
                    GrupoConfeccion TEXT,
                    Material TEXT,
                    CosteRecoleccionTransporteCentKg REAL,
                    CosteConfeccionCentKg REAL,
                    CosteGeneralCentKg REAL,
                    CosteTotalCentKg REAL,
                    CosteRecoleccionTransporteEurKg REAL,
                    CosteConfeccionEurKg REAL,
                    CosteGeneralEurKg REAL,
                    CosteTotalEurKg REAL,
                    OrigenArchivo TEXT,
                    FechaImportacion TEXT,
                    GrupoForfait TEXT,
                    KgForfait REAL,
                    UnidadesForfait INTEGER,
                    KgUnidad REAL,
                    Medidas TEXT,
                    TipoEnvase TEXT,
                    MaterialEnvase TEXT,
                    ColorEnvase TEXT,
                    DescripcionNormalizada TEXT,
                    ClaveForfait TEXT
                )
                """
            )
            conn.execute(
                f"""
                CREATE TABLE IF NOT EXISTS "{self.TABLE_EQUIV}" (
                    Id INTEGER PRIMARY KEY AUTOINCREMENT,
                    Cultivo TEXT NOT NULL,
                    "Campaña" TEXT NOT NULL,
                    ConfeccionPedido TEXT NOT NULL,
                    GrupoConfeccion TEXT,
                    NombreConfeccion TEXT,
                    DescripcionCorta TEXT,
                    Neto REAL,
                    Marca TEXT,
                    DescripcionForfait TEXT,
                    CosteConfeccionEurKg REAL,
                    CosteTotalEurKg REAL,
                    Estado TEXT,
                    Observaciones TEXT,
                    FechaActualizacion TEXT,
                    GrupoForfait TEXT,
                    KgForfait REAL,
                    UnidadesForfait INTEGER,
                    KgUnidad REAL,
                    Medidas TEXT,
                    TipoEnvase TEXT,
                    MaterialEnvase TEXT,
                    ClaveForfait TEXT,
                    ConfianzaSugerencia REAL
                )
                """
            )
            conn.execute(
                f"""
                CREATE TABLE IF NOT EXISTS "{self.TABLE_RELATED}" (
                    Id INTEGER PRIMARY KEY AUTOINCREMENT,
                    "Campaña" TEXT NOT NULL,
                    Cultivo TEXT NOT NULL,
                    Variedad TEXT NOT NULL,
                    Condicion1 TEXT NOT NULL,
                    IdConfeccion TEXT NOT NULL,
                    Grupo TEXT,
                    NombreConfeccion TEXT,
                    GrupoConfeccion TEXT,
                    Marca TEXT,
                    CosteMaterialEurKg REAL,
                    CosteRecoleccionTransporteEurKg REAL,
                    CosteGastosGeneralesEurKg REAL,
                    CosteManoObraEurKg REAL,
                    CosteTotalEurKg REAL,
                    Estado TEXT,
                    Observaciones TEXT,
                    OrigenArchivo TEXT,
                    HojaOrigen TEXT,
                    FechaImportacion TEXT
                )
                """
            )
            self._ensure_columns(conn, self.TABLE_FORFAIT, self.FORFAIT_EXTRA_COLUMNS)
            self._ensure_columns(conn, self.TABLE_EQUIV, self.EQUIV_EXTRA_COLUMNS)
            self._ensure_columns(conn, self.TABLE_RELATED, self.RELATED_EXTRA_COLUMNS)
            conn.execute("DROP INDEX IF EXISTS ux_forfait_confeccion")
            conn.execute(
                f"""
                CREATE UNIQUE INDEX IF NOT EXISTS ux_forfait_confeccion_clave
                ON "{self.TABLE_FORFAIT}" (Cultivo, "Campaña", ClaveForfait)
                """
            )
            conn.execute(
                f"""
                CREATE UNIQUE INDEX IF NOT EXISTS ux_equiv_forfait_confeccion
                ON "{self.TABLE_EQUIV}" (Cultivo, "Campaña", ConfeccionPedido)
                """
            )
            conn.execute(
                f"""
                DELETE FROM "{self.TABLE_RELATED}"
                WHERE Id NOT IN (
                    SELECT MAX(Id)
                    FROM "{self.TABLE_RELATED}"
                    GROUP BY "Campaña", Cultivo, IdConfeccion
                )
                """
            )
            conn.execute(
                f"""
                CREATE UNIQUE INDEX IF NOT EXISTS ux_forfait_relacionado_confeccion
                ON "{self.TABLE_RELATED}" ("Campaña", Cultivo, IdConfeccion)
                """
            )
            conn.execute("DROP INDEX IF EXISTS ux_forfait_related_logic")
            conn.commit()

    def fetch_excel_sheet_names(self, file_path: str) -> list[str]:
        wb = load_workbook(Path(file_path), data_only=True, read_only=True)
        return list(wb.sheetnames)

    def validate_related_forfait_sheet(self, file_path: str, sheet_name: str) -> tuple[bool, list[str], list[str]]:
        wb = load_workbook(Path(file_path), data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"No existe la hoja {sheet_name}.")
        ws = wb[sheet_name]
        header_values = [self._clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        expected = list(self.RELATED_REQUIRED_COLUMNS)
        normalized = [h.replace("Eur/kg Recolección y Transporte", "Eur/kg Recoleción y Transporte") for h in header_values]
        missing = [col for col in expected if col not in normalized]
        return not missing, expected, header_values

    def import_related_forfait_excel(self, file_path: str, sheet_name: str) -> dict[str, Any]:
        ok, expected, found = self.validate_related_forfait_sheet(file_path, sheet_name)
        if not ok:
            raise ValueError(
                "La hoja seleccionada no tiene la estructura de forfait relacionado.\n"
                f"Esperadas: {expected}\nEncontradas: {found}"
            )
        wb = load_workbook(Path(file_path), data_only=True, read_only=True)
        ws = wb[sheet_name]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nuevos = actualizados = revisar = errores = 0
        imported_keys: list[tuple[str, str, str]] = []
        with self._connect_calc() as conn:
            header_values = [self._clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
            col_map = {name: idx for idx, name in enumerate(header_values)}
            if "Eur/kg Recoleción y Transporte" not in col_map and "Eur/kg Recolección y Transporte" in col_map:
                col_map["Eur/kg Recoleción y Transporte"] = col_map["Eur/kg Recolección y Transporte"]
            for row_num in range(2, ws.max_row + 1):
                vals = [ws.cell(row_num, i).value for i in range(1, ws.max_column + 1)]
                if not any(str(v or "").strip() for v in vals):
                    continue
                campana = self._clean_text(vals[col_map["Campaña"]]) if "Campaña" in col_map else ""
                cultivo = self._clean_text(vals[col_map["Cultivo"]]) if "Cultivo" in col_map else ""
                id_conf = self._clean_text(vals[col_map["IdConfeccion"]]) if "IdConfeccion" in col_map else ""
                if not id_conf:
                    errores += 1
                    continue
                grupo = self._clean_text(vals[col_map["GRUPO"]]) if "GRUPO" in col_map else ""
                coste_material = self._to_float(vals[col_map["Eur/kg Material"]]) if "Eur/kg Material" in col_map else None
                coste_recoleccion = self._to_float(vals[col_map["Eur/kg Recoleción y Transporte"]]) if "Eur/kg Recoleción y Transporte" in col_map else None
                coste_gastos = self._to_float(vals[col_map["Eur/kg Gastos Generales"]]) if "Eur/kg Gastos Generales" in col_map else None
                coste_mano_obra = self._to_float(vals[col_map["Eur/kg Mano obra"]]) if "Eur/kg Mano obra" in col_map else None
                coste_total = self._to_float(vals[col_map["Eur/kg total"]]) if "Eur/kg total" in col_map else None
                missing_any_cost = any(v is None for v in [coste_material, coste_recoleccion, coste_gastos, coste_mano_obra, coste_total])
                estado = "IMPORTADO" if (coste_total is not None and not missing_any_cost) else "REVISAR"
                if estado != "IMPORTADO":
                    revisar += 1
                variedad = self._clean_text(vals[col_map["Variedad"]]) if "Variedad" in col_map else "TODAS"
                condicion1 = self._clean_text(vals[col_map["Condicion1"]]) if "Condicion1" in col_map else "TODAS"
                exists = conn.execute(
                    f'SELECT 1 FROM "{self.TABLE_RELATED}" WHERE "Campaña"=? AND Cultivo=? AND IdConfeccion=?',
                    [campana, cultivo, id_conf],
                ).fetchone()
                conn.execute(
                    f"""
                    INSERT INTO "{self.TABLE_RELATED}" ("Campaña", Cultivo, Variedad, Condicion1, IdConfeccion, Grupo, NombreConfeccion, Marca,
                        CosteMaterialEurKg, CosteRecoleccionTransporteEurKg, CosteGastosGeneralesEurKg, CosteManoObraEurKg, CosteTotalEurKg,
                        Estado, OrigenArchivo, HojaOrigen, FechaImportacion)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT("Campaña", Cultivo, IdConfeccion) DO UPDATE SET
                        Grupo=excluded.Grupo,
                        CosteMaterialEurKg=excluded.CosteMaterialEurKg,
                        CosteRecoleccionTransporteEurKg=excluded.CosteRecoleccionTransporteEurKg,
                        CosteGastosGeneralesEurKg=excluded.CosteGastosGeneralesEurKg,
                        CosteManoObraEurKg=excluded.CosteManoObraEurKg,
                        CosteTotalEurKg=excluded.CosteTotalEurKg,
                        Estado=excluded.Estado,
                        OrigenArchivo=excluded.OrigenArchivo,
                        HojaOrigen=excluded.HojaOrigen,
                        FechaImportacion=excluded.FechaImportacion
                    """,
                    [campana, cultivo, variedad or "TODAS", condicion1 or "TODAS", id_conf, grupo, "", "", coste_material, coste_recoleccion, coste_gastos, coste_mano_obra, coste_total, estado, str(file_path), sheet_name, now],
                )
                imported_keys.append((campana, cultivo, id_conf))
                actualizados += 1 if exists else 0
                nuevos += 0 if exists else 1
            conn.commit()
            rows = []
            for key in imported_keys:
                rec = conn.execute(
                    f'SELECT * FROM "{self.TABLE_RELATED}" WHERE "Campaña"=? AND Cultivo=? AND IdConfeccion=?',
                    list(key),
                ).fetchone()
                if rec:
                    rows.append(dict(rec))
        return {"nuevos": nuevos, "actualizados": actualizados, "revisar": revisar, "errores": errores, "rows": rows}

    def import_forfait_excel(
        self,
        file_path: str,
        cultivo: str,
        campana: str,
        sheet_name: str = "NARANJA",
    ) -> tuple[int, int, list[dict[str, Any]]]:
        cultivo_norm = self._norm_key(cultivo)
        campana_norm = str(campana or "").strip()
        if not cultivo_norm or not campana_norm:
            raise ValueError("Cultivo y campaña son obligatorios.")

        path = Path(file_path)
        wb = load_workbook(path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"No existe la hoja {sheet_name}. Hojas: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]

        header_row = 2
        material_row = 3
        row_recoleccion = self._find_row(ws, "TOTAL M.O. RECOLECCION Y TRANSPORTE")
        row_confeccion = self._find_row(ws, "TOTAL COSTES DE CONFECCION")
        row_general = self._find_row(ws, "3. COSTES GENERALES")
        row_total = self._find_row(ws, "TOTAL", exact=True)
        if not row_confeccion or not row_total:
            raise ValueError("No se encontraron las filas TOTAL COSTES DE CONFECCION y TOTAL.")

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows: list[dict[str, Any]] = []
        seen_keys: set[str] = set()
        for col in range(2, ws.max_column + 1):
            desc_base = self._clean_text(ws.cell(header_row, col).value)
            material_raw = self._clean_text(ws.cell(material_row, col).value)
            if not desc_base:
                continue
            coste_confeccion = self._to_float(ws.cell(row_confeccion, col).value)
            coste_total = self._to_float(ws.cell(row_total, col).value)
            if coste_confeccion is None and coste_total is None:
                continue
            if (coste_confeccion or 0) <= 0 and (coste_total or 0) <= 0:
                continue

            parsed = self.parse_forfait_description(desc_base, material_raw, cultivo_norm, campana_norm)
            if parsed["ClaveForfait"] in seen_keys:
                continue
            seen_keys.add(parsed["ClaveForfait"])

            coste_recoleccion = self._to_float(ws.cell(row_recoleccion, col).value) if row_recoleccion else None
            coste_general = self._to_float(ws.cell(row_general, col).value) if row_general else None
            row = {
                "Cultivo": cultivo_norm,
                "Campaña": campana_norm,
                "NombreForfait": sheet_name,
                "DescripcionForfait": parsed["DescripcionForfait"],
                "GrupoConfeccion": parsed["GrupoForfait"],
                "Material": material_raw,
                "CosteRecoleccionTransporteCentKg": coste_recoleccion,
                "CosteConfeccionCentKg": coste_confeccion,
                "CosteGeneralCentKg": coste_general,
                "CosteTotalCentKg": coste_total,
                "CosteRecoleccionTransporteEurKg": self._cent_to_eur(coste_recoleccion),
                "CosteConfeccionEurKg": self._cent_to_eur(coste_confeccion),
                "CosteGeneralEurKg": self._cent_to_eur(coste_general),
                "CosteTotalEurKg": self._cent_to_eur(coste_total),
                "OrigenArchivo": str(path),
                "FechaImportacion": now,
                **parsed,
            }
            rows.append(row)

        inserted = 0
        updated = 0
        with self._connect_calc() as conn:
            for row in rows:
                exists = conn.execute(
                    f"""
                    SELECT 1 FROM "{self.TABLE_FORFAIT}"
                    WHERE Cultivo = ? AND "Campaña" = ? AND ClaveForfait = ?
                    """,
                    [row["Cultivo"], row["Campaña"], row["ClaveForfait"]],
                ).fetchone()
                conn.execute(
                    f"""
                    INSERT INTO "{self.TABLE_FORFAIT}" (
                        Cultivo, "Campaña", NombreForfait, DescripcionForfait, GrupoConfeccion, Material,
                        CosteRecoleccionTransporteCentKg, CosteConfeccionCentKg, CosteGeneralCentKg,
                        CosteTotalCentKg, CosteRecoleccionTransporteEurKg, CosteConfeccionEurKg,
                        CosteGeneralEurKg, CosteTotalEurKg, OrigenArchivo, FechaImportacion,
                        GrupoForfait, KgForfait, UnidadesForfait, KgUnidad, Medidas, TipoEnvase,
                        MaterialEnvase, ColorEnvase, DescripcionNormalizada, ClaveForfait
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(Cultivo, "Campaña", ClaveForfait) DO UPDATE SET
                        NombreForfait=excluded.NombreForfait,
                        DescripcionForfait=excluded.DescripcionForfait,
                        GrupoConfeccion=excluded.GrupoConfeccion,
                        Material=excluded.Material,
                        CosteRecoleccionTransporteCentKg=excluded.CosteRecoleccionTransporteCentKg,
                        CosteConfeccionCentKg=excluded.CosteConfeccionCentKg,
                        CosteGeneralCentKg=excluded.CosteGeneralCentKg,
                        CosteTotalCentKg=excluded.CosteTotalCentKg,
                        CosteRecoleccionTransporteEurKg=excluded.CosteRecoleccionTransporteEurKg,
                        CosteConfeccionEurKg=excluded.CosteConfeccionEurKg,
                        CosteGeneralEurKg=excluded.CosteGeneralEurKg,
                        CosteTotalEurKg=excluded.CosteTotalEurKg,
                        OrigenArchivo=excluded.OrigenArchivo,
                        FechaImportacion=excluded.FechaImportacion,
                        GrupoForfait=excluded.GrupoForfait,
                        KgForfait=excluded.KgForfait,
                        UnidadesForfait=excluded.UnidadesForfait,
                        KgUnidad=excluded.KgUnidad,
                        Medidas=excluded.Medidas,
                        TipoEnvase=excluded.TipoEnvase,
                        MaterialEnvase=excluded.MaterialEnvase,
                        ColorEnvase=excluded.ColorEnvase,
                        DescripcionNormalizada=excluded.DescripcionNormalizada
                    """,
                    [
                        row["Cultivo"],
                        row["Campaña"],
                        row["NombreForfait"],
                        row["DescripcionForfait"],
                        row["GrupoConfeccion"],
                        row["Material"],
                        row["CosteRecoleccionTransporteCentKg"],
                        row["CosteConfeccionCentKg"],
                        row["CosteGeneralCentKg"],
                        row["CosteTotalCentKg"],
                        row["CosteRecoleccionTransporteEurKg"],
                        row["CosteConfeccionEurKg"],
                        row["CosteGeneralEurKg"],
                        row["CosteTotalEurKg"],
                        row["OrigenArchivo"],
                        row["FechaImportacion"],
                        row["GrupoForfait"],
                        row["KgForfait"],
                        row["UnidadesForfait"],
                        row["KgUnidad"],
                        row["Medidas"],
                        row["TipoEnvase"],
                        row["MaterialEnvase"],
                        row["ColorEnvase"],
                        row["DescripcionNormalizada"],
                        row["ClaveForfait"],
                    ],
                )
                if exists:
                    updated += 1
                else:
                    inserted += 1
            conn.commit()
            if rows:
                keys = [row["ClaveForfait"] for row in rows]
                placeholders = ",".join("?" for _ in keys)
                db_rows = conn.execute(
                    f"""
                    SELECT *
                    FROM "{self.TABLE_FORFAIT}"
                    WHERE Cultivo = ? AND "Campaña" = ? AND ClaveForfait IN ({placeholders})
                    """,
                    [cultivo_norm, campana_norm, *keys],
                ).fetchall()
                by_key = {str(row["ClaveForfait"]): dict(row) for row in db_rows}
                rows = [by_key.get(str(row["ClaveForfait"]), row) for row in rows]
        return inserted, updated, rows

    def fetch_forfaits(self, cultivo: str, campana: str) -> list[dict[str, Any]]:
        with self._connect_calc() as conn:
            rows = conn.execute(
                f"""
                SELECT *
                FROM "{self.TABLE_FORFAIT}"
                WHERE Cultivo = ? AND "Campaña" = ?
                ORDER BY GrupoForfait, KgForfait, Medidas, MaterialEnvase, DescripcionForfait
                """,
                [self._norm_key(cultivo), str(campana or "").strip()],
            ).fetchall()
        return [dict(r) for r in rows]

    def update_forfait_field(self, id_forfait: int, field_name: str, value: Any) -> dict[str, Any]:
        if field_name not in self.EDITABLE_FORFAIT_FIELDS:
            raise ValueError(f"El campo {field_name} no es editable.")
        row = self._get_forfait_by_id(id_forfait)
        if not row:
            raise ValueError("No se encontró el forfait a editar.")
        row[field_name] = self._prepare_forfait_value(field_name, value)
        return self.update_forfait_row(id_forfait, row)

    def update_forfait_row(self, id_forfait: int, data: dict[str, Any]) -> dict[str, Any]:
        current = self._get_forfait_by_id(id_forfait)
        if not current:
            raise ValueError("No se encontró el forfait a editar.")
        row = dict(current)
        for field in self.EDITABLE_FORFAIT_FIELDS:
            if field in data:
                row[field] = self._prepare_forfait_value(field, data.get(field))
        new_key = self.regenerate_forfait_key(row)
        old_key = str(current.get("ClaveForfait") or "")
        if not self.validate_unique_forfait_key(row.get("Cultivo"), row.get("Campaña"), new_key, id_forfait):
            raise ValueError(f"Ya existe un forfait con la clave {new_key} para este cultivo y campaña.")
        with self._connect_calc() as conn:
            conn.execute(
                f"""
                UPDATE "{self.TABLE_FORFAIT}"
                SET DescripcionForfait = ?, GrupoForfait = ?, KgForfait = ?,
                    UnidadesForfait = ?, KgUnidad = ?, Medidas = ?, TipoEnvase = ?,
                    MaterialEnvase = ?, DescripcionNormalizada = ?, ClaveForfait = ?
                WHERE Id = ?
                """,
                [
                    row.get("DescripcionForfait"),
                    row.get("GrupoForfait"),
                    row.get("KgForfait"),
                    row.get("UnidadesForfait"),
                    row.get("KgUnidad"),
                    row.get("Medidas"),
                    row.get("TipoEnvase"),
                    row.get("MaterialEnvase"),
                    row.get("DescripcionNormalizada"),
                    new_key,
                    id_forfait,
                ],
            )
            if old_key and old_key != new_key:
                conn.execute(
                    f"""
                    UPDATE "{self.TABLE_EQUIV}"
                    SET ClaveForfait = ?, DescripcionForfait = ?, CosteConfeccionEurKg = ?,
                        CosteTotalEurKg = ?, GrupoForfait = ?, KgForfait = ?,
                        UnidadesForfait = ?, KgUnidad = ?, Medidas = ?, TipoEnvase = ?,
                        MaterialEnvase = ?
                    WHERE Cultivo = ? AND "Campaña" = ? AND ClaveForfait = ?
                    """,
                    [
                        new_key,
                        row.get("DescripcionForfait"),
                        row.get("CosteConfeccionEurKg"),
                        row.get("CosteTotalEurKg"),
                        row.get("GrupoForfait"),
                        row.get("KgForfait"),
                        row.get("UnidadesForfait"),
                        row.get("KgUnidad"),
                        row.get("Medidas"),
                        row.get("TipoEnvase"),
                        row.get("MaterialEnvase"),
                        row.get("Cultivo"),
                        row.get("Campaña"),
                        old_key,
                    ],
                )
            conn.commit()
        updated = self._get_forfait_by_id(id_forfait)
        if not updated:
            raise ValueError("No se pudo recargar el forfait editado.")
        return updated

    def regenerate_forfait_key(self, row: dict[str, Any]) -> str:
        cultivo = self._norm_key(row.get("Cultivo"))
        campana = self._clean_text(row.get("Campaña"))
        grupo = self._norm_key(row.get("GrupoForfait")) or "PENDIENTE_REVISION"
        unidades = self._to_float(row.get("UnidadesForfait"))
        kg_unidad = self._to_float(row.get("KgUnidad"))
        kg_total = self._to_float(row.get("KgForfait"))
        if unidades and kg_unidad:
            kg_part = f"{int(unidades)}x{self._format_number(kg_unidad)}"
        else:
            kg_part = self._format_number(kg_total) or "SIN_KG"
        medidas = self._clean_text(row.get("Medidas")).lower() or "SIN_MEDIDAS"
        tipo_envase = self._norm_key(row.get("TipoEnvase")) or "SIN_TIPO"
        material_envase = self._norm_key(row.get("MaterialEnvase")) or "SIN_MATERIAL"
        return "|".join([cultivo, campana, grupo, kg_part, medidas, tipo_envase, material_envase])

    def validate_unique_forfait_key(
        self,
        cultivo: str,
        campana: str,
        clave_forfait: str,
        exclude_id: int | None = None,
    ) -> bool:
        params: list[Any] = [
            self._norm_key(cultivo),
            self._clean_text(campana),
            self._clean_text(clave_forfait),
        ]
        extra_where = ""
        if exclude_id is not None:
            extra_where = "AND Id <> ?"
            params.append(int(exclude_id))
        with self._connect_calc() as conn:
            row = conn.execute(
                f"""
                SELECT 1
                FROM "{self.TABLE_FORFAIT}"
                WHERE Cultivo = ? AND "Campaña" = ? AND ClaveForfait = ?
                {extra_where}
                """,
                params,
            ).fetchone()
        return row is None

    def fetch_mapping_rows(self, cultivo: str, campana: str) -> list[dict[str, Any]]:
        cultivo_norm = self._norm_key(cultivo)
        campana_norm = str(campana or "").strip()
        if not cultivo_norm or not campana_norm:
            return []
        self.ensure_equivalence_rows(cultivo_norm, campana_norm)
        forfaits = self.fetch_forfaits(cultivo_norm, campana_norm)
        with self._connect_calc() as conn:
            rows = conn.execute(
                f"""
                SELECT *
                FROM "{self.TABLE_EQUIV}"
                WHERE Cultivo = ? AND "Campaña" = ?
                ORDER BY ConfeccionPedido
                """,
                [cultivo_norm, campana_norm],
            ).fetchall()
        out = [dict(r) for r in rows]
        for row in out:
            suggestion = self.suggest_forfait(row, forfaits)
            suggestion_label = format_forfait_label(suggestion)
            row.update(
                {
                    "ForfaitSugerido": suggestion_label,
                    "ForfaitSugeridoLabel": suggestion_label,
                    "ClaveForfaitSugerida": suggestion.get("ClaveForfait", ""),
                    "ConfianzaSugerencia": suggestion.get("ConfianzaSugerencia", 0.0),
                    "GrupoForfaitSugerido": suggestion.get("GrupoForfait", ""),
                    "KgForfaitSugerido": suggestion.get("KgForfait"),
                    "MedidasSugeridas": suggestion.get("Medidas", ""),
                    "TipoEnvaseSugerido": suggestion.get("TipoEnvase", ""),
                    "MaterialEnvaseSugerido": suggestion.get("MaterialEnvase", ""),
                    "MotivoSugerencia": suggestion.get("MotivoSugerencia", ""),
                    "SugerenciaAmbigua": suggestion.get("SugerenciaAmbigua", False),
                }
            )
        return out


    def fetch_coverage_rows(self, cultivo: str, campana: str, only_missing: bool = False) -> list[dict[str, Any]]:
        cultivo_norm = self._norm_key(cultivo)
        campana_norm = str(campana or "").strip()
        if not cultivo_norm or not campana_norm:
            return []
        base_rows = self._fetch_pedido_confecciones(cultivo_norm, campana_norm)
        out: list[dict[str, Any]] = []
        for row in base_rows:
            record = self._resolve_related_forfait(campana_norm, cultivo_norm, str(row.get("ConfeccionPedido") or ""))
            coverage = {
                "Campaña": campana_norm,
                "Cultivo": cultivo_norm,
                "IdConfeccion": str(row.get("ConfeccionPedido") or ""),
                "NombreConfeccion": row.get("NombreConfeccion", ""),
                "GrupoConfeccion": row.get("GrupoConfeccion", ""),
                "Marca": row.get("Marca", ""),
                "CosteMaterialEurKg": record.get("CosteMaterialEurKg") if record else None,
                "CosteRecoleccionTransporteEurKg": record.get("CosteRecoleccionTransporteEurKg") if record else None,
                "CosteGastosGeneralesEurKg": record.get("CosteGastosGeneralesEurKg") if record else None,
                "CosteManoObraEurKg": record.get("CosteManoObraEurKg") if record else None,
                "CosteTotalEurKg": record.get("CosteTotalEurKg") if record else None,
                "Estado": record.get("Estado") if record else "SIN_FORFAIT",
                "OrigenCoste": record.get("OrigenCoste") if record else "SIN_FORFAIT",
            }
            if record:
                cost_fields = [
                    record.get("CosteMaterialEurKg"),
                    record.get("CosteRecoleccionTransporteEurKg"),
                    record.get("CosteGastosGeneralesEurKg"),
                    record.get("CosteManoObraEurKg"),
                    record.get("CosteTotalEurKg"),
                ]
                coverage["Estado"] = "OK" if all(v is not None for v in cost_fields) else "REVISAR"
            if not coverage["Estado"]:
                coverage["Estado"] = "SIN_FORFAIT"
            if only_missing and coverage["OrigenCoste"] != "SIN_FORFAIT":
                continue
            out.append(coverage)
        return out

    def _resolve_related_forfait(self, campana: str, cultivo: str, id_confeccion: str) -> dict[str, Any] | None:
        with self._connect_calc() as conn:
            rec = conn.execute(
                f'SELECT * FROM "{self.TABLE_RELATED}" WHERE "Campaña" = ? AND Cultivo = ? AND IdConfeccion = ? ORDER BY Id DESC LIMIT 1',
                [campana, cultivo, id_confeccion],
            ).fetchone()
            if rec:
                row = dict(rec)
                row["OrigenCoste"] = "EXACTO"
                cost_fields = [
                    row.get("CosteMaterialEurKg"),
                    row.get("CosteRecoleccionTransporteEurKg"),
                    row.get("CosteGastosGeneralesEurKg"),
                    row.get("CosteManoObraEurKg"),
                    row.get("CosteTotalEurKg"),
                ]
                row["Estado"] = "OK" if all(v is not None for v in cost_fields) else "REVISAR"
                return row
        return None
    def reset_mapping_rows(self, cultivo: str, campana: str) -> int:
        cultivo_norm = self._norm_key(cultivo)
        campana_norm = str(campana or "").strip()
        if not cultivo_norm or not campana_norm:
            return 0
        with self._connect_calc() as conn:
            cursor = conn.execute(
                f"""
                DELETE FROM "{self.TABLE_EQUIV}"
                WHERE Cultivo = ? AND "Campaña" = ?
                """,
                [cultivo_norm, campana_norm],
            )
            conn.commit()
            return int(cursor.rowcount or 0)

    def fetch_related_forfait(self, cultivo: str | None = None, campana: str | None = None) -> list[dict[str, Any]]:
        where = []
        params: list[Any] = []
        if cultivo:
            where.append("Cultivo = ?")
            params.append(self._clean_text(cultivo))
        if campana:
            where.append('"Campaña" = ?')
            params.append(self._clean_text(campana))
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        with self._connect_calc() as conn:
            rows = conn.execute(
                f'SELECT * FROM "{self.TABLE_RELATED}" {where_sql} ORDER BY "Campaña", Cultivo, IdConfeccion',
                params,
            ).fetchall()
        return [dict(r) for r in rows]

    def update_related_forfait_field(self, id_forfait: int, field_name: str, value: Any) -> dict[str, Any]:
        editable = {
            "Grupo", "NombreConfeccion", "Marca", "CosteMaterialEurKg", "CosteManoObraEurKg", "CosteTotalEurKg", "Estado", "Observaciones",
        }
        if field_name not in editable:
            raise ValueError(f"El campo {field_name} no es editable.")
        cast_value = self._to_float(value) if field_name.startswith("Coste") else self._clean_text(value)
        with self._connect_calc() as conn:
            conn.execute(f'UPDATE "{self.TABLE_RELATED}" SET "{field_name}" = ? WHERE Id = ?', [cast_value, int(id_forfait)])
            conn.commit()
            row = conn.execute(f'SELECT * FROM "{self.TABLE_RELATED}" WHERE Id = ?', [int(id_forfait)]).fetchone()
        if not row:
            raise ValueError("No se encontró el registro.")
        return dict(row)

    def reset_related_forfait(self, cultivo: str, campana: str) -> int:
        with self._connect_calc() as conn:
            cur = conn.execute(f'DELETE FROM "{self.TABLE_RELATED}" WHERE Cultivo = ? AND "Campaña" = ?', [self._clean_text(cultivo), self._clean_text(campana)])
            conn.commit()
            return int(cur.rowcount or 0)

    def ensure_equivalence_rows(self, cultivo: str, campana: str) -> int:
        rows = self._fetch_pedido_confecciones(cultivo, campana)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        inserted = 0
        with self._connect_calc() as conn:
            for row in rows:
                exists = conn.execute(
                    f"""
                    SELECT 1 FROM "{self.TABLE_EQUIV}"
                    WHERE Cultivo = ? AND "Campaña" = ? AND ConfeccionPedido = ?
                    """,
                    [cultivo, campana, row["ConfeccionPedido"]],
                ).fetchone()
                if exists:
                    conn.execute(
                        f"""
                        UPDATE "{self.TABLE_EQUIV}"
                        SET GrupoConfeccion = ?, NombreConfeccion = ?, DescripcionCorta = ?,
                            Neto = ?, Marca = ?
                        WHERE Cultivo = ? AND "Campaña" = ? AND ConfeccionPedido = ?
                        """,
                        [
                            row.get("GrupoConfeccion"),
                            row.get("NombreConfeccion"),
                            row.get("DescripcionCorta"),
                            row.get("Neto"),
                            row.get("Marca"),
                            cultivo,
                            campana,
                            row["ConfeccionPedido"],
                        ],
                    )
                    continue
                conn.execute(
                    f"""
                    INSERT INTO "{self.TABLE_EQUIV}" (
                        Cultivo, "Campaña", ConfeccionPedido, GrupoConfeccion, NombreConfeccion,
                        DescripcionCorta, Neto, Marca, Estado, FechaActualizacion
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'PENDIENTE', ?)
                    """,
                    [
                        cultivo,
                        campana,
                        row["ConfeccionPedido"],
                        row.get("GrupoConfeccion"),
                        row.get("NombreConfeccion"),
                        row.get("DescripcionCorta"),
                        row.get("Neto"),
                        row.get("Marca"),
                        now,
                    ],
                )
                inserted += 1
            conn.commit()
        return inserted

    def update_equivalence(
        self,
        cultivo: str,
        campana: str,
        confeccion_pedido: str,
        clave_forfait: str,
        estado: str,
        observaciones: str = "",
    ) -> None:
        cultivo_norm = self._norm_key(cultivo)
        campana_norm = str(campana or "").strip()
        estado_norm = str(estado or "").strip().upper()
        if estado_norm not in {"PENDIENTE", "VALIDADO", "SIN_EQUIVALENCIA"}:
            raise ValueError("Estado no valido.")
        if estado_norm == "SIN_EQUIVALENCIA":
            clave_forfait = ""
        forfait = self._get_forfait(cultivo_norm, campana_norm, clave_forfait) if clave_forfait else None
        if estado_norm == "VALIDADO" and not forfait:
            raise ValueError("Para validar es obligatorio asignar un forfait del mismo cultivo y campaña.")
        row = {
            "Cultivo": cultivo_norm,
            "Campaña": campana_norm,
            "ConfeccionPedido": str(confeccion_pedido or "").strip(),
            "Estado": estado_norm,
            "Observaciones": observaciones,
            "FechaActualizacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ConfianzaSugerencia": None,
        }
        self._apply_forfait_to_row(row, forfait or {}, estado=estado_norm)
        with self._connect_calc() as conn:
            self._update_equivalence_row(conn, row)
            conn.commit()

    def suggest_forfait(self, mapping_row: dict[str, Any], forfaits: list[dict[str, Any]]) -> dict[str, Any]:
        features = self.parse_confeccion_features(mapping_row)
        candidates: list[tuple[float, dict[str, Any], str]] = []
        for forfait in forfaits:
            score, reasons = self._score_forfait(features, forfait)
            if score <= 0:
                continue
            candidates.append((score, dict(forfait), "; ".join(reasons)))
        if not candidates:
            return {"ConfianzaSugerencia": 0.0}
        candidates.sort(key=lambda item: item[0], reverse=True)
        best_score, best, motivo = candidates[0]
        if len(candidates) > 1 and best_score - candidates[1][0] < 5:
            best_score = min(best_score, 69.0)
            motivo = f"Sugerencia ambigua: {motivo}"
            best["SugerenciaAmbigua"] = True
        best["ConfianzaSugerencia"] = min(best_score, 100.0)
        best["MotivoSugerencia"] = motivo
        return best

    def parse_forfait_description(
        self,
        descripcion: str,
        material: str,
        cultivo: str,
        campana: str,
    ) -> dict[str, Any]:
        base_norm = self._norm_text(descripcion)
        material_norm = self._norm_text(material)
        full_desc = f"{self._clean_text(descripcion)} | {self._clean_text(material)}" if material else self._clean_text(descripcion)
        full_norm = self._norm_text(full_desc)
        grupo = self._detect_group(base_norm)
        unidades = self._extract_units(base_norm)
        kg_unidad = self._extract_kg(base_norm)
        kg_total = (unidades * kg_unidad) if unidades and kg_unidad else kg_unidad
        medidas = self._extract_medidas(base_norm)
        tipo_envase, material_envase, color_envase = self._detect_envase(full_norm)
        if material_norm:
            _, material_envase, color_envase = self._detect_envase(material_norm)
            tipo_envase = self._detect_envase(full_norm)[0]
        parsed = {
            "DescripcionForfait": full_desc,
            "GrupoForfait": grupo,
            "KgForfait": kg_total,
            "UnidadesForfait": unidades,
            "KgUnidad": kg_unidad if unidades else None,
            "Medidas": medidas,
            "TipoEnvase": tipo_envase,
            "MaterialEnvase": material_envase,
            "ColorEnvase": color_envase,
            "DescripcionNormalizada": full_norm,
        }
        parsed["ClaveForfait"] = self.regenerate_forfait_key(
            {"Cultivo": cultivo, "Campaña": campana, **parsed}
        )
        return parsed

    def parse_confeccion_features(self, row: dict[str, Any]) -> dict[str, Any]:
        text = " ".join(
            str(row.get(k) or "")
            for k in ("GrupoConfeccion", "NombreConfeccion", "DescripcionCorta", "Marca")
        )
        norm = self._norm_text(text)
        grupo = self._detect_group(norm)
        unidades = self._extract_units(norm)
        kg_unidad = self._extract_kg(norm)
        neto = self._to_float(row.get("Neto"))
        kg = neto or ((unidades * kg_unidad) if unidades and kg_unidad else kg_unidad)
        medidas = self._extract_medidas(norm)
        tipo_envase, material_envase, _color = self._detect_envase(norm)
        keywords = {t for t in re.split(r"[^A-Z0-9]+", norm) if len(t) >= 4}
        return {
            "GrupoForfait": grupo,
            "KgForfait": kg,
            "UnidadesForfait": unidades,
            "KgUnidad": kg_unidad if unidades else None,
            "Medidas": medidas,
            "TipoEnvase": tipo_envase,
            "MaterialEnvase": material_envase,
            "DescripcionNormalizada": norm,
            "Keywords": keywords,
        }

    def format_forfait_label(self, row: dict[str, Any]) -> str:
        return format_forfait_label(row)

    def format_forfait_option(self, row: dict[str, Any]) -> str:
        return format_forfait_label(row)

    def _fetch_pedido_confecciones(self, cultivo: str, campana: str) -> list[dict[str, Any]]:
        with self._connect_pedidos() as conn:
            rows = conn.execute(
                """
                SELECT
                    CAST(p."Confeccion" AS TEXT) AS ConfeccionPedido,
                    COALESCE(CAST(mc."GRUPO" AS TEXT), '') AS GrupoConfeccion,
                    COALESCE(CAST(mc."NOMBRE" AS TEXT), '') AS NombreConfeccion,
                    COALESCE(CAST(mc."DESCRIPCORTA" AS TEXT), '') AS DescripcionCorta,
                    COALESCE(CAST(mc."NETO" AS REAL), 0) AS Neto,
                    COALESCE(CAST(mc."MARCA" AS TEXT), '') AS Marca
                FROM "Pedidos" p
                LEFT JOIN "MConfecciones" mc
                  ON CAST(p."Confeccion" AS TEXT) = CAST(mc."CODIGO" AS TEXT)
                WHERE COALESCE(p."Cancelado", 0) = 0
                  AND CAST(p."Cultivo" AS TEXT) = ?
                  AND CAST(p."Campaña" AS TEXT) = ?
                  AND COALESCE(TRIM(CAST(p."Confeccion" AS TEXT)), '') <> ''
                GROUP BY p."Confeccion", mc."GRUPO", mc."NOMBRE", mc."DESCRIPCORTA", mc."NETO", mc."MARCA"
                ORDER BY CAST(p."Confeccion" AS TEXT)
                """,
                [cultivo, campana],
            ).fetchall()
        return [dict(r) for r in rows]

    def _score_forfait(self, features: dict[str, Any], forfait: dict[str, Any]) -> tuple[float, list[str]]:
        score = 0.0
        max_score = 100.0
        reasons: list[str] = []

        feature_group = str(features.get("GrupoForfait") or "")
        forfait_group = str(forfait.get("GrupoForfait") or "")
        if not feature_group or not forfait_group:
            return 0.0, ["grupo no detectado"]
        if feature_group != forfait_group:
            return 0.0, [f"grupo distinto ({feature_group} vs {forfait_group})"]
        score += 35
        reasons.append("grupo coincide")

        feature_kg = self._to_float(features.get("KgForfait"))
        forfait_kg = self._to_float(forfait.get("KgForfait"))
        if feature_kg is None or forfait_kg is None:
            max_score = min(max_score, 60.0)
            reasons.append("kg incompleto")
        elif abs(feature_kg - forfait_kg) < 0.05:
            score += 25
            reasons.append("kg coincide")
        else:
            return 0.0, [f"kg distinto ({feature_kg:g} vs {forfait_kg:g})"]

        feature_medidas = str(features.get("Medidas") or "")
        forfait_medidas = str(forfait.get("Medidas") or "")
        if feature_medidas and forfait_medidas and feature_medidas == forfait_medidas:
            score += 15
            reasons.append("medidas coinciden")
        elif feature_medidas:
            max_score = min(max_score, 75.0)
            reasons.append("medidas no coinciden")

        feature_material = str(features.get("MaterialEnvase") or "")
        forfait_material = str(forfait.get("MaterialEnvase") or "")
        feature_tipo = str(features.get("TipoEnvase") or "")
        forfait_tipo = str(forfait.get("TipoEnvase") or "")
        if feature_material:
            if feature_material == forfait_material:
                score += 20
                reasons.append("material/envase coincide")
            elif feature_tipo and feature_tipo == forfait_tipo:
                score += 10
                max_score = min(max_score, 65.0)
                reasons.append("solo coincide el tipo de envase")
            else:
                max_score = min(max_score, 65.0)
                reasons.append("material/envase no coincide")
        elif feature_tipo:
            if feature_tipo == forfait_tipo:
                score += 15
                max_score = min(max_score, 85.0)
                reasons.append("tipo de envase coincide")
            else:
                max_score = min(max_score, 65.0)
                reasons.append("tipo de envase no coincide")
        else:
            max_score = min(max_score, 75.0)
            reasons.append("envase no detectado")

        if (
            features.get("UnidadesForfait")
            and forfait.get("UnidadesForfait")
            and int(features["UnidadesForfait"]) == int(forfait["UnidadesForfait"])
        ):
            score += 5
            reasons.append("unidades coinciden")
        if (
            features.get("KgUnidad")
            and forfait.get("KgUnidad")
            and abs(float(features["KgUnidad"]) - float(forfait["KgUnidad"])) < 0.05
        ):
            score += 5
            reasons.append("kg por unidad coincide")
        elif features.get("UnidadesForfait"):
            max_score = min(max_score, 80.0)

        desc_norm = str(forfait.get("DescripcionNormalizada") or "")
        keyword_hits = sum(1 for token in features.get("Keywords", set()) if token in desc_norm)
        if keyword_hits:
            score += min(5, keyword_hits)
            reasons.append(f"{min(5, keyword_hits)} palabras clave coinciden")
        return min(score, max_score), reasons

    def _apply_forfait_to_row(self, row: dict[str, Any], forfait: dict[str, Any], estado: str) -> None:
        if not forfait:
            row.update(
                {
                    "DescripcionForfait": None,
                    "CosteConfeccionEurKg": None,
                    "CosteTotalEurKg": None,
                    "GrupoForfait": None,
                    "KgForfait": None,
                    "UnidadesForfait": None,
                    "KgUnidad": None,
                    "Medidas": None,
                    "TipoEnvase": None,
                    "MaterialEnvase": None,
                    "ClaveForfait": None,
                    "Estado": estado,
                }
            )
            return
        row.update(
            {
                "DescripcionForfait": forfait.get("DescripcionForfait"),
                "CosteConfeccionEurKg": forfait.get("CosteConfeccionEurKg"),
                "CosteTotalEurKg": forfait.get("CosteTotalEurKg"),
                "GrupoForfait": forfait.get("GrupoForfait"),
                "KgForfait": forfait.get("KgForfait"),
                "UnidadesForfait": forfait.get("UnidadesForfait"),
                "KgUnidad": forfait.get("KgUnidad"),
                "Medidas": forfait.get("Medidas"),
                "TipoEnvase": forfait.get("TipoEnvase"),
                "MaterialEnvase": forfait.get("MaterialEnvase"),
                "ClaveForfait": forfait.get("ClaveForfait"),
                "Estado": estado,
            }
        )

    def _update_equivalence_row(self, conn: sqlite3.Connection, row: dict[str, Any]) -> None:
        conn.execute(
            f"""
            UPDATE "{self.TABLE_EQUIV}"
            SET DescripcionForfait = ?, CosteConfeccionEurKg = ?, CosteTotalEurKg = ?,
                Estado = ?, Observaciones = COALESCE(?, Observaciones), FechaActualizacion = ?,
                GrupoForfait = ?, KgForfait = ?, UnidadesForfait = ?, KgUnidad = ?, Medidas = ?,
                TipoEnvase = ?, MaterialEnvase = ?, ClaveForfait = ?, ConfianzaSugerencia = ?
            WHERE Cultivo = ? AND "Campaña" = ? AND ConfeccionPedido = ?
            """,
            [
                row.get("DescripcionForfait"),
                row.get("CosteConfeccionEurKg"),
                row.get("CosteTotalEurKg"),
                row.get("Estado"),
                row.get("Observaciones"),
                row.get("FechaActualizacion") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                row.get("GrupoForfait"),
                row.get("KgForfait"),
                row.get("UnidadesForfait"),
                row.get("KgUnidad"),
                row.get("Medidas"),
                row.get("TipoEnvase"),
                row.get("MaterialEnvase"),
                row.get("ClaveForfait"),
                row.get("ConfianzaSugerencia"),
                row.get("Cultivo"),
                row.get("Campaña"),
                row.get("ConfeccionPedido"),
            ],
        )

    def _get_forfait(self, cultivo: str, campana: str, clave_forfait: str) -> dict[str, Any] | None:
        with self._connect_calc() as conn:
            row = conn.execute(
                f"""
                SELECT *
                FROM "{self.TABLE_FORFAIT}"
                WHERE Cultivo = ? AND "Campaña" = ? AND ClaveForfait = ?
                """,
                [cultivo, campana, str(clave_forfait or "").strip()],
            ).fetchone()
        return dict(row) if row else None

    def _get_forfait_by_id(self, id_forfait: int) -> dict[str, Any] | None:
        with self._connect_calc() as conn:
            row = conn.execute(
                f"""
                SELECT *
                FROM "{self.TABLE_FORFAIT}"
                WHERE Id = ?
                """,
                [int(id_forfait)],
            ).fetchone()
        return dict(row) if row else None

    def _prepare_forfait_value(self, field_name: str, value: Any) -> Any:
        if field_name in {"KgForfait", "KgUnidad"}:
            return self._to_float(value)
        if field_name == "UnidadesForfait":
            number = self._to_float(value)
            return int(number) if number is not None else None
        text = self._clean_text(value)
        if field_name in {"GrupoForfait", "TipoEnvase", "MaterialEnvase"}:
            return self._norm_key(text)
        if field_name == "DescripcionNormalizada":
            return self._norm_text(text)
        if field_name == "Medidas":
            return text.lower()
        return text

    def _connect_calc(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_calc))
        conn.row_factory = sqlite3.Row
        return conn

    def _connect_pedidos(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_pedidos))
        conn.row_factory = sqlite3.Row
        return conn

    @staticmethod
    def _ensure_columns(conn: sqlite3.Connection, table_name: str, columns: dict[str, str]) -> None:
        existing = {str(r["name"]) for r in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()}
        for column, col_type in columns.items():
            if column not in existing:
                conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{column}" {col_type}')

    @staticmethod
    def _find_row(ws: Any, needle: str, exact: bool = False) -> int | None:
        target = needle.upper()
        for row in ws.iter_rows(min_col=1, max_col=1):
            value = ForfaitRepository._clean_text(row[0].value).upper()
            if (exact and value == target) or (not exact and target in value):
                return int(row[0].row)
        return None

    @staticmethod
    def _cent_to_eur(value: float | None) -> float | None:
        return None if value is None else float(value) / 100.0

    @staticmethod
    def _to_float(value: Any) -> float | None:
        if value is None:
            return None
        try:
            return float(str(value).strip().replace(",", "."))
        except Exception:
            return None

    @staticmethod
    def _clean_text(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip())

    @staticmethod
    def _norm_key(value: Any) -> str:
        return ForfaitRepository._clean_text(value).upper()

    @staticmethod
    def _norm_text(value: Any) -> str:
        text = ForfaitRepository._clean_text(value).upper()
        replacements = str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN")
        return text.translate(replacements).replace(",", ".")

    @staticmethod
    def _detect_group(text: str) -> str:
        if "GIRS" in text or "GIRSA" in text:
            return "MALLAS/GIRSAC"
        if "MALL" in text or "MLL" in text:
            return "MALLAS"
        if "GRANEL" in text or "GRAN" in text:
            return "GRANEL"
        if "ENCAJADO" in text or "ENCA" in text:
            return "ENCAJADO"
        if "ALVEOLOS" in text or "ALVE" in text:
            return "ALVEOLOS"
        return ""

    @staticmethod
    def _detect_envase(text: str) -> tuple[str, str, str]:
        color = ""
        if "NEGRO" in text or "NEGRA" in text:
            color = "NEGRO"
        elif "VERDE" in text:
            color = "VERDE"
        elif "BLANCO" in text or "BLANCA" in text:
            color = "BLANCO"

        if "IFCO" in text:
            material = "IFCO" + (f" BLL {color}" if "BLL" in text and color else (" BLL" if "BLL" in text else ""))
            return "IFCO", material.strip(), color
        if "GREENBOX" in text:
            return "GREENBOX", "GREENBOX", color
        if "EPS" in text:
            return "EPS", "EPS", color
        if "MADERA" in text:
            material = "MADERA IMPRESA" if "IMPRESA" in text else "MADERA"
            return "MADERA", material, color
        if "PLASTICO" in text or "PLASTICA" in text:
            return "PLASTICO", "PLASTICO", color
        if "CARTON" in text or "CRT" in text or "CAJA" in text:
            return "CARTON", "CARTON", color
        if "MIXTO" in text:
            return "MIXTO", "MIXTO", color
        return "", "", color

    @staticmethod
    def _extract_kg(text: str) -> float | None:
        match = re.search(r"(\d+(?:[\.,]\d+)?)\s*K(?:G)?", str(text).upper())
        if not match:
            return None
        return ForfaitRepository._to_float(match.group(1))

    @staticmethod
    def _extract_units(text: str) -> int | None:
        match = re.search(
            r"\b(\d+)\s*[Xx]\s*(?:MALL|MLL|GIRS|GIRSA|ULPA|CLIP|PAP|PAP\.BAG)",
            str(text).upper(),
        )
        if not match:
            return None
        try:
            return int(match.group(1))
        except ValueError:
            return None

    @staticmethod
    def _extract_medidas(text: str) -> str:
        match = re.search(r"(\d{2,3}\s*[Xx]\s*\d{2,3})", str(text).upper())
        return match.group(1).replace(" ", "").lower() if match else ""

    @staticmethod
    def _format_number(value: Any) -> str:
        number = ForfaitRepository._to_float(value)
        if number is None:
            return ""
        return str(int(number)) if float(number).is_integer() else str(number).rstrip("0").rstrip(".")
