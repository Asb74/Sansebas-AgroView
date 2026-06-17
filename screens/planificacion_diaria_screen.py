import json
import logging
import os
import platform
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from services.planning_service import PlanningService
from services.runtime_database_service import RuntimeDatabaseService
from widgets.data_table import DataTable
from widgets.screen_header import ScreenHeader
from widgets.date_picker import DatePickerPopup
from widgets.multi_select_filter import MultiSelectFilter
from services.simulacion_asignacion import abrir_simulacion_asignacion, construir_panel_pedidos_previstos
from screens.operational_quality_settings_screen import OperationalQualitySettingsScreen
from services.production_capacity_service import ProductionCapacityService
from services.commercial_pdf_report_service import CommercialPdfReportService


class PlanificacionDiariaScreen(ttk.Frame):
    FILTERS_FILE = Path("config") / "planificacion_diaria_filters.json"
    BALANCE_SETTINGS_FILE = Path("config") / "planificacion_balance_settings.json"
    FILTER_KEYS = ["campana", "cultivo", "empresa", "semana", "var_coop", "grupo_varietal", "marca"]
    PEDIDOS_MODOS = [("Próximos 10 días", "10_dias"), ("Semana actual", "semana_actual"), ("Próximas semanas", "proximas_semanas"), ("Rango fechas", "rango"), ("Todos futuros", "todos_futuros"), ("Todos", "todos")]

    def __init__(self, master: tk.Misc, on_back) -> None:
        super().__init__(master, padding=16)
        self.on_back = on_back
        self.service = PlanningService()
        self.runtime_db_service = RuntimeDatabaseService()
        self.capacity_service = ProductionCapacityService()
        self.commercial_pdf_service = CommercialPdfReportService()
        self.fecha_desde_var = tk.StringVar()
        self.fecha_hasta_var = tk.StringVar()
        self.filter_widgets: dict[str, MultiSelectFilter] = {}
        self.pedidos_modo_var = tk.StringVar(value="10_dias")
        self.filters_status_var = tk.StringVar(value="Sin filtros activos")
        self.stock_campo_rows: list[dict] = []
        self.stock_almacen_rows: list[dict] = []
        self.stock_almacen_detalle_rows: list[dict] = []
        self.pedidos_pendientes_rows: list[dict] = []
        self.pedidos_pendientes_rows_raw: list[dict] = []
        self.pedidos_previstos_panel: dict | None = None
        self.balance_rows: list[dict] = []
        self.balance_rows_all: list[dict] = []
        self.last_capacity_simulation: dict | None = None
        self.last_capacity_payload: dict | None = None
        self._planning_cache = self._new_planning_cache()
        self.sim_policy_vars: dict[str, tk.BooleanVar] = {}
        self._build_ui()
        self._load_filters()
        self._load_filter_options(contextual=True)
        self._load_balance_settings()
        self.load_data(save_filters=False)

    def _load_filter_options(self, contextual: bool = True) -> None:
        payload = self._filters_payload() if contextual else {}
        for key in self.FILTER_KEYS:
            selected = self.filter_widgets[key].get_selected() if key in self.filter_widgets else []
            try:
                options = self.service.get_filter_options_contextual(key, payload) if contextual else self.service.get_filter_options(key)
                valid = [v for v in selected if v in options]
                self.filter_widgets[key].set_options(options)
                self.filter_widgets[key].set_selected(valid)
            except Exception:
                logging.getLogger(__name__).exception("No se pudo cargar opciones de filtro %s", key)

    def _build_ui(self) -> None:
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        ScreenHeader(self, title="Planificación diaria", on_back=self.on_back).grid(row=0, column=0, sticky="ew")

        filters_frame = ttk.LabelFrame(self, text="Filtros globales", padding=10)
        filters_frame.grid(row=1, column=0, sticky="ew", pady=(8, 8))
        fields = [("Campaña", "campana"), ("Cultivo", "cultivo"), ("Semana", "semana"), ("Fecha desde", "fecha_desde"), ("Fecha hasta", "fecha_hasta"), ("Empresa", "empresa"), ("Variedad Coop", "var_coop"), ("Grupo varietal", "grupo_varietal"), ("Marca", "marca")]
        for i, (label, key) in enumerate(fields):
            ttk.Label(filters_frame, text=label).grid(row=0, column=i, sticky="w", padx=4)
            if key == "fecha_desde":
                self._build_date_field(filters_frame, 1, i, self.fecha_desde_var)
            elif key == "fecha_hasta":
                self._build_date_field(filters_frame, 1, i, self.fecha_hasta_var)
            else:
                widget = MultiSelectFilter(filters_frame, title=label, width=16)
                widget.grid(row=1, column=i, padx=4, sticky="ew")
                self.filter_widgets[key] = widget
            filters_frame.grid_columnconfigure(i, weight=1)
        btns = ttk.Frame(filters_frame)
        btns.grid(row=2, column=0, columnspan=10, sticky="w", pady=(8, 0))
        ttk.Button(btns, text="Aplicar filtros", command=lambda: self._reload_with_invalidated_cache("aplicar_filtros", save_filters=True)).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Limpiar filtros", command=self.reset_filters).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Reaplicar filtros", command=lambda: self._reload_with_invalidated_cache("reaplicar_filtros", save_filters=True)).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Exportar Excel", command=self.export_excel).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Exportar diagnóstico", command=self.export_diagnostico).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Informe comercial PDF", command=self.export_informe_comercial_pdf).pack(side="left", padx=(0, 8))
        self._btn_actualizar_planificacion = ttk.Button(btns, text="Actualizar planificación", command=self._actualizar_planificacion)
        self._btn_actualizar_planificacion.pack(side="left", padx=(0, 8))
        self._btn_actualizar_foto = ttk.Button(btns, text="Actualizar foto local", command=self._actualizar_foto_local)
        self._btn_actualizar_foto.pack(side="left")
        ttk.Label(filters_frame, textvariable=self.filters_status_var).grid(row=3, column=0, columnspan=10, sticky="w", padx=4, pady=(6, 0))

        self.tabs = ttk.Notebook(self)
        self.tabs.grid(row=2, column=0, sticky="nsew")

        self.campo_tab = ttk.Frame(self.tabs, padding=8)
        self.almacen_tab = ttk.Frame(self.tabs, padding=8)
        self.tabs.add(self.campo_tab, text="Stock campo")
        self.tabs.add(self.almacen_tab, text="Stock almacén")
        self.pedidos_tab = ttk.Frame(self.tabs, padding=8)
        self.tabs.add(self.pedidos_tab, text="Pedidos pendientes")
        self.previstos_tab = ttk.Frame(self.tabs, padding=8)
        self.tabs.add(self.previstos_tab, text="Pedidos previstos")
        self.balance_tab = ttk.Frame(self.tabs, padding=8)
        self.capacidad_tab = ttk.Frame(self.tabs, padding=8)
        self.tabs.add(self.balance_tab, text="Balance")
        self.tabs.add(self.capacidad_tab, text="Capacidad productiva")
        self.tabs.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.kpi_campo = tk.StringVar(value="Kg campo total: 0 | Nº partidas: 0 | Nº variedades: 0")
        self.kpi_almacen = tk.StringVar(value="Kg stock almacén: 0 | Nº grupos: 0 | Nº variedades: 0 | Nº calibres: 0")
        self.last_update = tk.StringVar(value="")
        self.snapshot_info_var = tk.StringVar(value="Foto de datos: No disponible")
        self.kpi_pedidos = tk.StringVar(value="Kg pedido teórico total: 0 | Kg hecho real total: 0 | Kg pendiente total: 0 | Merma kg total: 0 | % merma total: 0 | Nº pedidos: 0 | Nº líneas: 0 | Nº líneas sin datos: 0 | Nº líneas parciales: 0")
        self.kpi_balance = tk.StringVar(value="Kg stock comercial: 0 | Kg pedidos pendientes: 0 | Diferencia comercial: 0 | Kg stock industrial almacén: 0 | Kg entrada estimada: 0 | Kg base total estimada: 0 | Kg cobertura exacta: 0 | Kg cobertura agrupada: 0 | Kg cobertura potencial total: 0 | Nº faltantes comerciales: 0 | Nº faltantes con cobertura agrupada: 0 | Nº faltantes con cobertura: 0 | Nº faltantes sin cobertura: 0 | Nº sobrantes comerciales: 0")

        ttk.Label(self.campo_tab, textvariable=self.kpi_campo, style="KPI.TLabel").pack(anchor="w", pady=(0, 2))
        ttk.Label(self.campo_tab, textvariable=self.last_update).pack(anchor="w", pady=(0, 2))
        ttk.Label(self.campo_tab, textvariable=self.snapshot_info_var).pack(anchor="w", pady=(0, 6))
        self.campo_table = DataTable(self.campo_tab, ["Cultivo", "Campaña", "Fecha carga", "Semana", "Socio", "Variedad", "Grupo varietal", "Boleta", "Plataforma", "Empresa", "Restricciones", "Color", "Kg campo", "Estado aprovechamiento", "Nº calibres aprovechamiento", "Kg estimados calculados"])
        self.campo_table.pack(fill="both", expand=True)
        ttk.Button(self.campo_tab, text="Ver aprovechamiento", command=self._ver_aprovechamiento_campo).pack(anchor="e", pady=(6, 0))

        header_almacen = ttk.Frame(self.almacen_tab)
        header_almacen.pack(fill="x", pady=(0, 6))
        ttk.Label(header_almacen, textvariable=self.kpi_almacen, style="KPI.TLabel").pack(side="left")
        ttk.Button(header_almacen, text="Ver detalle palets", command=self.show_detalle_palets).pack(side="right")
        self.almacen_table = DataTable(self.almacen_tab, ["Cultivo", "Campaña", "Variedad", "Grupo varietal", "Calibre", "Categoría", "Marca", "IdConfeccion", "Confección", "Palets", "Cajas", "Kg stock", "Agrupado"])
        self.almacen_table.pack(fill="both", expand=True)

        pedidos_filters = ttk.Frame(self.pedidos_tab)
        pedidos_filters.pack(fill="x", pady=(0, 6))
        ttk.Label(pedidos_filters, text="Modo pedidos").pack(side="left", padx=(0, 6))
        self._pedidos_modo_combo = ttk.Combobox(pedidos_filters, state="readonly", width=24, values=[m[0] for m in self.PEDIDOS_MODOS])
        self._pedidos_modo_combo.pack(side="left")
        self._sync_pedidos_mode_combo()
        self._pedidos_modo_combo.bind("<<ComboboxSelected>>", self._on_pedidos_modo_changed)
        ttk.Label(pedidos_filters, text="Grupo confección").pack(side="left", padx=(10, 6))
        self.pedidos_grupo_confeccion_filter = MultiSelectFilter(pedidos_filters, title="Grupo confección", width=18)
        self.pedidos_grupo_confeccion_filter.pack(side="left")
        ttk.Label(pedidos_filters, text="Perfil confección").pack(side="left", padx=(10, 6))
        self.pedidos_perfil_confeccion_filter = MultiSelectFilter(pedidos_filters, title="Perfil confección", width=18)
        self.pedidos_perfil_confeccion_filter.pack(side="left")

        ttk.Label(self.pedidos_tab, textvariable=self.kpi_pedidos, style="KPI.TLabel").pack(anchor="w", pady=(0, 6))
        self.pedidos_table = DataTable(
            self.pedidos_tab,
            ["Semana", "Fecha salida", "Cliente", "IdPedidoLora", "Línea", "Cultivo", "Campaña", "Variedad Coop", "Grupo varietal", "Calibre", "Categoría", "Marca", "IdConfeccion", "Confección", "Grupo confección", "Perfil confección", "Palets pedido", "Palets hechos", "Palets pendientes", "Cajas/palet", "Cajas pedido", "Cajas hechas", "Cajas pendientes", "Kg pedido teórico", "Kg hecho real", "Kg pendiente", "Merma kg", "% hecho", "% merma", "Estado", "Aviso"],
        )
        self.pedidos_table.pack(fill="both", expand=True)

        self.pedidos_previstos_panel = construir_panel_pedidos_previstos(
            self.previstos_tab,
            owner=self,
            filters_payload=self._filters_payload,
            refresh_command=lambda: self._reload_with_invalidated_cache("pedidos_previstos", save_filters=True),
            refresh_button_text="Refrescar planificación",
        )

        sim_frame = ttk.LabelFrame(self.balance_tab, text="Simulación / flexibilidad", padding=8)
        sim_frame.pack(fill="x", pady=(0, 6))
        sim_defaults = [("mismo_grupo_varietal","Mismo grupo varietal", True),("permitir_variedad_alternativa","Permitir variedad alternativa dentro del grupo", True),("permitir_grupo_varietal_alternativo","Permitir grupo varietal alternativo", False),("permitir_calibre_admitido","Permitir calibre admitido", True),("permitir_calibre_agrupado","Permitir calibre agrupado", True),("permitir_solape_parcial","Permitir solape parcial de calibre", False),("permitir_categoria_inferior","Permitir categoría inferior", False),("permitir_categoria_superior","Permitir categoría superior", False),("usar_stock_industrial","Usar stock industrial", True),("usar_stock_comercial","Usar stock comercial S/P", False),("usar_entrada_estimada","Usar entrada estimada / campo", False),("usar_reservas_amplias","Usar reservas amplias", False)]
        for i,(k,lbl,dv) in enumerate(sim_defaults):
            self.sim_policy_vars[k] = tk.BooleanVar(value=dv)
            ttk.Checkbutton(sim_frame, text=lbl, variable=self.sim_policy_vars[k]).grid(row=i//3, column=i%3, sticky="w", padx=4, pady=2)
        ttk.Label(sim_frame, text="Modo simulación: no descuenta stock ni crea reservas").grid(row=4, column=0, columnspan=2, sticky="w", pady=(6,0))
        ttk.Button(sim_frame, text="Recalcular simulación", command=self._recalcular_simulacion).grid(row=4, column=2, sticky="e")
        ttk.Button(sim_frame, text="Configurar calidad", command=self._open_config_calidad).grid(row=4, column=1, sticky="e", padx=(0, 8))

        balance_header = ttk.Frame(self.balance_tab)
        balance_header.pack(fill="x", pady=(0, 6))
        ttk.Label(balance_header, textvariable=self.kpi_balance, style="KPI.TLabel").pack(side="left", anchor="w")
        acciones_frame = ttk.Frame(balance_header)
        acciones_frame.pack(side="right")
        ttk.Button(acciones_frame, text="Simular asignación", command=self._open_simulacion_asignacion).pack(side="left", padx=4)
        ttk.Button(acciones_frame, text="Ver cobertura (pedidos)", command=self._open_selected_balance_coverage).pack(side="left", padx=4)
        self.balance_table = DataTable(
            self.balance_tab,
            ["Cultivo", "Campaña", "Grupo varietal", "Variedad", "Calibre", "Categoría", "Marca", "IdConfeccion", "Confección", "Kg stock comercial", "Kg pedidos pendientes", "Diferencia comercial", "Tipo línea", "Estado comercial", "Cobertura posible", "Kg disponibilidad compatible", "Mejor cobertura", "Calibres coincidentes", "Flexibilidad aplicada", "Score cobertura", "Explicación"],
        )
        self.balance_table.pack(fill="both", expand=True)
        self.balance_table.tree.bind("<Double-1>", self._on_balance_double_click)
        self.balance_table.tree.tag_configure("tipo_venta", foreground="#0d47a1")
        self.balance_table.tree.tag_configure("tipo_industria", foreground="#6a1b9a")
        self.balance_table.tree.tag_configure("estado_faltante", foreground="#b71c1c")
        self.balance_table.tree.tag_configure("estado_sobrante", foreground="#1b5e20")

        self.kpi_capacidad = tk.StringVar(value="Capacidad productiva: sin calcular")
        ttk.Label(self.capacidad_tab, textvariable=self.kpi_capacidad, style="KPI.TLabel").pack(anchor="w", pady=(0, 6))
        self.capacidad_subtabs = ttk.Notebook(self.capacidad_tab)
        self.capacidad_subtabs.pack(fill="both", expand=True)
        self.capacidad_familias_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_lineas_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_recursos_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_personal_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_bottleneck_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_incidencias_tab = ttk.Frame(self.capacidad_subtabs, padding=4)
        self.capacidad_subtabs.add(self.capacidad_familias_tab, text="Familias")
        self.capacidad_subtabs.add(self.capacidad_lineas_tab, text="Líneas")
        self.capacidad_subtabs.add(self.capacidad_recursos_tab, text="Recursos")
        self.capacidad_subtabs.add(self.capacidad_personal_tab, text="Personal requerido")
        self.capacidad_subtabs.add(self.capacidad_bottleneck_tab, text="Cuellos de botella")
        self.capacidad_subtabs.add(self.capacidad_incidencias_tab, text="Incidencias")
        self.capacidad_family_table = DataTable(self.capacidad_familias_tab, ["Familia", "Kg reales", "Kg previstos", "Kg total", "Horas necesarias", "Horas disponibles", "Ocupación %", "Rendimiento medio", "Personal estimado", "Estado"])
        self.capacidad_family_table.pack(fill="both", expand=True)
        self.capacidad_line_table = DataTable(self.capacidad_lineas_tab, ["Línea productiva", "Kg", "Horas necesarias", "Horas disponibles línea", "Ocupación %", "Pedidos", "Cambios formato estimados", "Estado"])
        self.capacidad_line_table.pack(fill="both", expand=True)
        ttk.Label(self.capacidad_recursos_tab, text="Recursos utilizados / cuellos de botella", style="KPI.TLabel").pack(anchor="w", pady=(0, 6))
        self.capacidad_resource_table = DataTable(self.capacidad_recursos_tab, ["Recurso", "Tipo recurso", "Línea productiva", "Modo uso", "Kg asignados", "Capacidad kg/h", "Horas necesarias", "Horas disponibles", "Ocupación %", "Personal mínimo", "Personal óptimo", "Estado"])
        self.capacidad_resource_table.pack(fill="both", expand=True)
        ttk.Label(self.capacidad_personal_tab, text="Dotación completa requerida por flujo productivo", style="KPI.TLabel").pack(anchor="w", pady=(0, 6))
        self.capacidad_staffing_table = DataTable(self.capacidad_personal_tab, ["Línea productiva", "Área / puesto", "Tipo personal", "Mínimo", "Óptimo", "Ocupación %", "Necesario estimado", "Disponible base", "Polivalente", "Disponible efectivo", "Origen polivalencia", "Disponible", "Diferencia", "Estado"])
        self.capacidad_staffing_table.pack(fill="both", expand=True)
        ttk.Label(self.capacidad_bottleneck_tab, text="Capacidad humana estimada frente a dotación óptima", style="KPI.TLabel").pack(anchor="w", pady=(0, 6))
        self.capacidad_bottleneck_table = DataTable(self.capacidad_bottleneck_tab, ["Línea productiva", "Puesto limitante", "Tipo personal", "Mínimo", "Óptimo", "Disponible base", "Polivalente", "Disponible efectivo", "Déficit mínimo", "Déficit óptimo", "Factor capacidad %", "Kg línea", "Kg alcanzables estimados", "Kg no cubiertos estimados", "Horas reales actuales", "Estado", "Acción sugerida"])
        self.capacidad_bottleneck_table.pack(fill="both", expand=True)
        self.capacidad_inc_table = DataTable(self.capacidad_incidencias_tab, ["Tipo incidencia", "Pedido", "Cliente", "Confección", "Línea productiva", "Motivo", "Acción sugerida"])
        self.capacidad_inc_table.pack(fill="both", expand=True)

    def _build_date_field(self, parent: ttk.Frame, row: int, col: int, var: tk.StringVar) -> None:
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=col, sticky="ew", padx=4)
        frame.grid_columnconfigure(0, weight=1)
        ttk.Entry(frame, textvariable=var).grid(row=0, column=0, sticky="ew")
        btn = ttk.Button(frame, text="...", width=3)
        btn.configure(command=lambda v=var, b=btn: DatePickerPopup(self, target_var=v, anchor_widget=b))
        btn.grid(row=0, column=1, padx=(4, 0))

    def _open_config_calidad(self) -> None:
        win = tk.Toplevel(self)
        win.title("Configuración calidad operativa")
        win.geometry("1200x620")
        frame = OperationalQualitySettingsScreen(win, on_back=win.destroy)
        frame.pack(fill="both", expand=True)

    def _filters_payload(self) -> dict:
        return {
            "campana": self.filter_widgets["campana"].get_selected(),
            "cultivo": self.filter_widgets["cultivo"].get_selected(),
            "empresa": self.filter_widgets["empresa"].get_selected(),
            "semana": self.filter_widgets["semana"].get_selected(),
            "var_coop": self.filter_widgets["var_coop"].get_selected(),
            "grupo_varietal": self.filter_widgets["grupo_varietal"].get_selected(),
            "marca": self.filter_widgets["marca"].get_selected(),
            "fecha_desde": self.fecha_desde_var.get().strip(),
            "fecha_hasta": self.fecha_hasta_var.get().strip(),
            "pedidos_modo": self.pedidos_modo_var.get(),
        }

    def _new_planning_cache(self) -> dict:
        return {
            "filters_key": None,
            "stock_campo": None,
            "stock_almacen": None,
            "stock_almacen_detalle": None,
            "pedidos_pendientes": None,
            "pedidos_previstos": None,
            "balance": None,
            "capacidad_productiva": None,
            "aprovechamientos_resumen": None,
            "aprovechamientos_detalle": None,
        }

    def _get_filters_key(self, payload: dict | None = None) -> tuple:
        payload = payload or self._filters_payload()

        def stable(value):
            if isinstance(value, (list, tuple, set)):
                return tuple(sorted(str(v) for v in value))
            if isinstance(value, dict):
                return tuple((k, stable(v)) for k, v in sorted(value.items()))
            return str(value or "")

        policy = self._build_sim_policy() if self.sim_policy_vars else {}
        return tuple((key, stable(payload.get(key))) for key in (
            "campana", "cultivo", "semana", "fecha_desde", "fecha_hasta",
            "empresa", "var_coop", "grupo_varietal", "marca", "pedidos_modo",
        )) + (("sim_policy", stable(policy)),)

    def _invalidate_planning_cache(self, motivo: str, keys: list[str] | None = None) -> None:
        logger = logging.getLogger(__name__)
        logger.info("CACHE INVALIDADA motivo=%s", motivo)
        if keys is None:
            self._planning_cache = self._new_planning_cache()
            self.last_capacity_simulation = None
            self.last_capacity_payload = None
            return
        for key in keys:
            if key in self._planning_cache:
                self._planning_cache[key] = None
        self._planning_cache["aprovechamientos_resumen"] = None
        self._planning_cache["aprovechamientos_detalle"] = None
        self.last_capacity_simulation = None
        self.last_capacity_payload = None

    def _ensure_cache_key(self, payload: dict) -> None:
        filters_key = self._get_filters_key(payload)
        if self._planning_cache.get("filters_key") != filters_key:
            self._invalidate_planning_cache("cambio_filtros")
            self._planning_cache["filters_key"] = filters_key

    def _reload_with_invalidated_cache(self, motivo: str, save_filters: bool = True) -> None:
        self._invalidate_planning_cache(motivo)
        self.load_data(save_filters=save_filters)

    def _cache_get(self, key: str):
        logger = logging.getLogger(__name__)
        if self._planning_cache.get(key) is not None:
            logger.info("CACHE HIT %s", key)
            return self._planning_cache[key]
        logger.info("CACHE MISS %s", key)
        return None

    def _cache_set(self, key: str, value) -> None:
        self._planning_cache[key] = value

    def load_data(self, save_filters: bool = True) -> None:
        tab_activa = self.tabs.tab(self.tabs.select(), "text")
        payload = self._filters_payload()
        self._ensure_cache_key(payload)
        updated = None
        update_warning = False
        pedidos_kpi = {}
        if tab_activa == "Stock campo":
            cached = self._cache_get("stock_campo")
            if cached is not None:
                self.stock_campo_rows, updated, update_warning = cached
            else:
                try:
                    self.stock_campo_rows, updated, update_warning = self.service.load_stock_campo(payload)
                    self._cache_set("stock_campo", (self.stock_campo_rows, updated, update_warning))
                except Exception as exc:
                    self.stock_campo_rows = []
                    messagebox.showwarning("Planificación diaria", f"No se pudo cargar stock campo: {exc}")
        elif tab_activa == "Stock almacén":
            cached = self._cache_get("stock_almacen")
            if cached is not None:
                self.stock_almacen_rows, self.stock_almacen_detalle_rows, almacen_warning = cached
                if almacen_warning:
                    messagebox.showwarning("Planificación diaria", almacen_warning)
            else:
                try:
                    self.stock_almacen_rows, almacen_warning = self.service.load_stock_almacen(payload)
                    self.stock_almacen_detalle_rows = self.service.load_stock_almacen_detalle_palets(payload)
                    self._cache_set("stock_almacen", (self.stock_almacen_rows, self.stock_almacen_detalle_rows, almacen_warning))
                    if almacen_warning:
                        messagebox.showwarning("Planificación diaria", almacen_warning)
                except Exception as exc:
                    self.stock_almacen_rows = []
                    self.stock_almacen_detalle_rows = []
                    logging.getLogger(__name__).warning("No se pudo cargar stock almacén: %s", exc)
                    messagebox.showwarning("Planificación diaria", f"No se pudo cargar stock almacén: {exc}")
        elif tab_activa == "Pedidos pendientes":
            cached = self._cache_get("pedidos_pendientes")
            if cached is not None:
                self.pedidos_pendientes_rows_raw, self.pedidos_pendientes_rows, pedidos_kpi = cached
                self._refresh_pedidos_local_filter_options(self.pedidos_pendientes_rows_raw)
            else:
                modo_pedidos = self.pedidos_modo_var.get()
                try:
                    pedidos_rows, pedidos_kpi = self.service.load_pedidos_pendientes(payload, modo_pedidos)
                    self.pedidos_pendientes_rows_raw = [dict(r) for r in pedidos_rows]
                    self._refresh_pedidos_local_filter_options(pedidos_rows)
                    self.pedidos_pendientes_rows, pedidos_kpi = self._apply_pedidos_local_filters(pedidos_rows, pedidos_kpi)
                    self._cache_set("pedidos_pendientes", (self.pedidos_pendientes_rows_raw, self.pedidos_pendientes_rows, pedidos_kpi))
                except Exception as exc:
                    self.pedidos_pendientes_rows = []
                    self.pedidos_pendientes_rows_raw = []
                    logging.getLogger(__name__).warning("No se pudo cargar pedidos pendientes: %s", exc)
                    messagebox.showwarning("Pedidos pendientes", f"No se pudo cargar pedidos pendientes: {exc}")
        elif tab_activa == "Pedidos previstos":
            cached = self._cache_get("pedidos_previstos")
            if self.pedidos_previstos_panel:
                if cached is None:
                    self.pedidos_previstos_panel["refresh_rows"]()
                    self._cache_set("pedidos_previstos", True)
        elif tab_activa == "Balance":
            cached = self._cache_get("balance")
            if cached is not None:
                self.balance_rows_all, self.balance_rows = cached
            else:
                try:
                    self.balance_rows_all = self.service.load_balance_planificacion(payload, policy=self._build_sim_policy())
                    self.balance_rows = self._build_balance_view_rows(self.balance_rows_all)
                    self._cache_set("balance", (self.balance_rows_all, self.balance_rows))
                except Exception as exc:
                    self.balance_rows_all = []
                    self.balance_rows = []
                    logging.getLogger(__name__).warning("No se pudo cargar balance: %s", exc)
                    messagebox.showwarning("Balance", f"No se pudo cargar balance: {exc}")
        elif tab_activa == "Capacidad productiva":
            cached = self._cache_get("capacidad_productiva")
            if cached is not None:
                cap = cached
                self.last_capacity_simulation = cap
                self.last_capacity_payload = dict(payload)
                self._render_capacidad(cap)
            else:
                try:
                    cap = self.capacity_service.build_capacity_simulation(payload, self.pedidos_modo_var.get())
                    self.last_capacity_simulation = cap
                    self.last_capacity_payload = dict(payload)
                    self._cache_set("capacidad_productiva", cap)
                    self._render_capacidad(cap)
                except Exception as exc:
                    self.capacidad_family_table.set_rows([])
                    self.capacidad_line_table.set_rows([])
                    self.capacidad_resource_table.set_rows([])
                    self.capacidad_staffing_table.set_rows([])
                    self.capacidad_bottleneck_table.set_rows([])
                    self.capacidad_inc_table.set_rows([])
                    self.last_capacity_simulation = None
                    self.last_capacity_payload = None
                    messagebox.showwarning("Capacidad productiva", f"No se pudo calcular capacidad productiva: {exc}")
        self.campo_table.set_rows(self.stock_campo_rows)
        self.almacen_table.set_rows(self.stock_almacen_rows)
        self.pedidos_table.set_rows(self.pedidos_pendientes_rows)
        self.balance_table.set_rows(self.balance_rows)
        self._update_kpis(updated, pedidos_kpi)
        if update_warning:
            messagebox.showwarning("Planificación diaria", "No se pudo leer un archivo auxiliar de actualización. Se continuará sin ese dato.")
        if save_filters:
            self._save_filters(payload)
        self._load_filter_options(contextual=True)
        payload = self._filters_payload()
        self.filters_status_var.set(self._format_filters_status(payload))
        self._refresh_snapshot_info_label()

    def _render_capacidad(self, cap: dict) -> None:
        s = cap["summary"]
        self.kpi_capacidad.set(
            f"Kg reales pendientes: {s['Kg reales pendientes']:,.2f} | Kg previstos: {s['Kg previstos']:,.2f} | "
            f"Kg total simulación: {s['Kg total simulación']:,.2f} | Horas necesarias estimadas: {s['Horas necesarias estimadas']:,.2f} | "
            f"Horas disponibles: {s['Horas disponibles']:,.2f} | Ocupación %: {s['Ocupación %']:,.2f}% | "
            f"Turnos equivalentes: {s.get('turnos_equivalentes', 0):,.2f} | Personal total/directo/soporte/indirecto: {s['Personal disponible total']}/{s['Personal directo disponible']}/{s['Personal soporte disponible']}/{s['Personal indirecto disponible']} | "
            f"Personal requerido mín/ópt/estimado: {s.get('personal_minimo_flujo', 0)}/{s.get('personal_optimo_flujo', 0)}/{s.get('personal_estimado_flujo', 0)} | Déficit personal: {s.get('deficit_personal_flujo', 0)} | "
            f"Personal min/ópt recursos: {s.get('personal_minimo_recursos', 0)}/{s.get('personal_optimo_recursos', 0)} | Cuello botella: {s.get('linea_cuello_botella_principal', '') or '-'} / {s.get('puesto_cuello_botella_principal', '') or s.get('motivo_cuello_botella', '') or 'Sin datos'} | "
            f"Capacidad alcanzable: {s.get('capacidad_alcanzable_pct', 0):,.2f}% | Kg alcanzables: {s.get('kg_alcanzables_estimados', 0):,.0f} | Estado capacidad: {s['Estado capacidad']}"
        )
        self.capacidad_family_table.set_rows(cap["family_rows"])
        self.capacidad_line_table.set_rows(cap["line_rows"])
        self.capacidad_resource_table.set_rows(cap.get("resource_rows", []))
        self.capacidad_staffing_table.set_rows(cap.get("staffing_rows", []))
        self.capacidad_bottleneck_table.set_rows(cap.get("bottleneck_rows", []))
        self.capacidad_inc_table.set_rows(cap["incidencias"])

    def _update_kpis(self, updated, pedidos_kpi: dict) -> None:
        self.kpi_campo.set(
            f"Kg campo total: {sum(float(r.get('Kg campo', 0) or 0) for r in self.stock_campo_rows):,.2f} | "
            f"Nº partidas: {len(self.stock_campo_rows)} | Nº variedades: {len({r.get('Variedad') for r in self.stock_campo_rows})}"
        )
        self.kpi_almacen.set(
            f"Kg stock almacén: {sum(float(r.get('Kg stock', 0) or 0) for r in self.stock_almacen_rows):,.2f} | "
            f"Nº grupos: {len(self.stock_almacen_rows)} | Nº variedades: {len({r.get('Variedad') for r in self.stock_almacen_rows})} | Nº calibres: {len({r.get('Calibre') for r in self.stock_almacen_rows})}"
        )
        self.last_update.set(f"Última actualización: {updated}" if updated else self.last_update.get() or "Última actualización: No disponible")
        self.kpi_pedidos.set(
            f"Kg pedido teórico total: {float(pedidos_kpi.get('Kg pedido teórico total', 0) or 0):,.2f} | "
            f"Kg hecho real total: {float(pedidos_kpi.get('Kg hecho real total', 0) or 0):,.2f} | Kg pendiente total: {float(pedidos_kpi.get('Kg pendiente total', 0) or 0):,.2f} | "
            f"Merma kg total: {float(pedidos_kpi.get('Merma kg total', 0) or 0):,.2f} | % merma total: {float(pedidos_kpi.get('% merma total', 0) or 0):,.2f}% | "
            f"Nº pedidos: {int(pedidos_kpi.get('Nº pedidos', 0) or 0)} | Nº líneas: {int(pedidos_kpi.get('Nº líneas', 0) or 0)} | Nº líneas sin datos: {int(pedidos_kpi.get('Nº líneas sin datos', 0) or 0)} | Nº líneas parciales: {int(pedidos_kpi.get('Nº líneas parciales', 0) or 0)}"
        )
        self.kpi_balance.set(self._format_balance_summary(self.balance_rows_all))

    def _build_balance_view_rows(self, rows: list[dict]) -> list[dict]:
        out: list[dict] = []
        for row in rows:
            estado = str(row.get("Estado comercial", "")).strip()
            if estado not in ("Faltante comercial", "Sobrante comercial"):
                continue
            item = dict(row)
            tags = []
            tipo = str(row.get("Tipo línea", "")).strip().lower()
            if "venta" in tipo:
                tags.append("tipo_venta")
            else:
                tags.append("tipo_industria")
            if estado == "Faltante comercial":
                tags.append("estado_faltante")
            elif estado == "Sobrante comercial":
                tags.append("estado_sobrante")
            item["__tags__"] = tuple(tags)
            out.append(item)
        return out

    def _format_balance_summary(self, rows: list[dict]) -> str:
        faltantes = [r for r in rows if str(r.get("Estado comercial", "")).strip() == "Faltante comercial"]
        sobrantes = [r for r in rows if str(r.get("Estado comercial", "")).strip() == "Sobrante comercial"]
        return (
            f"Nº faltantes: {len(faltantes)} | "
            f"Kg faltantes: {sum(max(0.0, -float(r.get('Diferencia comercial', 0) or 0)) for r in faltantes):,.2f} | "
            f"Nº sobrantes: {len(sobrantes)} | "
            f"Kg disponibles para venta: {sum(max(0.0, float(r.get('Diferencia comercial', 0) or 0)) for r in sobrantes):,.2f} | "
            f"Kg stock industrial almacén: {sum(float(r.get('Kg stock industrial almacén', 0) or 0) for r in rows if str(r.get('Tipo línea', '')).startswith('Pedido')):,.2f} | "
            f"Kg entrada estimada: {sum(float(r.get('Kg entrada estimada', 0) or 0) for r in rows if str(r.get('Tipo línea', '')).startswith('Pedido')):,.2f} | "
            f"Kg base total estimada: {sum(float(r.get('Kg base total estimada', 0) or 0) for r in rows if str(r.get('Tipo línea', '')).startswith('Pedido')):,.2f} | "
            f"Kg cobertura potencial total: {sum(float(r.get('Kg cobertura potencial total', 0) or 0) for r in rows if str(r.get('Tipo línea', '')).startswith('Pedido')):,.2f}"
        )

    def _on_balance_double_click(self, _event=None) -> None:
        self._open_selected_balance_coverage()


    def _open_simulacion_asignacion(self) -> None:
        pedidos = [
            r
            for r in self.balance_rows_all
            if str(r.get("Tipo línea", "")).strip() == "Pedido"
        ]

        def _candidatos_de_pedido(pedido: dict) -> list[dict]:
            return self.service.get_candidatos_compatibles_para_pedido(self._filters_payload(), pedido, policy_cfg=self._build_sim_policy())

        def _inventario_global() -> list[dict]:
            pools = self.service.get_inventario_operativo_global(
                self._filters_payload(),
                policy=self._build_sim_policy(),
            )
            logger = logging.getLogger(__name__)
            logger.info(
                "Inventario operativo global: pools=%s calibres=%s",
                len(pools),
                sorted(set(str(p.get("calibre", "")) for p in pools)),
            )
            logger.info(
                "Inventario operativo global origenes=%s",
                sorted(set(str(p.get("origen", "")) for p in pools)),
            )
            return pools

        inventario_global = _inventario_global()
        if not pedidos and not inventario_global:
            messagebox.showinfo(
                "Simulación de asignación",
                "No hay pedidos pendientes ni stock analizable para simular.",
                parent=self,
            )
            return

        if not self.pedidos_pendientes_rows_raw:
            try:
                modo_pedidos = self.pedidos_modo_var.get()
                pedidos_rows, _kpi = self.service.load_pedidos_pendientes(self._filters_payload(), modo_pedidos)
                self.pedidos_pendientes_rows_raw = [dict(r) for r in pedidos_rows]
            except Exception:
                logging.getLogger(__name__).exception("No se pudo cargar pedidos raw para horizonte")

        logger = logging.getLogger(__name__)
        pedidos_detalle_horizonte = [dict(r) for r in self.pedidos_pendientes_rows_raw]
        if not pedidos_detalle_horizonte:
            pedidos_detalle_horizonte = [dict(r) for r in self.pedidos_pendientes_rows]
        logger.info(
            "Simulación horizonte raw: filas=%s fechas=%s kg_total=%s",
            len(pedidos_detalle_horizonte),
            sorted(set(str(r.get("Fecha salida", "")) for r in pedidos_detalle_horizonte)),
            sum(float(r.get("Kg pendiente", 0) or 0) for r in pedidos_detalle_horizonte),
        )
        if not pedidos:
            logger.info("Simulación abierta sin pedidos: modo análisis stock/sobrantes")

        filtros = self._filters_payload()
        cultivos = filtros.get("cultivo", [])
        cultivos_validos = [str(c or "").strip() for c in cultivos if str(c or "").strip() and str(c or "").strip().upper() != "TODOS"]
        if len(cultivos_validos) == 0:
            opciones_cultivo = self.service.get_filter_options_contextual("cultivo", filtros)
            if len(opciones_cultivo) == 1:
                cultivo_auto = str(opciones_cultivo[0]).strip()
                self.filter_widgets["cultivo"].set_selected([cultivo_auto])
                filtros = self._filters_payload()
                cultivos_validos = [cultivo_auto]
                logging.getLogger(__name__).info("Simulación cultivo autoseleccionado=%s por opción única", cultivo_auto)
            elif len(opciones_cultivo) > 1:
                messagebox.showwarning("Simulación de asignación", "Seleccione un único cultivo para simular.", parent=self)
                return
        if len(cultivos_validos) != 1:
            messagebox.showwarning("Simulación de asignación", "Seleccione un único cultivo para simular.", parent=self)
            return

        campanas = filtros.get("campana", [])
        campanas_validas = [str(c or "").strip() for c in campanas if str(c or "").strip() and str(c or "").strip().upper() != "TODOS"]
        if len(campanas_validas) != 1:
            messagebox.showwarning("Simulación de asignación", "Seleccione una única campaña para simular.", parent=self)
            return
        avisos: list[str] = []
        if not pedidos:
            avisos.append("No hay pedidos para esta campaña/cultivo.")
        if not self.stock_campo_rows:
            avisos.append("No hay stock de campo para esta campaña/cultivo.")
        if self.stock_almacen_rows:
            avisos.append("Hay stock de almacén disponible.")
        if (not pedidos and not self.stock_campo_rows) and self.stock_almacen_rows:
            avisos.append("Simulación sin datos operativos completos; se permite selección desde maestro.")
        if avisos:
            messagebox.showinfo("Simulación de asignación", "\n".join(avisos), parent=self)

        empresas = filtros.get("empresa", [])
        empresas_validas = [str(e or "").strip() for e in empresas if str(e or "").strip() and str(e or "").strip().upper() != "TODOS"]
        empresa_actual = empresas_validas[0] if len(empresas_validas) == 1 else ""

        abrir_simulacion_asignacion(
            self,
            pedidos,
            _candidatos_de_pedido,
            get_inventario_global_cb=lambda: inventario_global,
            pedidos_detalle_horizonte=pedidos_detalle_horizonte,
            cultivo_actual=cultivos_validos[0],
            campana_actual=campanas_validas[0],
            empresa_actual=empresa_actual,
            filters_payload=filtros,
        )

    def _open_selected_balance_coverage(self) -> None:
        sel = self.balance_table.tree.selection()
        if not sel:
            return
        values = self.balance_table.tree.item(sel[0], "values")
        id_conf = str(values[7]) if len(values) > 7 else ""
        variedad = str(values[3]) if len(values) > 3 else ""
        calibre = str(values[4]) if len(values) > 4 else ""
        categoria = str(values[5]) if len(values) > 5 else ""
        marca = str(values[6]) if len(values) > 6 else ""
        popup = tk.Toplevel(self)
        popup.title("Cobertura por pedidos")
        popup.geometry("1380x700")
        info = next((r for r in self.balance_rows_all if str(r.get("IdConfeccion", "")) == id_conf and str(r.get("Variedad", "")) == variedad and str(r.get("Calibre", "")) == calibre and str(r.get("Categoría", "")) == categoria and str(r.get("Marca", "")) == marca), None)
        if info:
            if str(info.get("Tipo línea", "")).strip() == "Sobrante comercial":
                messagebox.showinfo("Cobertura", "Esta línea es stock comercial disponible para venta. No requiere cobertura.", parent=popup)
                popup.destroy()
                return
        if not info:
            messagebox.showinfo("Cobertura", "No se pudo localizar la línea de balance seleccionada.", parent=popup)
            popup.destroy()
            return

        detalle_frame = ttk.LabelFrame(popup, text="Pedido seleccionado", padding=8)
        detalle_frame.pack(fill="x", padx=8, pady=(8, 4))
        ttk.Label(
            detalle_frame,
            text=(
                f"Variedad: {info.get('Variedad', '')} | Grupo varietal: {info.get('Grupo varietal', '')} | "
                f"Calibre pedido: {info.get('Calibre', '')} | Categoría: {info.get('Categoría', '')} | "
                f"Kg pendiente: {float(info.get('Kg pedidos pendientes', 0) or 0):,.2f} | "
                f"Cobertura posible: {info.get('Cobertura posible', '')} | "
                f"Kg cobertura exacta: {float(info.get('Kg cobertura exacta', 0) or 0):,.2f} | "
                f"Kg cobertura agrupada: {float(info.get('Kg cobertura agrupada', 0) or 0):,.2f} | "
                f"Kg cobertura potencial total: {float(info.get('Kg cobertura potencial total', 0) or 0):,.2f} | "
                f"Aviso: {info.get('Aviso', '')}"
            ),
            wraplength=1340,
        ).pack(anchor="w")

        stock_frame = ttk.LabelFrame(popup, text="Stock industrial compatible", padding=8)
        stock_frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        cols = ["Origen", "Tipo cobertura", "Cultivo", "Campaña", "Grupo varietal", "Variedad stock", "Marca stock", "Calibre stock", "Calibres coincidentes", "Categoría", "IdConfeccion stock", "Confección stock", "Palets", "Cajas", "Kg disponibles", "Score", "Flexibilidad aplicada", "Aviso", "Explicación"]
        tbl = DataTable(stock_frame, cols)
        tbl.pack(fill="both", expand=True)
        cobertura_rows = self.service.get_balance_cobertura_detalle(self._filters_payload(), info, policy=self._build_sim_policy())
        tbl.set_rows(cobertura_rows)


    def _ver_aprovechamiento_campo(self) -> None:
        sel = self.campo_table.tree.selection()
        if not sel:
            messagebox.showwarning("Stock campo", "Selecciona una partida de stock campo.", parent=self)
            return
        values = self.campo_table.tree.item(sel[0], "values")
        selected_row = {col: values[idx] if idx < len(values) else "" for idx, col in enumerate(self.campo_table.columns)}
        boleta = str(selected_row.get("Boleta", ""))

        def normalize_kg(value) -> float:
            try:
                return round(float(str(value).replace(".", "").replace(",", ".") if isinstance(value, str) and "," in value else value or 0), 3)
            except (TypeError, ValueError):
                return 0.0

        def partida_key(row: dict) -> tuple:
            return (
                str(row.get("Boleta", "")).strip(),
                str(row.get("Fecha carga", "")).strip(),
                normalize_kg(row.get("Kg campo", 0)),
                str(row.get("Socio", "")).strip(),
                str(row.get("Variedad", "")).strip(),
                str(row.get("Grupo varietal", "")).strip(),
            )

        selected_key = partida_key(selected_row)
        partida = next((r for r in self.stock_campo_rows if partida_key(r) == selected_key), None)
        if not partida:
            messagebox.showwarning("Stock campo", "No se pudo localizar la partida seleccionada.", parent=self)
            return

        popup = tk.Toplevel(self)
        popup.title(f"Aprovechamiento boleta {boleta}")
        popup.geometry("1200x650")
        view_mode = tk.StringVar(value="partida")
        current_rows: list[dict] = []

        header = ttk.LabelFrame(popup, text="Partida seleccionada", padding=8)
        header.pack(fill="x", padx=8, pady=(8, 4))
        header_lbl = ttk.Label(header, wraplength=1150)
        header_lbl.pack(anchor="w")

        body = ttk.LabelFrame(popup, text="Detalle", padding=8)
        body.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        msg = ttk.Label(body, text="")
        msg.pack(anchor="w", pady=(0, 6))
        cols = ["Boleta", "Calibre", "Categoría", "Kg real", "% aprovechamiento", "Kg campo aplicado", "Kg estimado", "Origen", "Explicación"]
        tbl = DataTable(body, cols)
        tbl.pack(fill="both", expand=True)

        def load_rows() -> tuple[list[dict], bool]:
            _resumen, detalle_map = self.service.get_aprovechamiento_stock_campo([partida], self._filters_payload())
            rows = detalle_map.get(boleta, [])
            tiene_real = any(str(r.get("Origen aprovechamiento", "")).upper() in {"REAL", "REAL_PESOSFRES", "LOTEADO", "HARVESTSYNC"} for r in rows)
            return rows, tiene_real

        def selected_estimated_id() -> int | None:
            sel_row = tbl.tree.selection()
            if not sel_row:
                return None
            vals = tbl.tree.item(sel_row[0], "values")
            if len(vals) < 8 or str(vals[7]).upper() != "ESTIMADO_MANUAL":
                return None
            calibre = str(vals[1])
            categoria = str(vals[2])
            for r in current_rows:
                if str(r.get("Origen aprovechamiento", "")).upper() == "ESTIMADO_MANUAL" and str(r.get("Calibre", "")) == calibre and str(r.get("Categoría", "")) == categoria:
                    return int(r.get("Id")) if r.get("Id") else None
            return None

        def open_form(existing: dict | None = None) -> None:
            rows_now, tiene_real_now = load_rows()
            if tiene_real_now:
                messagebox.showinfo("Aprovechamiento boleta", "El aprovechamiento real tiene prioridad", parent=popup)
                return
            form = tk.Toplevel(popup)
            form.title("Editar estimado" if existing else "Añadir estimado")
            form.geometry("420x310")
            form.transient(popup)
            form.grab_set()
            calibre_var = tk.StringVar(value=str(existing.get("Calibre", "") if existing else ""))
            categoria_var = tk.StringVar(value=str(existing.get("Categoría", "NORMAL") if existing else "NORMAL"))
            pct_var = tk.StringVar(value=str(existing.get("% aprovechamiento", "") if existing else ""))
            kg_campo = float(partida.get("Kg campo", 0) or 0)
            obs_var = tk.StringVar(value=str(existing.get("Observaciones", "") if existing else ""))
            kg_est_var = tk.StringVar(value="0.00")

            frm = ttk.Frame(form, padding=12)
            frm.pack(fill="both", expand=True)
            for idx, (label, var) in enumerate((("Calibre", calibre_var), ("Categoría", categoria_var), ("% aprovechamiento", pct_var), ("Kg campo aplicado", tk.StringVar(value=f"{kg_campo:.2f}")), ("Kg estimado", kg_est_var), ("Observaciones", obs_var))):
                ttk.Label(frm, text=label).grid(row=idx, column=0, sticky="w", pady=4)
                state = "readonly" if label in {"Kg campo aplicado", "Kg estimado"} else "normal"
                ttk.Entry(frm, textvariable=var, state=state).grid(row=idx, column=1, sticky="ew", pady=4)
            frm.columnconfigure(1, weight=1)

            def recalc(*_args) -> None:
                try:
                    pct = float(str(pct_var.get()).replace(",", "."))
                except ValueError:
                    pct = 0.0
                kg_est_var.set(f"{kg_campo * pct / 100:.2f}")
            pct_var.trace_add("write", recalc)
            recalc()

            def save() -> None:
                calibre = calibre_var.get().strip()
                categoria = categoria_var.get().strip().upper() or "NORMAL"
                try:
                    pct = float(pct_var.get().replace(",", "."))
                except ValueError:
                    messagebox.showwarning("Aprovechamiento estimado", "El porcentaje debe ser numérico.", parent=form)
                    return
                calibres_norm = self.service.normalizar_calibre_a_set(calibre)
                if not calibres_norm:
                    messagebox.showwarning("Aprovechamiento estimado", "Calibre obligatorio.", parent=form)
                    return
                if pct <= 0 or pct > 100:
                    messagebox.showwarning("Aprovechamiento estimado", "El porcentaje debe ser > 0 y <= 100.", parent=form)
                    return
                existentes = [r for r in rows_now if str(r.get("Origen aprovechamiento", "")).upper() == "ESTIMADO_MANUAL"]
                duplicados = []
                edit_id = existing.get("Id") if existing else None
                for r in existentes:
                    if edit_id and str(r.get("Id")) == str(edit_id):
                        continue
                    r_cal = self.service.normalizar_calibre_a_set(str(r.get("Calibre", "")))
                    if r_cal.intersection(calibres_norm) and str(r.get("Categoría", "")).upper() == categoria:
                        duplicados.append(r.get("Calibre"))
                if duplicados and not messagebox.askyesno("Duplicado", "Ya existe un estimado activo para el mismo calibre/categoría. ¿Deseas continuar?", parent=form):
                    return
                try:
                    self.service.upsert_aprovechamiento_estimado({
                        "Id": edit_id,
                        "Boleta": boleta,
                        "Campana": partida.get("Campaña", ""),
                        "Cultivo": partida.get("Cultivo", ""),
                        "Variedad": partida.get("Variedad", ""),
                        "GrupoVarietal": partida.get("Grupo varietal", ""),
                        "Categoria": categoria,
                        "Calibre": calibre,
                        "KgCampoAplicado": kg_campo,
                        "Porcentaje": pct,
                        "Observaciones": obs_var.get().strip(),
                    })
                except Exception as exc:
                    messagebox.showerror("Aprovechamiento estimado", str(exc), parent=form)
                    return
                form.destroy()
                self._invalidate_planning_cache("aprovechamiento_estimado_manual", keys=["stock_campo", "balance", "capacidad_productiva"])
                render()
                self.load_data(save_filters=False)

            btns = ttk.Frame(frm)
            btns.grid(row=6, column=0, columnspan=2, sticky="e", pady=(12, 0))
            ttk.Button(btns, text="Cancelar", command=form.destroy).pack(side="right", padx=(6, 0))
            ttk.Button(btns, text="Guardar", command=save).pack(side="right")

        def edit_selected() -> None:
            est_id = selected_estimated_id()
            if not est_id:
                messagebox.showwarning("Aprovechamiento estimado", "Selecciona una línea estimada manual.", parent=popup)
                return
            existing = next((r for r in current_rows if str(r.get("Id")) == str(est_id)), None)
            if existing:
                open_form(existing)

        def delete_selected() -> None:
            est_id = selected_estimated_id()
            if not est_id:
                messagebox.showwarning("Aprovechamiento estimado", "Selecciona una línea estimada manual.", parent=popup)
                return
            if not messagebox.askyesno("Eliminar estimado", "¿Eliminar el aprovechamiento estimado seleccionado?", parent=popup):
                return
            self.service.delete_aprovechamiento_estimado(est_id)
            self._invalidate_planning_cache("aprovechamiento_estimado_manual", keys=["stock_campo", "balance", "capacidad_productiva"])
            render()
            self.load_data(save_filters=False)

        def render() -> None:
            nonlocal current_rows
            current_rows, tiene_real = load_rows()
            header_lbl.configure(text=(f"Cultivo: {partida.get('Cultivo', '')} | Campaña: {partida.get('Campaña', '')} | Fecha carga: {partida.get('Fecha carga', '')} | "
                                   f"Semana: {partida.get('Semana', '')} | Socio: {partida.get('Socio', '')} | Variedad: {partida.get('Variedad', '')} | "
                                   f"Grupo varietal: {partida.get('Grupo varietal', '')} | Boleta: {partida.get('Boleta', '')} | Kg campo: {float(partida.get('Kg campo', 0) or 0):,.2f} | "
                                   f"Estado aprovechamiento: {('Real Loteado' if any(str(r.get('Origen aprovechamiento', '')).upper() == 'LOTEADO' for r in current_rows) else ('HarvestSync' if any(str(r.get('Origen aprovechamiento', '')).upper() == 'HARVESTSYNC' for r in current_rows) else ('Real PesosFres' if tiene_real else ('Estimado Manual' if current_rows else 'Sin aprovechamiento'))))}"))
            body.configure(text="Vista por partida" if view_mode.get() == "partida" else "Vista aprovechamiento medio")
            if tiene_real:
                msg.configure(text="El aprovechamiento real tiene prioridad")
            elif not current_rows:
                msg.configure(text="Esta partida no tiene aprovechamiento real ni estimado manual")
            else:
                msg.configure(text="Aprovechamiento estimado manual activo")
            rows = []
            for r in current_rows:
                origen = str(r.get("Origen aprovechamiento", "")).upper()
                kg = r.get("Kg disponibles", 0)
                rows.append({
                    "Boleta": boleta,
                    "Calibre": r.get("Calibre", ""),
                    "Categoría": r.get("Categoría", ""),
                    "Kg real": kg if origen in {"REAL", "REAL_PESOSFRES", "LOTEADO", "HARVESTSYNC"} else 0,
                    "% aprovechamiento": r.get("% aprovechamiento", 0),
                    "Kg campo aplicado": r.get("Kg campo origen", 0),
                    "Kg estimado": kg,
                    "Origen": origen if origen in {"REAL", "REAL_PESOSFRES", "LOTEADO", "HARVESTSYNC"} else "ESTIMADO_MANUAL",
                    "Explicación": r.get("Explicación", ""),
                })
            tbl.set_rows(rows)
            add_btn.configure(state="disabled" if tiene_real else "normal")
            edit_btn.configure(state="disabled" if tiene_real else "normal")
            del_btn.configure(state="disabled" if tiene_real else "normal")

        btnbar = ttk.Frame(popup)
        btnbar.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(btnbar, text="Alternar vista", command=lambda: (view_mode.set("media" if view_mode.get()=="partida" else "partida"), render())).pack(side="left")
        add_btn = ttk.Button(btnbar, text="Añadir estimado", command=lambda: open_form())
        add_btn.pack(side="right", padx=(6, 0))
        edit_btn = ttk.Button(btnbar, text="Editar estimado", command=edit_selected)
        edit_btn.pack(side="right", padx=(6, 0))
        del_btn = ttk.Button(btnbar, text="Eliminar estimado", command=delete_selected)
        del_btn.pack(side="right", padx=(6, 0))
        render()

    def _refresh_snapshot_info_label(self) -> None:
        info = self.runtime_db_service.get_snapshot_info()
        self.snapshot_info_var.set(info.get("label", "Foto de datos: No disponible"))

    def _actualizar_foto_local(self) -> None:
        ok, errors = self.runtime_db_service.prepare_runtime_databases(force=True)
        if not ok:
            messagebox.showwarning("Planificación diaria", self.runtime_db_service.WARNING_MESSAGE, parent=self)
            logging.getLogger(__name__).warning("No se pudo actualizar foto local en: %s", errors)
        self._refresh_snapshot_info_label()
        self._reload_with_invalidated_cache("actualizar_foto_local", save_filters=True)

    def _on_tab_changed(self, _event=None) -> None:
        self.load_data(save_filters=False)

    def reset_filters(self) -> None:
        for widget in self.filter_widgets.values():
            widget.clear()
        self.fecha_desde_var.set("")
        self.fecha_hasta_var.set("")
        self.pedidos_modo_var.set("10_dias")
        self._sync_pedidos_mode_combo()
        self.pedidos_grupo_confeccion_filter.clear()
        self.pedidos_perfil_confeccion_filter.clear()
        self._clear_saved_filters()
        self._load_filter_options(contextual=True)
        self.filters_status_var.set("Sin filtros activos")
        self._reload_with_invalidated_cache("limpiar_filtros", save_filters=True)

    def _diagnostico_resumen_rows(self, summary: dict) -> list[dict]:
        return [
            {"KPI": "Kg reales pendientes", "Valor": summary.get("Kg reales pendientes", 0)},
            {"KPI": "Kg previstos", "Valor": summary.get("Kg previstos", 0)},
            {"KPI": "Kg total simulación", "Valor": summary.get("Kg total simulación", 0)},
            {"KPI": "Horas necesarias", "Valor": summary.get("Horas necesarias estimadas", summary.get("Horas necesarias", 0))},
            {"KPI": "Horas disponibles", "Valor": summary.get("Horas disponibles", 0)},
            {"KPI": "Ocupación %", "Valor": summary.get("Ocupación %", 0)},
            {"KPI": "Turnos equivalentes", "Valor": summary.get("turnos_equivalentes", 0)},
            {"KPI": "Personal total", "Valor": summary.get("Personal disponible total", 0)},
            {"KPI": "Personal directo", "Valor": summary.get("Personal directo disponible", 0)},
            {"KPI": "Personal soporte", "Valor": summary.get("Personal soporte disponible", 0)},
            {"KPI": "Personal indirecto", "Valor": summary.get("Personal indirecto disponible", 0)},
            {"KPI": "Estado capacidad", "Valor": summary.get("Estado capacidad", "")},
        ]

    def _diagnostico_filter_rows(self, payload: dict) -> list[dict]:
        labels = {
            "campana": "Campaña",
            "cultivo": "Cultivo",
            "empresa": "Empresa",
            "semana": "Semana",
            "var_coop": "Variedad Coop",
            "grupo_varietal": "Grupo varietal",
            "marca": "Marca",
            "fecha_desde": "Fecha desde",
            "fecha_hasta": "Fecha hasta",
            "pedidos_modo": "Modo pedidos",
        }
        modo_label = next((label for label, value in self.PEDIDOS_MODOS if value == self.pedidos_modo_var.get()), self.pedidos_modo_var.get())
        rows = []
        for key, label in labels.items():
            value = modo_label if key == "pedidos_modo" else payload.get(key, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value) if value else "TODOS"
            rows.append({"Filtro": label, "Valor": value or "TODOS"})
        rows.append({"Filtro": "Grupo confección pedidos", "Valor": ", ".join(self.pedidos_grupo_confeccion_filter.get_selected()) or "TODOS"})
        rows.append({"Filtro": "Perfil confección pedidos", "Valor": ", ".join(self.pedidos_perfil_confeccion_filter.get_selected()) or "TODOS"})
        return rows

    def _diagnostico_pedidos_previstos_rows(self, cap: dict) -> list[dict]:
        rows = []
        for mapped in cap.get("pedidos", []):
            if str(mapped.get("tipo_pedido", "")).strip() != "Previsto":
                continue
            pedido = mapped.get("pedido", {}) or {}
            rows.append({
                "Id previsto": pedido.get("id_previsto", pedido.get("IdPedidoLora", "")),
                "Estado": pedido.get("estado", ""),
                "Fecha salida": pedido.get("Fecha salida", pedido.get("fecha_salida", "")),
                "Cliente": pedido.get("Cliente", pedido.get("cliente", "")),
                "Cultivo": pedido.get("Cultivo", pedido.get("cultivo", "")),
                "Campaña": pedido.get("Campaña", pedido.get("campana", "")),
                "Empresa": pedido.get("Empresa", pedido.get("empresa", "")),
                "Grupo varietal": pedido.get("Grupo varietal", pedido.get("grupo_varietal", "")),
                "Variedad": pedido.get("Variedad", pedido.get("variedad", "")),
                "Calibre": pedido.get("Calibre", pedido.get("calibre", "")),
                "Categoría": pedido.get("Categoría", pedido.get("categoria", "")),
                "Marca": pedido.get("Marca", pedido.get("marca", "")),
                "Confección prevista": pedido.get("Confección", pedido.get("confeccion_prevista", pedido.get("descripcion_base_packaging", ""))),
                "Grupo confección": pedido.get("Grupo confección", pedido.get("grupo_confeccion", "")),
                "Perfil confección": pedido.get("Perfil confección", pedido.get("perfil_confeccion", "")),
                "Kg estimados": mapped.get("kg", pedido.get("kg_estimados", 0)),
                "Palets estimados": pedido.get("palets_estimados", ""),
                "Familia productiva": mapped.get("familia", pedido.get("familia_productiva", "")),
                "Línea productiva": mapped.get("linea", pedido.get("linea_productiva", "")),
                "Observaciones": pedido.get("observaciones", ""),
            })
        return rows

    def _current_capacity_simulation(self, payload: dict) -> dict:
        if self.last_capacity_simulation is not None and self.last_capacity_payload == payload:
            return self.last_capacity_simulation
        cap = self.capacity_service.build_capacity_simulation(payload, self.pedidos_modo_var.get())
        self.last_capacity_simulation = cap
        self.last_capacity_payload = dict(payload)
        return cap

    def _write_excel_sheet(self, wb: Workbook, title: str, rows: list[dict], headers: list[str]) -> None:
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for idx, header in enumerate(headers, start=1):
            if any(token in header.lower() for token in ("kg", "hora", "ocupación", "%", "personal", "turnos", "palets", "cajas", "déficit", "factor", "capacidad")):
                for row_idx in range(2, ws.max_row + 1):
                    value = ws.cell(row_idx, idx).value
                    if isinstance(value, (int, float)):
                        ws.cell(row_idx, idx).number_format = "#,##0.00"
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_len + 2, 38)

    def _rows_from_table(self, table: DataTable) -> list[dict]:
        rows: list[dict] = []
        for item in table.tree.get_children():
            values = table.tree.item(item, "values")
            rows.append({col: values[idx] if idx < len(values) else "" for idx, col in enumerate(table.columns)})
        return rows

    def _commercial_pdf_filters(self) -> dict:
        payload = dict(self._filters_payload())
        payload["pedidos_modo_label"] = self._pedidos_mode_label(self.pedidos_modo_var.get())
        return payload

    def _ensure_commercial_pdf_rows_loaded(self) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], dict, dict]:
        payload = self._filters_payload()
        stock_campo_rows: list[dict] = []
        stock_almacen_rows: list[dict] = []
        pedidos_pendientes_rows: list[dict] = []
        prevision_recoleccion_rows: list[dict] = []
        aprovechamiento_volcado: dict = {}
        aprovechamiento_campo_detalle: dict = {}
        try:
            stock_campo_rows, _updated, _warning = self.service.load_stock_campo(payload)
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar stock campo para PDF comercial")
        try:
            stock_almacen_rows, _warning = self.service.load_stock_almacen(payload)
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar stock almacén para PDF comercial")
        try:
            prevision_recoleccion_rows = self.service.load_prevision_recoleccion(payload, today=datetime.now())
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar previsión de recolección para PDF comercial")
        try:
            pedidos_rows, pedidos_kpi = self.service.load_pedidos_pendientes(payload, self.pedidos_modo_var.get())
            self._refresh_pedidos_local_filter_options(pedidos_rows)
            pedidos_pendientes_rows, _ = self._apply_pedidos_local_filters(pedidos_rows, pedidos_kpi)
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar pedidos pendientes para PDF comercial")
        try:
            aprovechamiento_volcado = self.service.load_aprovechamiento_volcado(payload, today=datetime.now())
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar aprovechamiento de volcado para PDF comercial")
        try:
            _resumen_aprov, aprovechamiento_campo_detalle = self.service.get_aprovechamiento_stock_campo(stock_campo_rows, payload)
        except Exception:
            logging.getLogger(__name__).exception("No se pudo asegurar detalle de aprovechamiento campo para PDF comercial")
            aprovechamiento_campo_detalle = {}
        pedidos_previstos_rows = self._rows_from_table(self.pedidos_previstos_panel["table"]) if self.pedidos_previstos_panel else []
        return stock_campo_rows, stock_almacen_rows, prevision_recoleccion_rows, pedidos_pendientes_rows, pedidos_previstos_rows, aprovechamiento_volcado, aprovechamiento_campo_detalle

    def export_informe_comercial_pdf(self) -> None:
        generated_at = datetime.now()
        suggested_filename = self.commercial_pdf_service.default_filename(
            self.filter_widgets["cultivo"].get_selected(),
            now=generated_at,
        )
        should_save = messagebox.askyesno(
            "Informe comercial PDF",
            "¿Quieres guardar el informe?",
            parent=self,
        )
        if should_save:
            target = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=suggested_filename,
                parent=self,
            )
            if not target:
                messagebox.showinfo("Informe comercial PDF", "Guardado cancelado. No se generó el informe.", parent=self)
                return
        else:
            temp_dir = Path(tempfile.gettempdir()) / "SansebasAgroView"
            temp_dir.mkdir(exist_ok=True)
            target = temp_dir / suggested_filename

        try:
            stock_campo, stock_almacen, prevision_recoleccion, pedidos_pendientes, pedidos_previstos, aprovechamiento_volcado, aprovechamiento_campo_detalle = self._ensure_commercial_pdf_rows_loaded()
        except Exception as exc:
            messagebox.showerror("Informe comercial PDF", f"No se pudieron preparar los datos del informe: {exc}", parent=self)
            return
        try:
            path = self.commercial_pdf_service.generate(
                target,
                filters=self._commercial_pdf_filters(),
                stock_campo_rows=[dict(r) for r in stock_campo],
                stock_almacen_rows=[dict(r) for r in stock_almacen],
                prevision_recoleccion_rows=[dict(r) for r in prevision_recoleccion],
                pedidos_pendientes_rows=[dict(r) for r in pedidos_pendientes],
                pedidos_previstos_rows=[dict(r) for r in pedidos_previstos],
                aprovechamiento_volcado=aprovechamiento_volcado,
                aprovechamiento_campo_detalle={k: [dict(r) for r in v] for k, v in aprovechamiento_campo_detalle.items()},
                generated_at=generated_at,
            )
            self._open_pdf_preview(path)
        except Exception as exc:
            logging.getLogger(__name__).exception("No se pudo generar informe comercial PDF")
            messagebox.showerror("Informe comercial PDF", f"No se pudo generar el PDF: {exc}", parent=self)
            return
        if should_save:
            messagebox.showinfo("Informe comercial PDF", f"Informe guardado y abierto:\n{path}", parent=self)
        else:
            messagebox.showinfo(
                "Informe comercial PDF",
                "Informe generado en vista previa temporal. Guárdelo desde el visor PDF si lo desea.",
                parent=self,
            )

    def _open_pdf_preview(self, path: str | Path) -> None:
        path_str = str(path)
        system = platform.system()
        if system == "Windows":
            os.startfile(path_str)  # type: ignore[attr-defined]
        elif system == "Darwin":
            subprocess.call(["open", path_str])
        else:
            subprocess.call(["xdg-open", path_str])

    def _diagnostico_aprovechamiento_campo_rows(self, payload: dict) -> list[dict]:
        _resumen, detalle_map = self.service.get_aprovechamiento_stock_campo(self.stock_campo_rows, payload)
        rows: list[dict] = []
        for detalle_rows in detalle_map.values():
            for r in detalle_rows:
                rows.append({
                    "Boleta": r.get("Boleta", ""),
                    "Cultivo": r.get("Cultivo", ""),
                    "Campaña": r.get("Campaña", ""),
                    "Fecha carga": r.get("Fecha carga", ""),
                    "Calibre": r.get("Calibre", ""),
                    "Categoría": r.get("Categoría", ""),
                    "Kg disponibles": r.get("Kg disponibles", 0),
                    "% aprovechamiento": r.get("% aprovechamiento", 0),
                    "Origen aprovechamiento": r.get("Origen aprovechamiento", ""),
                    "Explicación": r.get("Explicación", ""),
                })
        return rows

    def export_diagnostico(self) -> None:
        payload = self._filters_payload()
        try:
            cap = self._current_capacity_simulation(payload)
            pedidos_rows, pedidos_kpi = self.service.load_pedidos_pendientes(payload, self.pedidos_modo_var.get())
            pedidos_pendientes, _ = self._apply_pedidos_local_filters(pedidos_rows, pedidos_kpi)
        except Exception as exc:
            messagebox.showerror("Exportar diagnóstico", f"No se pudo preparar el diagnóstico: {exc}", parent=self)
            return

        default_name = f"Diagnostico_Planificacion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        target = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel", "*.xlsx")])
        if not target:
            return

        sheets = [
            ("Resumen_KPI", self._diagnostico_resumen_rows(cap.get("summary", {})), ["KPI", "Valor"]),
            ("Familias", cap.get("family_rows", []), list(self.capacidad_family_table.columns)),
            ("Lineas", cap.get("line_rows", []), list(self.capacidad_line_table.columns)),
            ("Recursos", cap.get("resource_rows", []), list(self.capacidad_resource_table.columns)),
            ("Personal_requerido", cap.get("staffing_rows", []), list(self.capacidad_staffing_table.columns)),
            ("Cuellos_botella", cap.get("bottleneck_rows", []), list(self.capacidad_bottleneck_table.columns)),
            ("Incidencias", cap.get("incidencias", []), list(self.capacidad_inc_table.columns)),
            ("Pedidos_pendientes", pedidos_pendientes, list(self.pedidos_table.columns)),
            ("Pedidos_previstos", self._diagnostico_pedidos_previstos_rows(cap), ["Id previsto", "Estado", "Fecha salida", "Cliente", "Cultivo", "Campaña", "Empresa", "Grupo varietal", "Variedad", "Calibre", "Categoría", "Marca", "Confección prevista", "Grupo confección", "Perfil confección", "Kg estimados", "Palets estimados", "Familia productiva", "Línea productiva", "Observaciones"]),
            ("Aprovechamientos_estimados", self.service.get_aprovechamientos_estimados_filtrados(payload, boletas=[r.get("Boleta", "") for r in self.stock_campo_rows]), ["Id", "Boleta", "Campana", "Cultivo", "Variedad", "GrupoVarietal", "Categoria", "Calibre", "KgCampoAplicado", "Porcentaje", "KgEstimado", "Origen", "Activo", "Observaciones", "FechaCreacion", "FechaModificacion"]),
            ("Aprovechamiento_campo", self._diagnostico_aprovechamiento_campo_rows(payload), ["Boleta", "Cultivo", "Campaña", "Fecha carga", "Calibre", "Categoría", "Kg disponibles", "% aprovechamiento", "Origen aprovechamiento", "Explicación"]),
            ("Filtros_aplicados", self._diagnostico_filter_rows(payload), ["Filtro", "Valor"]),
        ]
        wb = Workbook()
        wb.remove(wb.active)
        for title, rows, headers in sheets:
            self._write_excel_sheet(wb, title, rows, headers)
        wb.save(Path(target))
        messagebox.showinfo("Exportar diagnóstico", f"Archivo guardado en:\n{target}", parent=self)


    def export_excel(self) -> None:
        tab = self.tabs.tab(self.tabs.select(), "text")
        if tab == "Stock campo":
            rows = self.stock_campo_rows
        elif tab == "Stock almacén":
            rows = self.stock_almacen_rows
        else:
            rows = self.pedidos_pendientes_rows if tab == "Pedidos pendientes" else self.balance_rows
        cultivos = self.filter_widgets["cultivo"].get_selected()
        campanas = self.filter_widgets["campana"].get_selected()
        path = self.service.export_rows_to_excel(rows, tab, cultivos, campanas)
        if path:
            messagebox.showinfo("Exportación", f"Archivo guardado en:\n{path}")



    def _actualizar_planificacion(self) -> None:
        confirm = messagebox.askyesno(
            "Actualizar planificación",
            "¿Quieres actualizar los datos de planificación desde hoy?\n\n"
            "Esto refrescará pedidos, loteado, lote y pesos frescos desde el día actual.\n"
            "Si estás trabajando con una foto fija, cancela esta acción.",
            parent=self,
        )
        if not confirm:
            return
        self._invalidate_planning_cache("actualizar_planificacion")
        self.stock_campo_rows = []
        self.stock_almacen_rows = []
        self.pedidos_pendientes_rows = []
        self.pedidos_pendientes_rows_raw = []
        self._btn_actualizar_planificacion.configure(state="disabled", text="Actualizando...")
        try:
            ok, msg = self.service.actualizar_planificacion_hoy_en_adelante()
            if not ok:
                messagebox.showerror("Actualizar planificación", msg, parent=self)
                return
            messagebox.showinfo("Actualizar planificación", msg.replace(" | ", "\n").replace("Planificación rápida OK. ", ""), parent=self)
            self._load_filter_options()
            self._reload_with_invalidated_cache("actualizar_planificacion", save_filters=True)
        except ValueError as exc:
            msg = str(exc)
            if "No existe configuración legacy" in msg and "DBPedidos.sqlite:Pedidos" in msg:
                messagebox.showwarning(
                    "Actualizar planificación",
                    "No hay configuración legacy para DBPedidos.sqlite:Pedidos. "
                    "Configúrala en Configuración > Actualización tablas legacy para habilitar esta acción.",
                    parent=self,
                )
                logging.getLogger(__name__).warning(msg)
                return
            logging.getLogger(__name__).exception("Error de configuración al actualizar planificación")
            messagebox.showerror("Actualizar planificación", msg, parent=self)
        except Exception as exc:
            logging.getLogger(__name__).exception("Error inesperado al actualizar planificación")
            messagebox.showerror("Actualizar planificación", f"No se pudo actualizar la planificación: {exc}", parent=self)
        finally:
            self._btn_actualizar_planificacion.configure(state="normal", text="Actualizar planificación")

    def _pedidos_mode_label(self, mode_key: str) -> str:
        for label, key in self.PEDIDOS_MODOS:
            if key == mode_key:
                return label
        return "Próximos 10 días"

    def _sync_pedidos_mode_combo(self) -> None:
        if hasattr(self, "_pedidos_modo_combo"):
            self._pedidos_modo_combo.set(self._pedidos_mode_label(self.pedidos_modo_var.get()))

    def _on_pedidos_modo_changed(self, _event=None) -> None:
        label = self._pedidos_modo_combo.get()
        lookup = {lbl: key for lbl, key in self.PEDIDOS_MODOS}
        self.pedidos_modo_var.set(lookup.get(label, "10_dias"))
        self._reload_with_invalidated_cache("modo_pedidos", save_filters=True)

    def _format_filters_status(self, payload: dict) -> str:
        labels = {
            "pedidos_modo": "Modo pedidos",
            "campana": "Campaña",
            "cultivo": "Cultivo",
            "empresa": "Empresa",
            "semana": "Semana",
            "var_coop": "Variedad Coop",
            "grupo_varietal": "Grupo varietal",
            "marca": "Marca",
            "fecha_desde": "Fecha desde",
            "fecha_hasta": "Fecha hasta",
        }
        has_values = any(payload.get(k) for k in labels)
        if not has_values:
            return "Sin filtros activos"
        parts: list[str] = []
        for key in ("pedidos_modo", "campana", "cultivo", "empresa", "semana", "var_coop", "grupo_varietal", "marca", "fecha_desde", "fecha_hasta"):
            value = payload.get(key, []) if key in self.FILTER_KEYS else payload.get(key, "")
            if key == "pedidos_modo":
                display = self._pedidos_mode_label(str(value).strip())
            elif isinstance(value, list):
                display = ",".join(value) if value else "Todos"
            else:
                display = str(value).strip() or "Todos"
            parts.append(f"{labels[key]}={display}")
        return "Filtros activos: " + " | ".join(parts)

    def _save_filters(self, payload: dict) -> None:
        enriched = dict(payload)
        enriched["pedidos_grupo_confeccion"] = self.pedidos_grupo_confeccion_filter.get_selected()
        enriched["pedidos_perfil_confeccion"] = self.pedidos_perfil_confeccion_filter.get_selected()
        self.FILTERS_FILE.parent.mkdir(parents=True, exist_ok=True)
        self.FILTERS_FILE.write_text(json.dumps(enriched, ensure_ascii=False, indent=2), encoding="utf-8")

    def _load_filters(self) -> None:
        if not self.FILTERS_FILE.exists():
            self.filters_status_var.set("Sin filtros activos")
            return
        try:
            payload = json.loads(self.FILTERS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return
        for key in self.FILTER_KEYS:
            raw = payload.get(key, [])
            values = raw if isinstance(raw, list) else ([str(raw)] if str(raw or "").strip() else [])
            self.filter_widgets[key].set_selected([str(v).strip() for v in values if str(v or "").strip()])
        self.fecha_desde_var.set(str(payload.get("fecha_desde", "") or "").strip())
        self.fecha_hasta_var.set(str(payload.get("fecha_hasta", "") or "").strip())
        self.pedidos_modo_var.set(str(payload.get("pedidos_modo", "10_dias") or "10_dias").strip() or "10_dias")
        pedidos_gc = payload.get("pedidos_grupo_confeccion", [])
        pedidos_pc = payload.get("pedidos_perfil_confeccion", [])
        gc_vals = pedidos_gc if isinstance(pedidos_gc, list) else ([str(pedidos_gc)] if str(pedidos_gc or "").strip() else [])
        pc_vals = pedidos_pc if isinstance(pedidos_pc, list) else ([str(pedidos_pc)] if str(pedidos_pc or "").strip() else [])
        self.pedidos_grupo_confeccion_filter.set_selected([str(v).strip() for v in gc_vals if str(v or "").strip()])
        self.pedidos_perfil_confeccion_filter.set_selected([str(v).strip() for v in pc_vals if str(v or "").strip()])
        self._sync_pedidos_mode_combo()
        self.filters_status_var.set(self._format_filters_status(self._filters_payload()))

    def _refresh_pedidos_local_filter_options(self, pedidos_rows: list[dict]) -> None:
        selected_gc = self.pedidos_grupo_confeccion_filter.get_selected()
        selected_pc = self.pedidos_perfil_confeccion_filter.get_selected()
        gc_options = sorted({str(r.get("Grupo confección", "")).strip() for r in pedidos_rows if str(r.get("Grupo confección", "")).strip()})
        pc_options = sorted({str(r.get("Perfil confección", "")).strip() for r in pedidos_rows if str(r.get("Perfil confección", "")).strip()})
        self.pedidos_grupo_confeccion_filter.set_options(gc_options)
        self.pedidos_perfil_confeccion_filter.set_options(pc_options)
        self.pedidos_grupo_confeccion_filter.set_selected([v for v in selected_gc if v in gc_options])
        self.pedidos_perfil_confeccion_filter.set_selected([v for v in selected_pc if v in pc_options])

    def _apply_pedidos_local_filters(self, pedidos_rows: list[dict], pedidos_kpi: dict) -> tuple[list[dict], dict]:
        gc_selected = set(self.pedidos_grupo_confeccion_filter.get_selected())
        pc_selected = set(self.pedidos_perfil_confeccion_filter.get_selected())
        filtered = [
            row for row in pedidos_rows
            if (not gc_selected or str(row.get("Grupo confección", "")).strip() in gc_selected)
            and (not pc_selected or str(row.get("Perfil confección", "")).strip() in pc_selected)
        ]
        if not gc_selected and not pc_selected:
            return filtered, pedidos_kpi
        kg_pedido = sum(float(r.get("Kg pedido teórico", 0) or 0) for r in filtered)
        merma = sum(float(r.get("Merma kg", 0) or 0) for r in filtered)
        return filtered, {
            "Kg pedido teórico total": kg_pedido,
            "Kg hecho real total": sum(float(r.get("Kg hecho real", 0) or 0) for r in filtered),
            "Kg pendiente total": sum(float(r.get("Kg pendiente", 0) or 0) for r in filtered),
            "Merma kg total": merma,
            "% merma total": (merma / max(kg_pedido, 1e-9)) * 100,
            "Nº pedidos": len({str(r.get("IdPedidoLora", "")).strip() for r in filtered if str(r.get("IdPedidoLora", "")).strip()}),
            "Nº líneas": len(filtered),
            "Nº líneas sin datos": sum(1 for r in filtered if str(r.get("Estado", "")).strip() == "Sin datos"),
            "Nº líneas parciales": sum(1 for r in filtered if str(r.get("Estado", "")).strip() == "Parcial"),
        }

    def _clear_saved_filters(self) -> None:
        if self.FILTERS_FILE.exists():
            self.FILTERS_FILE.unlink()

    def show_detalle_palets(self) -> None:
        if not self.stock_almacen_detalle_rows:
            messagebox.showinfo("Planificación diaria", "No hay detalle de palets para los filtros actuales.")
            return
        modal = tk.Toplevel(self)
        modal.title("Detalle palets - Stock almacén")
        modal.geometry("1320x620")
        modal.transient(self.winfo_toplevel())
        modal.grab_set()
        table = DataTable(
            modal,
            ["IdPalet", "Pedido", "FechaAlmacen", "Estado", "Terminado", "Variedad", "Grupo varietal", "Calibre", "Categoria", "Marca", "IdConfeccion", "Confeccion", "Cajas", "Neto"],
        )
        table.pack(fill="both", expand=True, padx=10, pady=10)
        rows = []
        for row in self.stock_almacen_detalle_rows:
            rows.append(
                {
                    "IdPalet": row.get("IdPalet", ""),
                    "Pedido": row.get("Pedido", ""),
                    "FechaAlmacen": row.get("FechaAlmacen", ""),
                    "Estado": row.get("Estado", ""),
                    "Terminado": row.get("Terminado", ""),
                    "Variedad": row.get("Variedad", ""),
                    "Grupo varietal": row.get("GrupoVarietal", ""),
                    "Calibre": row.get("Calibre", ""),
                    "Categoria": row.get("Categoria", ""),
                    "Marca": row.get("Marca", ""),
                    "IdConfeccion": row.get("IdConfeccion", ""),
                    "Confeccion": row.get("Confeccion", ""),
                    "Cajas": round(float(row.get("Cajas") or 0), 2),
                    "Neto": round(float(row.get("Neto") or 0), 2),
                }
            )
        table.set_rows(rows)


    def _build_sim_policy(self) -> dict:
        return {k: bool(v.get()) for k, v in self.sim_policy_vars.items()}

    def _recalcular_simulacion(self) -> None:
        self._save_balance_settings()
        self._reload_with_invalidated_cache("recalcular_simulacion", save_filters=True)

    def _save_balance_settings(self) -> None:
        self.BALANCE_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        self.BALANCE_SETTINGS_FILE.write_text(json.dumps(self._build_sim_policy(), ensure_ascii=False, indent=2), encoding="utf-8")

    def _load_balance_settings(self) -> None:
        if not self.BALANCE_SETTINGS_FILE.exists():
            return
        try:
            payload = json.loads(self.BALANCE_SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return
        for k, var in self.sim_policy_vars.items():
            if k in payload:
                var.set(bool(payload.get(k)))
