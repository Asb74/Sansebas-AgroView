import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from services.forfait_service import ForfaitService
from widgets.data_table import DataTable


class MapForfaitScreen(ttk.Frame):
    PENDING_EXPORT_COLUMNS = [
        "Campaña",
        "Cultivo",
        "IdConfeccion",
        "GRUPO",
        "Eur/kg Material",
        "Eur/kg Recoleción y Transporte",
        "Eur/kg Gastos Generales",
        "Eur/kg Mano obra",
        "Eur/kg total",
    ]
    TABLE_COLUMNS = [
        "Campaña", "Cultivo", "IdConfeccion", "NombreConfeccion",
        "GrupoConfeccion", "Marca", "CosteMaterialEurKg", "CosteManoObraEurKg", "CosteTotalEurKg", "Estado", "OrigenCoste",
    ]

    def __init__(self, master: tk.Misc, on_back) -> None:
        super().__init__(master, padding=16)
        self.on_back = on_back
        self.service = ForfaitService()
        self.cultivo_var = tk.StringVar(value="CITRICOS")
        self.campana_var = tk.StringVar(value="2025")
        self.only_missing_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="")
        self.rows: list[dict[str, Any]] = []
        self._build_ui()

    def _build_ui(self) -> None:
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        header = ttk.Frame(self)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)
        ttk.Label(header, text="Control cobertura confecciones", style="Section.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Button(header, text="Volver", command=self.on_back).grid(row=0, column=1, sticky="e")

        controls = ttk.LabelFrame(self, text="Filtro obligatorio", padding=12)
        controls.grid(row=1, column=0, sticky="ew", pady=(8, 8))
        ttk.Label(controls, text="Cultivo").grid(row=0, column=0, sticky="w", padx=(0, 6))
        ttk.Entry(controls, textvariable=self.cultivo_var, width=18).grid(row=0, column=1, sticky="w", padx=(0, 12))
        ttk.Label(controls, text="Campaña").grid(row=0, column=2, sticky="w", padx=(0, 6))
        ttk.Entry(controls, textvariable=self.campana_var, width=18).grid(row=0, column=3, sticky="w", padx=(0, 12))
        ttk.Button(controls, text="Cargar cobertura", command=self.load_rows).grid(row=0, column=4, padx=(0, 8))
        ttk.Checkbutton(controls, text="Mostrar solo sin forfait", variable=self.only_missing_var, command=self.load_rows).grid(row=0, column=5, padx=(0, 8))
        ttk.Button(controls, text="Exportar pendientes", command=self.export_pending).grid(row=0, column=6, padx=(0, 8))
        ttk.Button(controls, text="Reiniciar forfait", command=self.reset_forfait).grid(row=0, column=7)

        self.table = DataTable(self, columns=self.TABLE_COLUMNS)
        self.table.grid(row=2, column=0, sticky="nsew")
        ttk.Label(self, textvariable=self.status_var).grid(row=3, column=0, sticky="w", pady=(8, 0))

    def load_rows(self) -> None:
        cultivo = self.cultivo_var.get().strip()
        campana = self.campana_var.get().strip()
        if not cultivo or not campana:
            messagebox.showwarning("Filtro obligatorio", "Indica cultivo y campaña.", parent=self)
            return
        try:
            self.rows = self.service.fetch_coverage_rows(cultivo, campana, self.only_missing_var.get())
        except Exception as exc:
            messagebox.showerror("Cobertura confecciones", str(exc), parent=self)
            return
        self.table.set_rows(self._map_rows(self.rows))
        self._set_counter_status()

    def reset_forfait(self) -> None:
        cultivo = self.cultivo_var.get().strip()
        campana = self.campana_var.get().strip()
        if not cultivo or not campana:
            messagebox.showwarning("Filtro obligatorio", "Indica cultivo y campaña.", parent=self)
            return
        confirm = (
            "Vas a eliminar el forfait importado para:\n"
            f"Cultivo: {cultivo}\n"
            f"Campaña: {campana}\n\n"
            "Esto pondrá a cero los costes cargados para esta campaña/cultivo.\n"
            "No se eliminarán pedidos ni MConfecciones.\n\n"
            "¿Quieres continuar?"
        )
        if not messagebox.askyesno("Reiniciar forfait", confirm, parent=self):
            return
        try:
            self.service.reset_related_forfait(cultivo, campana)
        except Exception as exc:
            messagebox.showerror("Cobertura confecciones", str(exc), parent=self)
            return
        self.load_rows()
        messagebox.showinfo("Reiniciar forfait", f"Forfait reiniciado correctamente para {cultivo} {campana}.", parent=self)

    def export_pending(self) -> None:
        pending = [
            r for r in self.rows
            if self._is_pending_for_export(r)
        ]
        if not pending:
            messagebox.showinfo("Exportar pendientes", "No hay confecciones pendientes para exportar.", parent=self)
            return
        cultivo = (self.cultivo_var.get().strip() or "cultivo").replace(" ", "_")
        campana = (self.campana_var.get().strip() or "campana").replace(" ", "_")
        suggested_name = f"forfait_pendientes_{cultivo}_{campana}.xlsx"
        out_path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            initialfile=suggested_name,
            filetypes=[("Excel", "*.xlsx")],
            title="Exportar pendientes",
        )
        if not out_path:
            return
        try:
            self._export_pending_excel(out_path, pending)
        except Exception as exc:
            messagebox.showerror("Exportar pendientes", str(exc), parent=self)
            return
        self.status_var.set(f"Pendientes exportados: {len(pending)}")
        messagebox.showinfo(
            "Exportar pendientes",
            "Exportación generada correctamente. Rellena los costes y vuelve a importar el Excel.",
            parent=self,
        )

    def _export_pending_excel(self, out_path: str, pending_rows: list[dict[str, Any]]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ForfaitPendientes"

        worksheet.append(self.PENDING_EXPORT_COLUMNS)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for row in pending_rows:
            worksheet.append(
                [
                    row.get("Campaña", ""),
                    row.get("Cultivo", ""),
                    row.get("IdConfeccion", ""),
                    row.get("GrupoConfeccion", ""),
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

        for column_index, _column_name in enumerate(self.PENDING_EXPORT_COLUMNS, start=1):
            letter = get_column_letter(column_index)
            if column_index >= 5:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column_index).number_format = "#,##0.000000"
            max_length = 0
            for cell in worksheet[letter]:
                value_len = len(str(cell.value or ""))
                if value_len > max_length:
                    max_length = value_len
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)

        workbook.save(out_path)

    @staticmethod
    def _is_pending_for_export(row: dict[str, Any]) -> bool:
        estado = str(row.get("Estado") or "").upper()
        origen_coste = str(row.get("OrigenCoste") or "").upper()
        if estado in {"SIN_FORFAIT", "REVISAR"} or origen_coste == "SIN_FORFAIT":
            return True
        total = row.get("CosteTotalEurKg")
        if total in (None, ""):
            return True
        try:
            return float(total) == 0.0
        except (TypeError, ValueError):
            return True

    def _set_counter_status(self) -> None:
        exactos = sum(1 for r in self.rows if r.get("OrigenCoste") == "EXACTO")
        sinf = sum(1 for r in self.rows if r.get("OrigenCoste") == "SIN_FORFAIT")
        self.status_var.set(f"Confecciones usadas: {len(self.rows)} | Con coste exacto: {exactos} | Sin forfait: {sinf}")

    def _map_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        out = []
        for row in rows:
            estado = str(row.get("Estado") or "")
            if estado in {"", "IMPORTADO"}:
                estado = "OK" if row.get("OrigenCoste") == "EXACTO" else ("APROXIMADO" if row.get("OrigenCoste") != "SIN_FORFAIT" else "SIN_FORFAIT")
            tag = "tag_green" if estado == "OK" else ("tag_red" if estado == "SIN_FORFAIT" else ("tag_orange" if estado == "REVISAR" else "tag_yellow"))
            out.append({
                "Campaña": row.get("Campaña", ""),
                "Cultivo": row.get("Cultivo", ""),
                "IdConfeccion": row.get("IdConfeccion", ""),
                "NombreConfeccion": row.get("NombreConfeccion", ""),
                "GrupoConfeccion": row.get("GrupoConfeccion", ""),
                "Marca": row.get("Marca", ""),
                "CosteMaterialEurKg": self._fmt_optional(row.get("CosteMaterialEurKg"), 4),
                "CosteManoObraEurKg": self._fmt_optional(row.get("CosteManoObraEurKg"), 4),
                "CosteTotalEurKg": self._fmt_optional(row.get("CosteTotalEurKg"), 4),
                "Estado": estado,
                "OrigenCoste": row.get("OrigenCoste", "SIN_FORFAIT"),
                "__tags__": tag,
            })
        return out

    @staticmethod
    def _fmt_optional(value: Any, decimals: int) -> str:
        if value is None or value == "":
            return ""
        try:
            return f"{float(value):,.{decimals}f}"
        except (TypeError, ValueError):
            return ""
