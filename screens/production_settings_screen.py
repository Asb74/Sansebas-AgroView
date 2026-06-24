from __future__ import annotations

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from db.production_settings_repository import infer_default_staff_type, staff_type_flags
from services.production_settings_service import ProductionSettingsService
from utils.master_excel_io import export_master_to_excel, import_master_from_excel, is_blank, normalize_bool, normalize_headers
from utils.help_dialog import show_tab_help
from utils.production_master_excel_configs import PRODUCTION_MASTER_EXCEL_CONFIGS
from utils.production_help_texts import PRODUCTION_CALIBER_FACTORS_HELP, PRODUCTION_CALIBER_FACTORS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP, PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_FIELD_HELP, PRODUCTION_LINES_HELP, PRODUCTION_LINES_HELP_KEYS, PRODUCTION_PACKAGING_HELP, PRODUCTION_PACKAGING_HELP_KEYS, PRODUCTION_PACKAGING_MAPPING_HELP, PRODUCTION_PACKAGING_MAPPING_HELP_KEYS, PRODUCTION_PENALTIES_HELP, PRODUCTION_PENALTIES_HELP_KEYS, PRODUCTION_PERFORMANCE_HELP, PRODUCTION_PERFORMANCE_HELP_KEYS, PRODUCTION_PERSONAL_HELP, PRODUCTION_PERSONAL_HELP_KEYS, PRODUCTION_FLOW_STAFFING_HELP, PRODUCTION_FLOW_STAFFING_HELP_KEYS, PRODUCTION_SEMAPHORE_HELP, PRODUCTION_SEMAPHORE_HELP_KEYS, PRODUCTION_RESOURCES_HELP, PRODUCTION_RESOURCES_HELP_KEYS, get_help_items
from widgets.screen_header import ScreenHeader


class ProductionSettingsScreen(ttk.Frame):
    VOLCADO_OPTIONS = ["Compacta", "Línea invierno", "Línea verano", "Tolva", "Manual"]
    STAFF_TYPES = ["Directo", "Soporte", "Indirecto"]
    PACKAGING_FAMILIES = ["Malla", "Encajado", "Granel", "Granelera", "Flowpack", "Otro"]
    PACKAGING_SUBTYPES = ["Tradicional", "Clip-to-clip", "Girsac", "Bolsa", "Caja cartón", "Caja madera", "Granel", "Granelera", "Otro"]
    PACKAGING_MATERIALS = ["Plástico", "Cartón", "Madera", "Malla", "Mixto", "Sin material", "Otro"]
    PACKAGING_MESH_TYPES = ["No aplica", "Tradicional", "Clip-to-clip", "Girsac", "Bolsa", "Otro"]
    MAPPING_FAMILIES = ["Malla", "Encajado", "Granel", "Granelera", "Flowpack", "Alvéolo", "Cestas", "Otro"]
    MAPPING_SUBTYPES = ["Tradicional", "Clip-to-clip", "Girsac", "Flowpack", "Caja", "Alvéolo", "Granel", "Granelera", "Cesta", "Otro"]
    MAPPING_MESH_TYPES = ["No aplica", "Tradicional", "Clip-to-clip", "Girsac", "Flowpack", "Otro"]
    LINE_TYPES = ["Volcado", "Malla", "Encajado", "Granel", "Granelera", "Calibrador", "Final línea", "Expedición", "Soporte", "Otro"]
    LINE_FAMILIES = ["Entrada fruta", "Producción directa", "Clasificación", "Envasado", "Salida / expedición", "Apoyo"]
    PERFORMANCE_FAMILIES = ["Malla", "Encajado", "Granel", "Granelera", "Volcado", "Calibrador", "Final línea", "Otro"]
    PERFORMANCE_LINE_TYPES = PERFORMANCE_FAMILIES
    PERFORMANCE_CONDITIONS = ["Normal", "Calibre pequeño", "Calibre grande", "Con BOX", "Con precalibrado", "Sin precalibrado", "Destrío alto", "Pedido pequeño", "Otro"]
    PERFORMANCE_DIFFICULTIES = ["Baja", "Media", "Alta", "Muy alta"]
    PENALTY_TYPES = ["Cambio cliente", "Cambio plataforma", "Cambio formato kg", "Cambio material", "Cambio tipo malla", "Cambio etiqueta", "Cambio confección", "Cambio calibre", "Cambio categoría", "Pedido pequeño", "Arranque línea", "Parada línea", "Limpieza", "Espera fruta", "Espera material", "Otro"]
    PENALTY_SCOPES = ["General", "Malla", "Encajado", "Granel", "Granelera", "Volcado", "Expedición", "Línea específica", "Otro"]
    PENALTY_APPLIES = ["Cada cambio", "Cada pedido", "Cada línea de pedido", "Cada cliente", "Cada plataforma", "Cada jornada", "Cada arranque", "Cada parada", "Otro"]
    SEMAPHORE_RULE_TYPES = ["Saturación capacidad", "Falta personal", "Exceso pedidos", "Exceso cambios", "Línea no operativa", "Fecha salida crítica", "Pedido previsto", "Cuello de botella", "Stock insuficiente", "Rendimiento bajo", "Otro"]
    SEMAPHORE_SCOPES = ["General", "Malla", "Encajado", "Granel", "Granelera", "Volcado", "Expedición", "Personal", "Línea específica", "Pedido", "Otro"]
    SEMAPHORE_METRICS = ["ocupacion_pct", "horas_faltantes", "personas_faltantes", "pedidos_dia", "cambios_formato", "cambios_cliente", "pedidos_pequenos", "kg_pendientes", "palets_pendientes", "dias_hasta_salida", "rendimiento_pct", "lineas_activas", "stock_cobertura_pct", "otro"]
    SEMAPHORE_OPERATORS = [">=", ">", "<=", "<", "=", "!="]
    CALIBER_FAMILIES = ["Malla", "Malla 1 kg", "Malla 1.5 kg", "Malla 2 kg", "Malla 3 kg", "Malla 4 kg", "Encajado", "Granel", "Granelera", "Otro"]
    CALIBER_GROUPS = ["Pequeño", "Medio", "Grande", "Personalizado"]
    CALIBER_APPLIES = ["OPH", "Kg/h", "Ambos"]

    def __init__(self, master: tk.Misc, on_back) -> None:
        super().__init__(master, padding=10)
        self.on_back = on_back
        self.service = ProductionSettingsService()

        self._general_vars: dict[str, tk.Variable] = {}
        self._tipos_volcado_vars: dict[str, tk.IntVar] = {}
        self._staff_summary_vars: dict[str, tk.Variable] = {}
        self._staff_editor_vars: dict[str, tk.Variable] = {}
        self._flow_staffing_editor_vars: dict[str, tk.Variable] = {}
        self._flow_staffing_tree: ttk.Treeview | None = None
        self._staff_tree: ttk.Treeview | None = None
        self._packaging_tree: ttk.Treeview | None = None
        self._packaging_editor_vars: dict[str, tk.Variable] = {}
        self._base_packaging_tree: ttk.Treeview | None = None
        self._base_packaging_editor_vars: dict[str, tk.Variable] = {}
        self._lines_tree: ttk.Treeview | None = None
        self._mapping_tree: ttk.Treeview | None = None
        self._mapping_editor_vars: dict[str, tk.Variable] = {}
        self._lines_editor_vars: dict[str, tk.Variable] = {}
        self._new_line_counter = 1
        self._new_packaging_counter = 1
        self._performance_tree: ttk.Treeview | None = None
        self._performance_editor_vars: dict[str, tk.Variable] = {}
        self._new_performance_counter = 1
        self._penalties_tree: ttk.Treeview | None = None
        self._penalties_editor_vars: dict[str, tk.Variable] = {}
        self._new_penalty_counter = 1
        self._semaphore_tree: ttk.Treeview | None = None
        self._semaphore_editor_vars: dict[str, tk.Variable] = {}
        self._new_semaphore_counter = 1
        self._caliber_tree: ttk.Treeview | None = None
        self._physical_resources_tree: ttk.Treeview | None = None
        self._resource_compatibilities_tree: ttk.Treeview | None = None
        self._resource_feeds_tree: ttk.Treeview | None = None
        self._resource_availability_tree: ttk.Treeview | None = None
        self._physical_resources_editor_vars: dict[str, tk.Variable] = {}
        self._resource_compatibilities_editor_vars: dict[str, tk.Variable] = {}
        self._resource_feeds_editor_vars: dict[str, tk.Variable] = {}
        self._resource_availability_editor_vars: dict[str, tk.Variable] = {}
        self._caliber_editor_vars: dict[str, tk.Variable] = {}
        self._productive_families_tree: ttk.Treeview | None = None
        self._line_capacity_config_tree: ttk.Treeview | None = None
        self._line_required_resources_tree: ttk.Treeview | None = None
        self._staff_area_equivalences_tree: ttk.Treeview | None = None
        self._staff_polyvalence_tree: ttk.Treeview | None = None
        self._capacity_productive_notebook: ttk.Notebook | None = None
        self._capacity_master_actions: dict[str, tuple[str, object, object]] = {}
        self._new_caliber_counter = 1
        self._build_ui()
        self._load_general_settings()
        self._load_staff_settings()
        self._load_flow_staffing_settings()
        self._load_packaging_settings()
        self._load_base_packaging_settings()
        self._load_lines_settings()
        self._load_physical_resources_settings()
        self._load_resource_compatibilities_settings()
        self._load_resource_feeds_settings()
        self._load_resource_availability_settings()
        self._load_packaging_mapping_settings()
        self._load_performance_settings()
        self._load_penalty_settings()
        self._load_semaphore_settings()
        self._load_caliber_factors_settings()
        self._load_capacity_productive_settings()

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        ScreenHeader(
            self,
            title="Configuración productiva",
            subtitle="Definición operativa diaria del almacén",
            on_back=self.on_back,
        ).grid(row=0, column=0, sticky="ew")

        notebook = ttk.Notebook(self)
        notebook.grid(row=1, column=0, sticky="nsew", pady=(8, 0))

        general_tab = ttk.Frame(notebook, padding=12)
        notebook.add(general_tab, text="General del día")
        personal_tab = ttk.Frame(notebook, padding=12)
        notebook.add(personal_tab, text="Personal")
        flow_staffing_tab = ttk.Frame(notebook, padding=12)
        notebook.add(flow_staffing_tab, text="Dotación flujos")
        packaging_tab = ttk.Frame(notebook, padding=12)
        notebook.add(packaging_tab, text="Confecciones")
        lines_tab = ttk.Frame(notebook, padding=12)
        notebook.add(lines_tab, text="Máquinas / líneas")
        resources_tab = ttk.Frame(notebook, padding=12)
        notebook.add(resources_tab, text="Recursos y flujos")
        capacity_tab = ttk.Frame(notebook, padding=12)
        notebook.add(capacity_tab, text="Capacidad productiva")
        base_packaging_tab = ttk.Frame(notebook, padding=12)
        notebook.add(base_packaging_tab, text="Confecciones base")
        mapping_tab = ttk.Frame(notebook, padding=12)
        notebook.add(mapping_tab, text="Mapeo confecciones")
        performance_tab = ttk.Frame(notebook, padding=12)
        notebook.add(performance_tab, text="Rendimientos")
        penalties_tab = ttk.Frame(notebook, padding=12)
        notebook.add(penalties_tab, text="Penalizaciones")
        semaphore_tab = ttk.Frame(notebook, padding=12)
        notebook.add(semaphore_tab, text="Reglas / semáforo")
        caliber_tab = ttk.Frame(notebook, padding=12)
        notebook.add(caliber_tab, text="Factores calibre")

        self._build_general_tab(general_tab)
        self._build_staff_tab(personal_tab)
        self._build_flow_staffing_tab(flow_staffing_tab)
        self._build_packaging_tab(packaging_tab)
        self._build_lines_tab(lines_tab)
        self._build_resources_flows_tab(resources_tab)
        self._build_capacity_productive_tab(capacity_tab)
        self._build_base_packaging_tab(base_packaging_tab)
        self._build_packaging_mapping_tab(mapping_tab)
        self._build_performance_tab(performance_tab)
        self._build_penalties_tab(penalties_tab)
        self._build_semaphore_tab(semaphore_tab)
        self._build_caliber_factors_tab(caliber_tab)

    # General tab (sin cambios funcionales)
    def _build_general_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        panel = ttk.LabelFrame(parent, text="Parámetros operativos diarios", padding=12)
        panel.grid(row=0, column=0, sticky="nsew")
        panel.grid_columnconfigure(1, weight=1)
        ttk.Button(panel, text="ⓘ Descripción de campos", command=self._show_general_day_help).grid(row=0, column=2, sticky="e", padx=(8, 0), pady=(0, 8))

        def add_entry(row: int, key: str, label: str, default: str = "") -> None:
            ttk.Label(panel, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
            var = tk.StringVar(value=default)
            ttk.Entry(panel, textvariable=var).grid(row=row, column=1, sticky="ew", pady=4)
            self._general_vars[key] = var

        def add_combo(row: int, key: str, label: str, values: list[str]) -> None:
            ttk.Label(panel, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
            var = tk.StringVar(value=values[0])
            ttk.Combobox(panel, textvariable=var, values=values, state="readonly").grid(row=row, column=1, sticky="ew", pady=4)
            self._general_vars[key] = var

        def add_check(row: int, key: str, label: str, default: int = 0) -> None:
            var = tk.IntVar(value=default)
            ttk.Checkbutton(panel, text=label, variable=var).grid(row=row, column=0, columnspan=2, sticky="w", pady=2)
            self._general_vars[key] = var

        def add_volcado_checks(row: int) -> None:
            ttk.Label(panel, text="Líneas de volcado activas").grid(row=row, column=0, sticky="nw", padx=(0, 8), pady=4)
            container = ttk.Frame(panel)
            container.grid(row=row, column=1, sticky="w", pady=4)
            for idx, option in enumerate(self.VOLCADO_OPTIONS):
                var = tk.IntVar(value=1 if option == "Compacta" else 0)
                self._tipos_volcado_vars[option] = var
                ttk.Checkbutton(container, text=option, variable=var).grid(row=idx, column=0, sticky="w", pady=1)

        add_entry(1, "horas_turno", "Horas por turno", "8")
        add_entry(2, "numero_turnos", "Número de turnos", "1")
        add_entry(3, "horas_descanso", "Horas de descanso", "0.5")
        add_combo(4, "tipo_campana", "Tipo de campaña", ["Baja actividad", "Normal", "Alta", "Pico campaña"])
        add_volcado_checks(5)
        add_entry(6, "saturacion_maxima_pct", "Saturación máxima %", "90")
        add_check(7, "permitir_horas_extra", "Permitir horas extra", 1)
        add_check(8, "permitir_segundo_turno", "Permitir segundo turno", 0)
        add_check(9, "priorizar_pedidos_reales", "Priorizar pedidos reales", 1)
        add_check(10, "permitir_adelantar_produccion", "Permitir adelantar producción", 1)
        add_check(11, "agrupar_pedidos_compatibles", "Agrupar pedidos compatibles", 1)
        add_check(12, "minimizar_cambios_formato", "Minimizar cambios de formato", 1)
        add_entry(13, "kg_objetivo_dia", "Kg objetivo día", "0")
        add_entry(14, "palets_objetivo_dia", "Palets objetivo día", "0")
        add_entry(15, "pedidos_maximos_recomendados", "Pedidos máximos recomendados", "0")

        calc = ttk.LabelFrame(parent, text="Campos calculados", padding=12)
        calc.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        calc.grid_columnconfigure(1, weight=1)
        self._general_vars["horas_brutas_dia"] = tk.StringVar(); self._general_vars["horas_utiles_dia"] = tk.StringVar(); self._general_vars["saturacion_util_objetivo"] = tk.StringVar()
        self._add_readonly(calc, 0, "horas_brutas_dia", "Horas brutas día", self._general_vars["horas_brutas_dia"])
        self._add_readonly(calc, 1, "horas_utiles_dia", "Horas útiles día", self._general_vars["horas_utiles_dia"])
        self._add_readonly(calc, 2, "saturacion_util_objetivo", "Saturación útil objetivo", self._general_vars["saturacion_util_objetivo"])
        btns = ttk.Frame(parent); btns.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Guardar configuración", command=self._save_general_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_defaults).pack(side="left", padx=4)
        ttk.Button(btns, text="Volver", command=self.on_back).pack(side="right", padx=4)
        for key in ("horas_turno", "numero_turnos", "horas_descanso", "saturacion_maxima_pct"):
            self._general_vars[key].trace_add("write", lambda *_: self._recalculate())

    def _build_staff_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_personal_help,
            export_command=lambda: self._export_master_excel("personal", self._load_personal_rows_for_excel),
            import_command=lambda: self._import_master_excel("personal", self._save_personal_rows_from_excel, self._load_staff_settings))

        summary = ttk.LabelFrame(parent, text="Resumen de plantilla diaria", padding=12)
        summary.grid(row=1, column=0, sticky="ew")
        summary.grid_columnconfigure(1, weight=1)
        fields = [("personal_total", "Personal disponible total", True), ("personal_directo", "Personal directo disponible", True), ("personal_soporte", "Personal soporte disponible", True), ("personal_indirecto", "Personal indirecto disponible", True), ("horas_por_persona", "Horas por persona", False), ("ausencias_previstas", "Ausencias previstas", False)]
        for i, (key, label, calculated) in enumerate(fields):
            ttk.Label(summary, text=label).grid(row=i, column=0, sticky="w", padx=(0, 8), pady=4)
            var = tk.StringVar(value="0")
            state = "readonly" if calculated else "normal"
            ttk.Entry(summary, textvariable=var, state=state).grid(row=i, column=1, sticky="ew", pady=4)
            self._staff_summary_vars[key] = var
        ttk.Label(summary, text="Observaciones").grid(row=6, column=0, sticky="nw", padx=(0, 8), pady=4)
        obs_var = tk.StringVar(value="")
        ttk.Entry(summary, textvariable=obs_var).grid(row=6, column=1, sticky="ew", pady=4)
        self._staff_summary_vars["observaciones"] = obs_var

        areas = ttk.LabelFrame(parent, text="Personal por área operativa", padding=12)
        areas.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        parent.grid_rowconfigure(2, weight=1)
        columns = ("id", "area", "tipo_personal", "disponible", "minimo_operativo", "optimo", "activo", "observaciones")
        tree = ttk.Treeview(areas, columns=columns, show="headings", height=10)
        self._staff_tree = tree
        headers = {"id": "ID", "area": "Área", "tipo_personal": "Tipo personal", "disponible": "Disponible", "minimo_operativo": "Mínimo operativo", "optimo": "Óptimo", "activo": "Activo", "observaciones": "Observaciones"}
        widths = {"id": 40, "area": 150, "tipo_personal": 110, "disponible": 85, "minimo_operativo": 110, "optimo": 75, "activo": 70, "observaciones": 260}
        for col in columns:
            tree.heading(col, text=headers[col]); tree.column(col, width=widths[col], anchor="w")
        yscroll = ttk.Scrollbar(areas, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(areas, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns"); xscroll.grid(row=1, column=0, sticky="ew")
        areas.grid_columnconfigure(0, weight=1); areas.grid_rowconfigure(0, weight=1)
        tree.bind("<<TreeviewSelect>>", self._on_staff_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar área seleccionada", padding=12)
        editor.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        editor.grid_columnconfigure(1, weight=1)
        self._staff_editor_vars = {"id": tk.StringVar(value=""), "area": tk.StringVar(value=""), "tipo_personal": tk.StringVar(value="Directo"), "disponible": tk.StringVar(value="0"), "minimo_operativo": tk.StringVar(value="0"), "optimo": tk.StringVar(value="0"), "activo": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor, text="Área").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(editor, textvariable=self._staff_editor_vars["area"]).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(editor, text="Tipo personal").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Combobox(editor, textvariable=self._staff_editor_vars["tipo_personal"], values=self.STAFF_TYPES, state="readonly").grid(row=1, column=1, sticky="ew", pady=4)
        for r, key, label in ((2, "disponible", "Disponible"), (3, "minimo_operativo", "Mínimo operativo"), (4, "optimo", "Óptimo")):
            ttk.Label(editor, text=label).grid(row=r, column=0, sticky="w", padx=(0, 8), pady=4)
            ttk.Entry(editor, textvariable=self._staff_editor_vars[key]).grid(row=r, column=1, sticky="ew", pady=4)
        ttk.Checkbutton(editor, text="Activo", variable=self._staff_editor_vars["activo"]).grid(row=5, column=1, sticky="w", pady=4)
        ttk.Label(editor, text="Observaciones").grid(row=6, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(editor, textvariable=self._staff_editor_vars["observaciones"]).grid(row=6, column=1, sticky="ew", pady=4)
        ttk.Button(editor, text="Aplicar cambios a la fila", command=self._apply_staff_row_changes).grid(row=7, column=1, sticky="e", pady=(8, 0))

        btns = ttk.Frame(parent); btns.grid(row=4, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Guardar personal", command=self._save_staff_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_staff_defaults).pack(side="left", padx=4)
        ttk.Button(btns, text="Añadir área", command=self._add_staff_area).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar área seleccionada", command=self._delete_staff_area).pack(side="left", padx=4)


    def _build_flow_staffing_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_flow_staffing_help,
            export_command=lambda: self._export_master_excel("flow_staffing", self.service.get_flow_staffing),
            import_command=lambda: self._import_master_excel("flow_staffing", self.service.save_flow_staffing, self._load_flow_staffing_settings))

        columns = ("id", "linea_productiva", "area_puesto", "tipo_personal", "minimo", "optimo", "escala_con_ocupacion", "factor_ocupacion", "obligatorio", "activo", "observaciones")
        headers = {"id": "ID", "linea_productiva": "Línea productiva", "area_puesto": "Área / puesto", "tipo_personal": "Tipo personal", "minimo": "Mínimo", "optimo": "Óptimo", "escala_con_ocupacion": "Escala con ocupación", "factor_ocupacion": "Factor ocupación", "obligatorio": "Obligatorio", "activo": "Activo", "observaciones": "Observaciones"}
        widths = {"id": 45, "linea_productiva": 150, "area_puesto": 170, "tipo_personal": 110, "minimo": 75, "optimo": 75, "escala_con_ocupacion": 145, "factor_ocupacion": 120, "obligatorio": 90, "activo": 70, "observaciones": 260}
        table_frame = ttk.LabelFrame(parent, text="Maestro de dotación por flujo", padding=12)
        table_frame.grid(row=1, column=0, sticky="nsew")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)
        self._flow_staffing_tree = tree
        for col in columns:
            tree.heading(col, text=headers[col]); tree.column(col, width=widths[col], anchor="w")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns"); xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.grid_columnconfigure(0, weight=1); table_frame.grid_rowconfigure(0, weight=1)
        tree.bind("<<TreeviewSelect>>", self._on_flow_staffing_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar dotación seleccionada", padding=12)
        editor.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        editor.grid_columnconfigure(1, weight=1); editor.grid_columnconfigure(3, weight=1)
        line_values = ["MALLAS_TRADICIONAL", "MALLAS_GIRSAC", "ENCAJADO", "GRANEL_MANUAL", "GRANELERA"]
        self._flow_staffing_editor_vars = {"id": tk.StringVar(value=""), "linea_productiva": tk.StringVar(value=line_values[0]), "area_puesto": tk.StringVar(value=""), "tipo_personal": tk.StringVar(value="Directo"), "minimo": tk.StringVar(value="0"), "optimo": tk.StringVar(value="0"), "escala_con_ocupacion": tk.IntVar(value=0), "factor_ocupacion": tk.StringVar(value="1.0"), "obligatorio": tk.IntVar(value=1), "activo": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor, text="Línea productiva").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Combobox(editor, textvariable=self._flow_staffing_editor_vars["linea_productiva"], values=line_values).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(editor, text="Área / puesto").grid(row=0, column=2, sticky="w", padx=(12, 8), pady=4)
        ttk.Entry(editor, textvariable=self._flow_staffing_editor_vars["area_puesto"]).grid(row=0, column=3, sticky="ew", pady=4)
        ttk.Label(editor, text="Tipo personal").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Combobox(editor, textvariable=self._flow_staffing_editor_vars["tipo_personal"], values=self.STAFF_TYPES, state="readonly").grid(row=1, column=1, sticky="ew", pady=4)
        for row, col, key, label in ((1, 2, "minimo", "Mínimo"), (2, 0, "optimo", "Óptimo"), (2, 2, "factor_ocupacion", "Factor ocupación")):
            ttk.Label(editor, text=label).grid(row=row, column=col, sticky="w", padx=(12 if col else 0, 8), pady=4)
            ttk.Entry(editor, textvariable=self._flow_staffing_editor_vars[key]).grid(row=row, column=col + 1, sticky="ew", pady=4)
        ttk.Checkbutton(editor, text="Escala con ocupación", variable=self._flow_staffing_editor_vars["escala_con_ocupacion"]).grid(row=3, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Checkbutton(editor, text="Obligatorio", variable=self._flow_staffing_editor_vars["obligatorio"]).grid(row=3, column=2, sticky="w", pady=2)
        ttk.Checkbutton(editor, text="Activo", variable=self._flow_staffing_editor_vars["activo"]).grid(row=3, column=3, sticky="w", pady=2)
        ttk.Label(editor, text="Observaciones").grid(row=4, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(editor, textvariable=self._flow_staffing_editor_vars["observaciones"]).grid(row=4, column=1, columnspan=3, sticky="ew", pady=4)

        btns = ttk.Frame(parent); btns.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Nueva dotación", command=self._add_flow_staffing_row).pack(side="left", padx=4)
        ttk.Button(btns, text="Aplicar cambios", command=self._apply_flow_staffing_row_changes).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar dotación", command=self._delete_flow_staffing_row).pack(side="left", padx=4)
        ttk.Button(btns, text="Guardar dotación", command=self._save_flow_staffing_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_flow_staffing_defaults).pack(side="left", padx=4)




    def _build_packaging_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_packaging_help,
            export_command=lambda: self._export_master_excel("packaging_types", self.service.get_packaging_types),
            import_command=lambda: self._import_master_excel("packaging_types", self.service.save_packaging_types, self._load_packaging_settings))

        catalog = ttk.LabelFrame(parent, text="Catálogo de confecciones", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew")
        catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols = ("id", "codigo", "descripcion", "familia", "subtipo", "kg_formato", "material", "tipo_malla", "requiere_precalibrado", "compatible_box", "activo", "observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=9)
        self._packaging_tree = tree
        for c,h,w in [("id","ID",40),("codigo","Código",150),("descripcion","Descripción",190),("familia","Familia",90),("subtipo","Subtipo",120),("kg_formato","Kg formato",80),("material","Material",90),("tipo_malla","Tipo malla",100),("requiere_precalibrado","Req. pre calib.",110),("compatible_box","Compatible BOX",110),("activo","Activo",60),("observaciones","Observaciones",220)]:
            tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        yscroll = ttk.Scrollbar(catalog, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns"); xscroll.grid(row=1, column=0, sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_packaging_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar confección seleccionada", padding=12)
        editor.grid(row=2, column=0, sticky="ew", pady=(10, 0)); editor.grid_columnconfigure(1, weight=1)
        self._packaging_editor_vars = {"id": tk.StringVar(value=""), "codigo": tk.StringVar(value=""), "descripcion": tk.StringVar(value=""), "familia": tk.StringVar(value=self.PACKAGING_FAMILIES[0]), "subtipo": tk.StringVar(value=self.PACKAGING_SUBTYPES[0]), "kg_formato": tk.StringVar(value="0"), "material": tk.StringVar(value=self.PACKAGING_MATERIALS[0]), "tipo_malla": tk.StringVar(value=self.PACKAGING_MESH_TYPES[0]), "requiere_precalibrado": tk.IntVar(value=0), "compatible_box": tk.IntVar(value=0), "activo": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor, text="Código").grid(row=0,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._packaging_editor_vars["codigo"]).grid(row=0,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Descripción").grid(row=1,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._packaging_editor_vars["descripcion"]).grid(row=1,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Familia").grid(row=2,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._packaging_editor_vars["familia"],values=self.PACKAGING_FAMILIES,state="readonly").grid(row=2,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Subtipo").grid(row=3,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._packaging_editor_vars["subtipo"],values=self.PACKAGING_SUBTYPES).grid(row=3,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Kg formato").grid(row=4,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._packaging_editor_vars["kg_formato"]).grid(row=4,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Material").grid(row=5,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._packaging_editor_vars["material"],values=self.PACKAGING_MATERIALS).grid(row=5,column=1,sticky="ew",pady=4)
        ttk.Label(editor, text="Tipo malla").grid(row=6,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._packaging_editor_vars["tipo_malla"],values=self.PACKAGING_MESH_TYPES).grid(row=6,column=1,sticky="ew",pady=4)
        ttk.Checkbutton(editor,text="Requiere pre calibrado",variable=self._packaging_editor_vars["requiere_precalibrado"]).grid(row=7,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor,text="Compatible BOX",variable=self._packaging_editor_vars["compatible_box"]).grid(row=8,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor,text="Activo",variable=self._packaging_editor_vars["activo"]).grid(row=9,column=1,sticky="w",pady=2)
        ttk.Label(editor, text="Observaciones").grid(row=10,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._packaging_editor_vars["observaciones"]).grid(row=10,column=1,sticky="ew",pady=4)

        btns = ttk.Frame(parent); btns.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Nueva confección", command=self._add_packaging_row).pack(side="left", padx=4)
        ttk.Button(btns, text="Aplicar cambios a selección", command=self._apply_packaging_row_changes).pack(side="left", padx=4)
        ttk.Button(btns, text="Guardar confecciones", command=self._save_packaging_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_packaging_defaults).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar confección seleccionada", command=self._delete_packaging_row).pack(side="left", padx=4)
    def _add_readonly(self, parent: ttk.Frame, row: int, key: str, label: str, variable: tk.StringVar) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(parent, textvariable=variable, state="readonly").grid(row=row, column=1, sticky="ew", pady=4)

    def _show_general_day_help(self) -> None:
        keys = ["horas_turno", "numero_turnos", "horas_descanso", "tipo_campana", "tipos_volcado_activos", "saturacion_maxima_pct", "permitir_horas_extra", "permitir_segundo_turno", "priorizar_pedidos_reales", "permitir_adelantar_produccion", "agrupar_pedidos_compatibles", "minimizar_cambios_formato", "kg_objetivo_dia", "palets_objetivo_dia", "pedidos_maximos_recomendados", "horas_brutas_dia", "horas_utiles_dia", "saturacion_util_objetivo"]
        show_tab_help(self, title="Descripción de campos - General del día", intro="Esta pestaña define los parámetros básicos de trabajo que se usarán para calcular la capacidad productiva diaria del almacén.", help_items=[PRODUCTION_FIELD_HELP[k] for k in keys if k in PRODUCTION_FIELD_HELP], heading="General del día")

    def _show_personal_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Personal", heading="Personal", intro="Esta pestaña define la plantilla disponible y los mínimos operativos por área de trabajo. Estos datos se usarán posteriormente para estimar si los pedidos previstos pueden producirse con el personal disponible.", help_items=get_help_items(PRODUCTION_PERSONAL_HELP_KEYS, PRODUCTION_PERSONAL_HELP))


    def _show_flow_staffing_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Dotación por flujo", heading="Dotación por flujo", intro="Este maestro define el equipo completo requerido por línea productiva y puesto operativo para calcular el personal total necesario del flujo.", help_items=get_help_items(PRODUCTION_FLOW_STAFFING_HELP_KEYS, PRODUCTION_FLOW_STAFFING_HELP))

    def _load_flow_staffing_settings(self) -> None:
        self._refresh_flow_staffing_tree(self.service.get_flow_staffing())

    def _refresh_flow_staffing_tree(self, rows: list[dict]) -> None:
        if not self._flow_staffing_tree:
            return
        self._flow_staffing_tree.delete(*self._flow_staffing_tree.get_children())
        for row in rows:
            self._flow_staffing_tree.insert("", "end", values=(row.get("id", ""), row.get("linea_productiva", ""), row.get("area_puesto", ""), row.get("tipo_personal", ""), row.get("minimo", 0), row.get("optimo", 0), row.get("escala_con_ocupacion", 0), row.get("factor_ocupacion", 1.0), row.get("obligatorio", 1), row.get("activo", 1), row.get("observaciones", "")))

    def _on_flow_staffing_row_selected(self, _event=None) -> None:
        if not self._flow_staffing_tree:
            return
        selected = self._flow_staffing_tree.selection()
        if not selected:
            return
        vals = self._flow_staffing_tree.item(selected[0], "values")
        keys = ("id", "linea_productiva", "area_puesto", "tipo_personal", "minimo", "optimo", "escala_con_ocupacion", "factor_ocupacion", "obligatorio", "activo", "observaciones")
        for i, key in enumerate(keys):
            self._flow_staffing_editor_vars[key].set(vals[i])

    def _validate_flow_staffing_values(self, values: tuple) -> tuple:
        linea = str(values[1]).strip(); area = str(values[2]).strip(); tipo = str(values[3]).strip()
        if not linea: raise ValueError("La línea productiva es obligatoria")
        if not area: raise ValueError("El área / puesto es obligatorio")
        if not tipo: raise ValueError("El tipo de personal es obligatorio")
        minimo = int(str(values[4]).strip()); optimo = int(str(values[5]).strip()); factor = float(str(values[7]).strip().replace(",", "."))
        if minimo < 0 or optimo < 0: raise ValueError("Mínimo y óptimo deben ser >= 0")
        if optimo < minimo: raise ValueError("Óptimo no puede ser menor que mínimo")
        if factor <= 0: raise ValueError("Factor ocupación debe ser > 0")
        return (values[0], linea, area, tipo, minimo, optimo, 1 if str(values[6]).strip() in ("1", "True", "true") else 0, factor, 1 if str(values[8]).strip() in ("1", "True", "true") else 0, 1 if str(values[9]).strip() in ("1", "True", "true") else 0, str(values[10]).strip())

    def _apply_flow_staffing_row_changes(self) -> None:
        if not self._flow_staffing_tree: return
        selected = self._flow_staffing_tree.selection()
        if not selected:
            messagebox.showerror("Dotación flujos", "Seleccione una fila para editar.", parent=self); return
        try:
            vals = self._validate_flow_staffing_values(tuple(self._flow_staffing_editor_vars[k].get() for k in ("id", "linea_productiva", "area_puesto", "tipo_personal", "minimo", "optimo", "escala_con_ocupacion", "factor_ocupacion", "obligatorio", "activo", "observaciones")))
            self._flow_staffing_tree.item(selected[0], values=vals)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _add_flow_staffing_row(self) -> None:
        if self._flow_staffing_tree:
            self._flow_staffing_tree.insert("", "end", values=("", "MALLAS_TRADICIONAL", "Nuevo puesto", "Directo", 1, 1, 0, 1.0, 1, 1, ""))

    def _delete_flow_staffing_row(self) -> None:
        if not self._flow_staffing_tree: return
        selected = self._flow_staffing_tree.selection()
        if not selected:
            messagebox.showerror("Dotación flujos", "Seleccione una fila para eliminar.", parent=self); return
        for item in selected:
            self._flow_staffing_tree.delete(item)

    def _collect_flow_staffing_rows_payload(self) -> list[dict]:
        rows=[]
        if not self._flow_staffing_tree:
            return rows
        for item in self._flow_staffing_tree.get_children():
            vals = self._validate_flow_staffing_values(self._flow_staffing_tree.item(item, "values"))
            rows.append({"id": vals[0], "linea_productiva": vals[1], "area_puesto": vals[2], "tipo_personal": vals[3], "minimo": vals[4], "optimo": vals[5], "escala_con_ocupacion": vals[6], "factor_ocupacion": vals[7], "obligatorio": vals[8], "activo": vals[9], "observaciones": vals[10]})
        return rows

    def _save_flow_staffing_settings(self) -> None:
        try:
            self.service.save_flow_staffing(self._collect_flow_staffing_rows_payload())
            self._load_flow_staffing_settings()
            messagebox.showinfo("Configuración productiva", "Dotación por flujo guardada correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_flow_staffing_defaults(self) -> None:
        self.service.reset_flow_staffing_defaults()
        self._load_flow_staffing_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de dotación por flujo restaurados.", parent=self)

    def _load_general_settings(self) -> None:
        data = self.service.get_general_settings()
        for key, var in self._general_vars.items():
            if key in data:
                var.set(int(data[key]) if isinstance(var, tk.IntVar) else str(data[key]))
        activos = set(data.get("tipos_volcado_activos", []))
        for option, var in self._tipos_volcado_vars.items():
            var.set(1 if option in activos else 0)
        self._recalculate()

    def _load_staff_settings(self) -> None:
        summary = self.service.get_staff_summary()
        for key, var in self._staff_summary_vars.items():
            var.set(str(summary.get(key, "")))
        self._refresh_staff_tree(self.service.get_staff_areas())
        self._refresh_staff_summary_totals()

    def _refresh_staff_tree(self, rows: list[dict]) -> None:
        if not self._staff_tree:
            return
        self._staff_tree.delete(*self._staff_tree.get_children())
        for row in rows:
            self._staff_tree.insert("", "end", values=(row.get("id", ""), row.get("area", ""), row.get("tipo_personal", ""), row.get("disponible", 0), row.get("minimo_operativo", 0), row.get("optimo", 0), row.get("activo", 0), row.get("observaciones", "")))

    def _parse_float(self, key: str, label: str) -> float:
        raw = str(self._general_vars[key].get()).strip().replace(",", ".")
        if raw == "": raise ValueError(f"{label}: valor obligatorio")
        return float(raw)

    def _parse_int(self, key: str, label: str) -> int:
        raw = str(self._general_vars[key].get()).strip()
        if raw == "": raise ValueError(f"{label}: valor obligatorio")
        return int(raw)

    def _parse_non_negative_int(self, value: str, label: str) -> int:
        number = int(str(value).strip())
        if number < 0: raise ValueError(f"{label} debe ser >= 0")
        return number

    def _collect_payload(self) -> dict:
        horas_turno = self._parse_float("horas_turno", "Horas por turno"); numero_turnos = self._parse_int("numero_turnos", "Número de turnos"); horas_descanso = self._parse_float("horas_descanso", "Horas de descanso"); saturacion = self._parse_float("saturacion_maxima_pct", "Saturación máxima %")
        if horas_turno < 0 or numero_turnos < 0 or horas_descanso < 0: raise ValueError("Horas por turno, número de turnos y horas de descanso deben ser >= 0")
        if saturacion < 0 or saturacion > 100: raise ValueError("Saturación máxima % debe estar entre 0 y 100")
        tipos_volcado_activos = [option for option, var in self._tipos_volcado_vars.items() if int(var.get()) == 1]
        if not tipos_volcado_activos: raise ValueError("Debe activar al menos una línea de volcado")
        return {"horas_turno": horas_turno, "numero_turnos": numero_turnos, "horas_descanso": horas_descanso, "tipo_campana": self._general_vars["tipo_campana"].get(), "tipos_volcado_activos": tipos_volcado_activos, "saturacion_maxima_pct": saturacion, "permitir_horas_extra": int(self._general_vars["permitir_horas_extra"].get()), "permitir_segundo_turno": int(self._general_vars["permitir_segundo_turno"].get()), "priorizar_pedidos_reales": int(self._general_vars["priorizar_pedidos_reales"].get()), "permitir_adelantar_produccion": int(self._general_vars["permitir_adelantar_produccion"].get()), "agrupar_pedidos_compatibles": int(self._general_vars["agrupar_pedidos_compatibles"].get()), "minimizar_cambios_formato": int(self._general_vars["minimizar_cambios_formato"].get()), "kg_objetivo_dia": self._parse_float("kg_objetivo_dia", "Kg objetivo día"), "palets_objetivo_dia": self._parse_float("palets_objetivo_dia", "Palets objetivo día"), "pedidos_maximos_recomendados": self._parse_int("pedidos_maximos_recomendados", "Pedidos máximos recomendados")}

    def _collect_staff_summary_payload(self) -> dict:
        calculated = self._calculate_staff_summary_from_tree()
        return {
            **calculated,
            "horas_por_persona": float(str(self._staff_summary_vars["horas_por_persona"].get()).replace(",", ".")),
            "ausencias_previstas": self._parse_non_negative_int(self._staff_summary_vars["ausencias_previstas"].get(), "Ausencias previstas"),
            "observaciones": str(self._staff_summary_vars["observaciones"].get()).strip(),
        }

    def _collect_staff_rows_payload(self) -> list[dict]:
        if not self._staff_tree: return []
        rows: list[dict] = []
        areas: set[str] = set()
        for item_id in self._staff_tree.get_children():
            values = self._staff_tree.item(item_id, "values")
            area = str(values[1]).strip()
            if not area: raise ValueError("El área no puede estar vacía")
            if area.lower() in areas: raise ValueError(f"Área duplicada: {area}")
            areas.add(area.lower())
            tipo_personal = self._normalize_staff_type(values[2])
            rows.append({"area": area, "tipo_personal": tipo_personal, "disponible": self._parse_non_negative_int(values[3], f"Disponible ({area})"), "minimo_operativo": self._parse_non_negative_int(values[4], f"Mínimo operativo ({area})"), "optimo": self._parse_non_negative_int(values[5], f"Óptimo ({area})"), "activo": 1 if str(values[6]).strip() in ("1", "True", "true", "Sí", "si") else 0, "observaciones": str(values[7]).strip()})
        return rows

    def _normalize_staff_type(self, value) -> str:
        text = str(value or "").strip().lower()
        if text == "directo":
            return "Directo"
        if text == "soporte":
            return "Soporte"
        if text == "indirecto":
            return "Indirecto"
        raise ValueError("Tipo personal debe ser Directo, Soporte o Indirecto")

    def _calculate_staff_summary_from_tree(self) -> dict:
        totals = {"personal_directo": 0, "personal_soporte": 0, "personal_indirecto": 0}
        if self._staff_tree:
            for item_id in self._staff_tree.get_children():
                values = self._staff_tree.item(item_id, "values")
                if str(values[6]).strip() not in ("1", "True", "true", "Sí", "si"):
                    continue
                tipo = self._normalize_staff_type(values[2])
                disponible = self._parse_non_negative_int(values[3], "Disponible")
                if tipo == "Directo":
                    totals["personal_directo"] += disponible
                elif tipo == "Soporte":
                    totals["personal_soporte"] += disponible
                else:
                    totals["personal_indirecto"] += disponible
        totals["personal_total"] = totals["personal_directo"] + totals["personal_soporte"] + totals["personal_indirecto"]
        return totals

    def _refresh_staff_summary_totals(self) -> None:
        for key, value in self._calculate_staff_summary_from_tree().items():
            if key in self._staff_summary_vars:
                self._staff_summary_vars[key].set(str(value))

    def _recalculate(self) -> None:
        try:
            horas_brutas = self._parse_float("horas_turno", "Horas por turno") * self._parse_int("numero_turnos", "Número de turnos")
            horas_utiles = horas_brutas - self._parse_float("horas_descanso", "Horas de descanso")
            saturacion_util = horas_utiles * self._parse_float("saturacion_maxima_pct", "Saturación máxima %") / 100.0
            self._general_vars["horas_brutas_dia"].set(f"{horas_brutas:.2f}"); self._general_vars["horas_utiles_dia"].set(f"{horas_utiles:.2f}"); self._general_vars["saturacion_util_objetivo"].set(f"{saturacion_util:.2f}")
        except Exception:
            self._general_vars["horas_brutas_dia"].set("-"); self._general_vars["horas_utiles_dia"].set("-"); self._general_vars["saturacion_util_objetivo"].set("-")

    def _save_general_settings(self) -> None:
        try:
            self.service.save_general_settings(self._collect_payload()); self._recalculate(); messagebox.showinfo("Configuración productiva", "Configuración guardada correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_defaults(self) -> None:
        self.service.reset_general_defaults(); self._load_general_settings(); messagebox.showinfo("Configuración productiva", "Valores por defecto restaurados.", parent=self)

    def _on_staff_row_selected(self, _event=None) -> None:
        if not self._staff_tree: return
        selected = self._staff_tree.selection()
        if not selected: return
        values = self._staff_tree.item(selected[0], "values")
        for idx, key in enumerate(("id", "area", "tipo_personal", "disponible", "minimo_operativo", "optimo", "activo", "observaciones")):
            self._staff_editor_vars[key].set(values[idx])

    def _apply_staff_row_changes(self) -> None:
        if not self._staff_tree: return
        selected = self._staff_tree.selection()
        if not selected:
            messagebox.showerror("Personal", "Seleccione una fila para editar.", parent=self); return
        try:
            area = str(self._staff_editor_vars["area"].get()).strip()
            if not area: raise ValueError("El área es obligatoria")
            values = (self._staff_editor_vars["id"].get(), area, self._staff_editor_vars["tipo_personal"].get(), self._parse_non_negative_int(self._staff_editor_vars["disponible"].get(), "Disponible"), self._parse_non_negative_int(self._staff_editor_vars["minimo_operativo"].get(), "Mínimo operativo"), self._parse_non_negative_int(self._staff_editor_vars["optimo"].get(), "Óptimo"), 1 if int(self._staff_editor_vars["activo"].get()) else 0, str(self._staff_editor_vars["observaciones"].get()).strip())
            self._staff_tree.item(selected[0], values=values)
            self._refresh_staff_summary_totals()
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _add_staff_area(self) -> None:
        if not self._staff_tree: return
        self._staff_tree.insert("", "end", values=("", "Nueva área", "Directo", 0, 0, 0, 1, ""))
        self._refresh_staff_summary_totals()

    def _delete_staff_area(self) -> None:
        if not self._staff_tree: return
        selected = self._staff_tree.selection()
        if not selected:
            messagebox.showerror("Personal", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar el área seleccionada?", parent=self):
            self._staff_tree.delete(selected[0])
            self._refresh_staff_summary_totals()

    def _save_staff_settings(self) -> None:
        try:
            summary = self._collect_staff_summary_payload()
            if summary["horas_por_persona"] < 0: raise ValueError("Horas por persona debe ser >= 0")
            rows = self._collect_staff_rows_payload()
            self.service.save_staff_areas(rows)
            self.service.save_staff_summary(summary)
            self._load_staff_settings()
            messagebox.showinfo("Configuración productiva", "Personal guardado correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_staff_defaults(self) -> None:
        self.service.reset_staff_defaults()
        self._load_staff_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de personal restaurados.", parent=self)


    def _show_packaging_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Confecciones", heading="Confecciones", intro="Esta pestaña define el catálogo de confecciones que podrá usar la planificación productiva. Cada confección representa una forma concreta de preparar la fruta: malla, encajado, granel, granelera u otros formatos.", help_items=get_help_items(PRODUCTION_PACKAGING_HELP_KEYS, PRODUCTION_PACKAGING_HELP))

    def _show_base_packaging_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Confecciones base", heading="Confecciones base", intro="Esta pestaña define el catálogo maestro de confecciones base y su interpretación productiva de referencia. Estos datos sirven como base para mapear y estandarizar confecciones comerciales.", help_items=get_help_items(PRODUCTION_PACKAGING_HELP_KEYS, PRODUCTION_PACKAGING_HELP))

    def _show_lines_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Máquinas / líneas", heading="Máquinas / líneas", intro="Esta pestaña define los recursos físicos y operativos disponibles en el almacén: líneas de volcado, máquinas de malla, encajado, granelera, calibrador y puntos de apoyo.", help_items=get_help_items(PRODUCTION_LINES_HELP_KEYS, PRODUCTION_LINES_HELP))

    def _load_lines_settings(self) -> None:
        self._refresh_lines_tree(self.service.get_lines())

    def _refresh_lines_tree(self, rows: list[dict]) -> None:
        if not self._lines_tree: return
        self._lines_tree.delete(*self._lines_tree.get_children())
        for row in rows:
            self._lines_tree.insert("", "end", values=(row.get("id",""), row.get("codigo",""), row.get("nombre",""), row.get("tipo_linea",""), row.get("familia_principal",""), row.get("numero_maquinas",0), row.get("activa",1), row.get("capacidad_kg_h_referencia",0), row.get("personal_minimo",0), row.get("personal_optimo",0), row.get("permite_precalibrado",0), row.get("permite_box",0), row.get("observaciones","")))

    def _on_line_row_selected(self, _event=None) -> None:
        if not self._lines_tree: return
        selected = self._lines_tree.selection()
        if not selected: return
        vals = self._lines_tree.item(selected[0], "values")
        for i, key in enumerate(("id", "codigo", "nombre", "tipo_linea", "familia_principal", "numero_maquinas", "activa", "capacidad_kg_h_referencia", "personal_minimo", "personal_optimo", "permite_precalibrado", "permite_box", "observaciones")):
            self._lines_editor_vars[key].set(vals[i])

    def _validate_line_values(self, values: tuple) -> tuple:
        codigo = str(values[1]).strip(); nombre = str(values[2]).strip()
        if not codigo: raise ValueError("El código es obligatorio")
        if not nombre: raise ValueError("El nombre es obligatorio")
        numero_maquinas = int(str(values[5]).strip()); capacidad = float(str(values[7]).strip().replace(",", ".")); pmin = int(str(values[8]).strip()); popt = int(str(values[9]).strip())
        if numero_maquinas < 0: raise ValueError("Nº máquinas debe ser >= 0")
        if capacidad < 0: raise ValueError("Capacidad kg/h referencia debe ser >= 0")
        if pmin < 0: raise ValueError("Personal mínimo debe ser >= 0")
        if popt < 0: raise ValueError("Personal óptimo debe ser >= 0")
        if popt < pmin: raise ValueError("Personal óptimo no debe ser menor que personal mínimo")
        return (values[0], codigo, nombre, str(values[3]).strip(), str(values[4]).strip(), numero_maquinas, 1 if str(values[6]).strip() in ("1","True","true") else 0, capacidad, pmin, popt, 1 if str(values[10]).strip() in ("1","True","true") else 0, 1 if str(values[11]).strip() in ("1","True","true") else 0, str(values[12]).strip())

    def _apply_line_row_changes(self) -> None:
        if not self._lines_tree: return
        selected = self._lines_tree.selection()
        if not selected: messagebox.showerror("Máquinas / líneas", "Seleccione una fila para editar.", parent=self); return
        try:
            values = self._validate_line_values((self._lines_editor_vars["id"].get(), self._lines_editor_vars["codigo"].get(), self._lines_editor_vars["nombre"].get(), self._lines_editor_vars["tipo_linea"].get(), self._lines_editor_vars["familia_principal"].get(), self._lines_editor_vars["numero_maquinas"].get(), self._lines_editor_vars["activa"].get(), self._lines_editor_vars["capacidad_kg_h_referencia"].get(), self._lines_editor_vars["personal_minimo"].get(), self._lines_editor_vars["personal_optimo"].get(), self._lines_editor_vars["permite_precalibrado"].get(), self._lines_editor_vars["permite_box"].get(), self._lines_editor_vars["observaciones"].get()))
            self._lines_tree.item(selected[0], values=values)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _add_line_row(self) -> None:
        if not self._lines_tree: return
        code = f"NUEVA_LINEA_{self._new_line_counter}"; self._new_line_counter += 1
        self._lines_tree.insert("", "end", values=("", code, "Nueva línea", "Otro", "Apoyo", 0, 1, 0.0, 0, 0, 0, 0, ""))

    def _delete_line_row(self) -> None:
        if not self._lines_tree: return
        selected = self._lines_tree.selection()
        if not selected: messagebox.showerror("Máquinas / líneas", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar la línea seleccionada?", parent=self):
            self._lines_tree.delete(selected[0])

    def _collect_lines_rows_payload(self) -> list[dict]:
        if not self._lines_tree: return []
        rows = []; codes = set()
        for item_id in self._lines_tree.get_children():
            clean = self._validate_line_values(self._lines_tree.item(item_id, "values"))
            code_key = clean[1].lower()
            if code_key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(code_key)
            rows.append({"codigo": clean[1], "nombre": clean[2], "tipo_linea": clean[3], "familia_principal": clean[4], "numero_maquinas": clean[5], "activa": clean[6], "capacidad_kg_h_referencia": clean[7], "personal_minimo": clean[8], "personal_optimo": clean[9], "permite_precalibrado": clean[10], "permite_box": clean[11], "observaciones": clean[12]})
        return rows

    def _save_lines_settings(self) -> None:
        try:
            self.service.save_lines(self._collect_lines_rows_payload())
            self._load_lines_settings()
            messagebox.showinfo("Configuración productiva", "Líneas guardadas correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_lines_defaults(self) -> None:
        self.service.reset_lines_defaults(); self._load_lines_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de líneas restaurados.", parent=self)

    def _load_packaging_settings(self) -> None:
        self._refresh_packaging_tree(self.service.get_packaging_types())

    def _refresh_packaging_tree(self, rows: list[dict]) -> None:
        if not self._packaging_tree:
            return
        self._packaging_tree.delete(*self._packaging_tree.get_children())
        for row in rows:
            self._packaging_tree.insert("", "end", values=(row.get("id", ""), row.get("codigo", ""), row.get("descripcion", ""), row.get("familia", ""), row.get("subtipo", ""), row.get("kg_formato", 0), row.get("material", ""), row.get("tipo_malla", ""), row.get("requiere_precalibrado", 0), row.get("compatible_box", 0), row.get("activo", 1), row.get("observaciones", "")))

    def _on_packaging_row_selected(self, _event=None) -> None:
        if not self._packaging_tree: return
        selected = self._packaging_tree.selection()
        if not selected: return
        vals = self._packaging_tree.item(selected[0], "values")
        for i, key in enumerate(("id", "codigo", "descripcion", "familia", "subtipo", "kg_formato", "material", "tipo_malla", "requiere_precalibrado", "compatible_box", "activo", "observaciones")):
            self._packaging_editor_vars[key].set(vals[i])

    def _validate_packaging_values(self, values: tuple) -> tuple:
        codigo = str(values[1]).strip(); descripcion = str(values[2]).strip(); familia = str(values[3]).strip()
        if not codigo: raise ValueError("El código es obligatorio")
        if not descripcion: raise ValueError("La descripción es obligatoria")
        if not familia: raise ValueError("La familia es obligatoria")
        kg = float(str(values[5]).replace(",", "."))
        if kg < 0: raise ValueError("Kg formato debe ser >= 0")
        return (values[0], codigo, descripcion, familia, str(values[4]).strip(), kg, str(values[6]).strip(), str(values[7]).strip(), 1 if str(values[8]).strip() in ("1", "True", "true") else 0, 1 if str(values[9]).strip() in ("1", "True", "true") else 0, 1 if str(values[10]).strip() in ("1", "True", "true") else 0, str(values[11]).strip())

    def _apply_packaging_row_changes(self) -> None:
        if not self._packaging_tree: return
        selected = self._packaging_tree.selection()
        if not selected:
            messagebox.showerror("Confecciones", "Seleccione una fila para editar.", parent=self); return
        try:
            values = self._validate_packaging_values((self._packaging_editor_vars["id"].get(), self._packaging_editor_vars["codigo"].get(), self._packaging_editor_vars["descripcion"].get(), self._packaging_editor_vars["familia"].get(), self._packaging_editor_vars["subtipo"].get(), self._packaging_editor_vars["kg_formato"].get(), self._packaging_editor_vars["material"].get(), self._packaging_editor_vars["tipo_malla"].get(), self._packaging_editor_vars["requiere_precalibrado"].get(), self._packaging_editor_vars["compatible_box"].get(), self._packaging_editor_vars["activo"].get(), self._packaging_editor_vars["observaciones"].get()))
            self._packaging_tree.item(selected[0], values=values)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _add_packaging_row(self) -> None:
        if not self._packaging_tree: return
        code = f"NUEVA_{self._new_packaging_counter}"
        self._new_packaging_counter += 1
        self._packaging_tree.insert("", "end", values=("", code, "Nueva confección", "Otro", "Otro", 0, "Otro", "No aplica", 0, 0, 1, ""))

    def _delete_packaging_row(self) -> None:
        if not self._packaging_tree: return
        selected = self._packaging_tree.selection()
        if not selected:
            messagebox.showerror("Confecciones", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar la confección seleccionada?", parent=self):
            self._packaging_tree.delete(selected[0])

    def _collect_packaging_rows_payload(self) -> list[dict]:
        if not self._packaging_tree: return []
        rows = []; codes = set()
        for item_id in self._packaging_tree.get_children():
            clean = self._validate_packaging_values(self._packaging_tree.item(item_id, "values"))
            code_key = clean[1].lower()
            if code_key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(code_key)
            rows.append({"codigo": clean[1], "descripcion": clean[2], "familia": clean[3], "subtipo": clean[4], "kg_formato": clean[5], "material": clean[6], "tipo_malla": clean[7], "requiere_precalibrado": clean[8], "compatible_box": clean[9], "activo": clean[10], "observaciones": clean[11]})
        return rows

    def _save_packaging_settings(self) -> None:
        try:
            rows = self._collect_packaging_rows_payload()
            self.service.save_packaging_types(rows)
            self._load_packaging_settings()
            messagebox.showinfo("Configuración productiva", "Confecciones guardadas correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_packaging_defaults(self) -> None:
        self.service.reset_packaging_defaults()
        self._load_packaging_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de confecciones restaurados.", parent=self)



    def _build_base_packaging_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent,
            help_command=self._show_base_packaging_help,
            export_command=lambda: self._export_master_excel("base_packaging", lambda: self.service.get_base_packaging(active_only=False)),
            import_command=lambda: self._import_master_excel("base_packaging", self.service.save_base_packaging, self._load_base_packaging_settings))

        catalog = ttk.LabelFrame(parent, text="Catálogo de confecciones base", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew")
        catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols = ("id", "codigo", "descripcion", "grupo_confeccion", "perfil_confeccion", "familia_productiva", "subtipo_productivo", "kg_formato", "tipo_malla", "linea_productiva", "requiere_precalibrado", "compatible_box", "activo", "observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=9)
        self._base_packaging_tree = tree
        headers = [("id", "ID", 40), ("codigo", "Código", 160), ("descripcion", "Descripción", 180), ("grupo_confeccion", "Grupo confección", 120), ("perfil_confeccion", "Perfil confección", 120), ("familia_productiva", "Familia productiva", 120), ("subtipo_productivo", "Subtipo productivo", 130), ("kg_formato", "Kg formato", 80), ("tipo_malla", "Tipo malla", 100), ("linea_productiva", "Línea productiva", 140), ("requiere_precalibrado", "Req. pre calib.", 110), ("compatible_box", "Compatible BOX", 110), ("activo", "Activo", 60), ("observaciones", "Observaciones", 180)]
        for c,h,w in headers:
            tree.heading(c, text=h); tree.column(c, width=w, anchor="w")
        yscroll = ttk.Scrollbar(catalog, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns"); xscroll.grid(row=1, column=0, sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_base_packaging_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar confección base seleccionada", padding=12)
        editor.grid(row=2, column=0, sticky="ew", pady=(10, 0)); editor.grid_columnconfigure(1, weight=1)
        self._base_packaging_editor_vars = {k: tk.StringVar(value="") for k in ("id", "codigo", "descripcion", "grupo_confeccion", "perfil_confeccion", "familia_productiva", "subtipo_productivo", "kg_formato", "tipo_malla", "linea_productiva", "observaciones")}
        self._base_packaging_editor_vars["requiere_precalibrado"] = tk.IntVar(value=0)
        self._base_packaging_editor_vars["compatible_box"] = tk.IntVar(value=0)
        self._base_packaging_editor_vars["activo"] = tk.IntVar(value=1)
        for r,(k,l) in enumerate((("codigo","Código"),("descripcion","Descripción"),("grupo_confeccion","Grupo confección"),("perfil_confeccion","Perfil confección"),("familia_productiva","Familia productiva"),("subtipo_productivo","Subtipo productivo"),("kg_formato","Kg formato"),("tipo_malla","Tipo malla"),("linea_productiva","Línea productiva"),("observaciones","Observaciones"))):
            ttk.Label(editor, text=l).grid(row=r, column=0, sticky="w", padx=(0,8), pady=4)
            ttk.Entry(editor, textvariable=self._base_packaging_editor_vars[k]).grid(row=r, column=1, sticky="ew", pady=4)
        ttk.Checkbutton(editor, text="Requiere pre calibrado", variable=self._base_packaging_editor_vars["requiere_precalibrado"]).grid(row=10,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor, text="Compatible BOX", variable=self._base_packaging_editor_vars["compatible_box"]).grid(row=11,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor, text="Activo", variable=self._base_packaging_editor_vars["activo"]).grid(row=12,column=1,sticky="w",pady=2)

        btns = ttk.Frame(parent); btns.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Nueva confección base", command=self._add_base_packaging_row).pack(side="left", padx=4)
        ttk.Button(btns, text="Aplicar cambios a selección", command=self._apply_base_packaging_row_changes).pack(side="left", padx=4)
        ttk.Button(btns, text="Guardar confecciones base", command=self._save_base_packaging_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_base_packaging_defaults).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar confección base seleccionada", command=self._delete_base_packaging_row).pack(side="left", padx=4)



    def _load_base_packaging_settings(self) -> None:
        self._refresh_base_packaging_tree(self.service.get_base_packaging(active_only=False))

    def _refresh_base_packaging_tree(self, rows: list[dict]) -> None:
        if not self._base_packaging_tree:
            return
        self._base_packaging_tree.delete(*self._base_packaging_tree.get_children())
        for row in rows:
            self._base_packaging_tree.insert("", "end", values=(row.get("id", ""), row.get("codigo", ""), row.get("descripcion", ""), row.get("grupo_confeccion", ""), row.get("perfil_confeccion", ""), row.get("familia_productiva", ""), row.get("subtipo_productivo", ""), row.get("kg_formato", 0), row.get("tipo_malla", ""), row.get("linea_productiva", ""), row.get("requiere_precalibrado", 0), row.get("compatible_box", 0), row.get("activo", 1), row.get("observaciones", "")))

    def _on_base_packaging_row_selected(self, _event=None) -> None:
        if not self._base_packaging_tree: return
        selected = self._base_packaging_tree.selection()
        if not selected: return
        vals = self._base_packaging_tree.item(selected[0], "values")
        for i, key in enumerate(("id", "codigo", "descripcion", "grupo_confeccion", "perfil_confeccion", "familia_productiva", "subtipo_productivo", "kg_formato", "tipo_malla", "linea_productiva", "requiere_precalibrado", "compatible_box", "activo", "observaciones")):
            self._base_packaging_editor_vars[key].set(vals[i])

    def _validate_base_packaging_values(self, values: tuple) -> tuple:
        codigo = str(values[1]).strip(); descripcion = str(values[2]).strip()
        if not codigo: raise ValueError("El código es obligatorio")
        if not descripcion: raise ValueError("La descripción es obligatoria")
        kg = float(str(values[7]).replace(",", "."))
        if kg < 0: raise ValueError("Kg formato debe ser >= 0")
        return (values[0], codigo, descripcion, str(values[3]).strip(), str(values[4]).strip(), str(values[5]).strip(), str(values[6]).strip(), kg, str(values[8]).strip(), str(values[9]).strip(), 1 if str(values[10]).strip() in ("1","True","true") else 0, 1 if str(values[11]).strip() in ("1","True","true") else 0, 1 if str(values[12]).strip() in ("1","True","true") else 0, str(values[13]).strip())

    def _apply_base_packaging_row_changes(self) -> None:
        if not self._base_packaging_tree: return
        selected = self._base_packaging_tree.selection()
        if not selected:
            messagebox.showerror("Confecciones base", "Seleccione una fila para editar.", parent=self); return
        try:
            values = self._validate_base_packaging_values((self._base_packaging_editor_vars["id"].get(), self._base_packaging_editor_vars["codigo"].get(), self._base_packaging_editor_vars["descripcion"].get(), self._base_packaging_editor_vars["grupo_confeccion"].get(), self._base_packaging_editor_vars["perfil_confeccion"].get(), self._base_packaging_editor_vars["familia_productiva"].get(), self._base_packaging_editor_vars["subtipo_productivo"].get(), self._base_packaging_editor_vars["kg_formato"].get(), self._base_packaging_editor_vars["tipo_malla"].get(), self._base_packaging_editor_vars["linea_productiva"].get(), self._base_packaging_editor_vars["requiere_precalibrado"].get(), self._base_packaging_editor_vars["compatible_box"].get(), self._base_packaging_editor_vars["activo"].get(), self._base_packaging_editor_vars["observaciones"].get()))
            self._base_packaging_tree.item(selected[0], values=values)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _add_base_packaging_row(self) -> None:
        if not self._base_packaging_tree: return
        code = f"NUEVA_BASE_{self._new_packaging_counter}"
        self._new_packaging_counter += 1
        self._base_packaging_tree.insert("", "end", values=("", code, "Nueva confección base", "", "", "", "", 0, "", "", 0, 0, 1, ""))

    def _delete_base_packaging_row(self) -> None:
        if not self._base_packaging_tree: return
        selected = self._base_packaging_tree.selection()
        if not selected:
            messagebox.showerror("Confecciones base", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar la confección base seleccionada?", parent=self):
            self._base_packaging_tree.delete(selected[0])

    def _collect_base_packaging_rows_payload(self) -> list[dict]:
        if not self._base_packaging_tree: return []
        rows = []; codes = set()
        for item_id in self._base_packaging_tree.get_children():
            clean = self._validate_base_packaging_values(self._base_packaging_tree.item(item_id, "values"))
            key = clean[1].lower()
            if key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(key)
            rows.append({"codigo": clean[1], "descripcion": clean[2], "grupo_confeccion": clean[3], "perfil_confeccion": clean[4], "familia_productiva": clean[5], "subtipo_productivo": clean[6], "kg_formato": clean[7], "tipo_malla": clean[8], "linea_productiva": clean[9], "requiere_precalibrado": clean[10], "compatible_box": clean[11], "activo": clean[12], "observaciones": clean[13]})
        return rows

    def _save_base_packaging_settings(self) -> None:
        try:
            self.service.save_base_packaging(self._collect_base_packaging_rows_payload())
            self._load_base_packaging_settings()
            messagebox.showinfo("Configuración productiva", "Confecciones base guardadas correctamente.", parent=self)
        except Exception as exc:
            messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_base_packaging_defaults(self) -> None:
        self.service.reset_base_packaging_defaults()
        self._load_base_packaging_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de confecciones base restaurados.", parent=self)

    def _build_lines_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_lines_help,
            export_command=lambda: self._export_master_excel("lines", self.service.get_lines),
            import_command=lambda: self._import_master_excel("lines", self.service.save_lines, self._load_lines_settings))
        catalog = ttk.LabelFrame(parent, text="Catálogo de máquinas y líneas", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew")
        catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols = ("id", "codigo", "nombre", "tipo_linea", "familia_principal", "numero_maquinas", "activa", "capacidad_kg_h_referencia", "personal_minimo", "personal_optimo", "permite_precalibrado", "permite_box", "observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=9)
        self._lines_tree = tree
        for c,h,w in [("id","ID",40),("codigo","Código",150),("nombre","Nombre",170),("tipo_linea","Tipo línea",110),("familia_principal","Familia principal",150),("numero_maquinas","Nº máquinas",90),("activa","Activa",60),("capacidad_kg_h_referencia","Capacidad kg/h referencia",160),("personal_minimo","Personal mínimo",100),("personal_optimo","Personal óptimo",100),("permite_precalibrado","Permite pre calibrado",130),("permite_box","Permite BOX",100),("observaciones","Observaciones",220)]:
            tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        yscroll = ttk.Scrollbar(catalog, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns"); xscroll.grid(row=1, column=0, sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_line_row_selected)
        editor = ttk.LabelFrame(parent, text="Editar línea seleccionada", padding=12)
        editor.grid(row=2, column=0, sticky="ew", pady=(10, 0)); editor.grid_columnconfigure(1, weight=1)
        self._lines_editor_vars = {"id": tk.StringVar(value=""), "codigo": tk.StringVar(value=""), "nombre": tk.StringVar(value=""), "tipo_linea": tk.StringVar(value=self.LINE_TYPES[0]), "familia_principal": tk.StringVar(value=self.LINE_FAMILIES[0]), "numero_maquinas": tk.StringVar(value="0"), "activa": tk.IntVar(value=1), "capacidad_kg_h_referencia": tk.StringVar(value="0"), "personal_minimo": tk.StringVar(value="0"), "personal_optimo": tk.StringVar(value="0"), "permite_precalibrado": tk.IntVar(value=0), "permite_box": tk.IntVar(value=0), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor,text="Código").grid(row=0,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._lines_editor_vars["codigo"]).grid(row=0,column=1,sticky="ew",pady=4)
        ttk.Label(editor,text="Nombre").grid(row=1,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._lines_editor_vars["nombre"]).grid(row=1,column=1,sticky="ew",pady=4)
        ttk.Label(editor,text="Tipo línea").grid(row=2,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._lines_editor_vars["tipo_linea"],values=self.LINE_TYPES,state="readonly").grid(row=2,column=1,sticky="ew",pady=4)
        ttk.Label(editor,text="Familia principal").grid(row=3,column=0,sticky="w",padx=(0,8),pady=4); ttk.Combobox(editor,textvariable=self._lines_editor_vars["familia_principal"],values=self.LINE_FAMILIES,state="readonly").grid(row=3,column=1,sticky="ew",pady=4)
        for r,k,l in ((4,"numero_maquinas","Nº máquinas"),(5,"capacidad_kg_h_referencia","Capacidad kg/h referencia"),(6,"personal_minimo","Personal mínimo"),(7,"personal_optimo","Personal óptimo")):
            ttk.Label(editor,text=l).grid(row=r,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._lines_editor_vars[k]).grid(row=r,column=1,sticky="ew",pady=4)
        ttk.Checkbutton(editor,text="Activa",variable=self._lines_editor_vars["activa"]).grid(row=8,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor,text="Permite pre calibrado",variable=self._lines_editor_vars["permite_precalibrado"]).grid(row=9,column=1,sticky="w",pady=2)
        ttk.Checkbutton(editor,text="Permite BOX",variable=self._lines_editor_vars["permite_box"]).grid(row=10,column=1,sticky="w",pady=2)
        ttk.Label(editor,text="Observaciones").grid(row=11,column=0,sticky="w",padx=(0,8),pady=4); ttk.Entry(editor,textvariable=self._lines_editor_vars["observaciones"]).grid(row=11,column=1,sticky="ew",pady=4)
        btns = ttk.Frame(parent); btns.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        ttk.Button(btns, text="Nueva línea", command=self._add_line_row).pack(side="left", padx=4)
        ttk.Button(btns, text="Aplicar cambios a selección", command=self._apply_line_row_changes).pack(side="left", padx=4)
        ttk.Button(btns, text="Guardar líneas", command=self._save_lines_settings).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=self._reset_lines_defaults).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar línea seleccionada", command=self._delete_line_row).pack(side="left", padx=4)


    def _build_performance_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1); parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_performance_help,
            export_command=lambda: self._export_master_excel("performance_rules", self.service.get_performance_rules),
            import_command=lambda: self._import_master_excel("performance_rules", self.service.save_performance_rules, self._load_performance_settings))
        catalog = ttk.LabelFrame(parent, text="Catálogo de rendimientos", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew"); catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols = ("id","codigo","familia","confeccion_formato","tipo_linea","condicion","oph_referencia","oph_minimo","oph_optimo","kg_h_referencia","factor_precalibrado","factor_destrio_alto","dificultad","activo","observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=8); self._performance_tree = tree
        for c,h,w in [("id","ID",40),("codigo","Código",150),("familia","Familia",90),("confeccion_formato","Confección / formato",180),("tipo_linea","Tipo línea",100),("condicion","Condición",120),("oph_referencia","OPH referencia",100),("oph_minimo","OPH mínimo",90),("oph_optimo","OPH óptimo",90),("kg_h_referencia","Kg/h referencia",110),("factor_precalibrado","Factor precalibrado",120),("factor_destrio_alto","Factor destrío alto",120),("dificultad","Dificultad",90),("activo","Activo",60),("observaciones","Observaciones",220)]:
            tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        yscroll = ttk.Scrollbar(catalog, orient="vertical", command=tree.yview); xscroll = ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0,column=0,sticky="nsew"); yscroll.grid(row=0,column=1,sticky="ns"); xscroll.grid(row=1,column=0,sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_performance_row_selected)
        editor = ttk.LabelFrame(parent, text="Editar rendimiento seleccionado", padding=12)
        editor.grid(row=2,column=0,sticky="ew",pady=(10,0)); editor.grid_columnconfigure(1, weight=1)
        self._performance_editor_vars = {"id": tk.StringVar(value=""), "codigo": tk.StringVar(value=""), "familia": tk.StringVar(value=self.PERFORMANCE_FAMILIES[0]), "confeccion_formato": tk.StringVar(value=""), "tipo_linea": tk.StringVar(value=self.PERFORMANCE_LINE_TYPES[0]), "condicion": tk.StringVar(value=self.PERFORMANCE_CONDITIONS[0]), "oph_referencia": tk.StringVar(value="0"), "oph_minimo": tk.StringVar(value="0"), "oph_optimo": tk.StringVar(value="0"), "kg_h_referencia": tk.StringVar(value="0"), "factor_precalibrado": tk.StringVar(value="1.0"), "factor_destrio_alto": tk.StringVar(value="1.0"), "dificultad": tk.StringVar(value=self.PERFORMANCE_DIFFICULTIES[0]), "activo": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        fields=[("codigo","Código"),("familia","Familia"),("confeccion_formato","Confección / formato"),("tipo_linea","Tipo línea"),("condicion","Condición"),("oph_referencia","OPH referencia"),("oph_minimo","OPH mínimo"),("oph_optimo","OPH óptimo"),("kg_h_referencia","Kg/h referencia"),("factor_precalibrado","Factor precalibrado"),("factor_destrio_alto","Factor destrío alto"),("dificultad","Dificultad")]
        for r,(k,l) in enumerate(fields):
            ttk.Label(editor,text=l).grid(row=r,column=0,sticky="w",padx=(0,8),pady=3)
            if k in ("familia","tipo_linea","condicion","dificultad"):
                vals = getattr(self, f"PERFORMANCE_{'LINE_TYPES' if k=='tipo_linea' else 'CONDITIONS' if k=='condicion' else 'DIFFICULTIES' if k=='dificultad' else 'FAMILIES'}")
                ttk.Combobox(editor,textvariable=self._performance_editor_vars[k],values=vals).grid(row=r,column=1,sticky="ew",pady=3)
            else: ttk.Entry(editor,textvariable=self._performance_editor_vars[k]).grid(row=r,column=1,sticky="ew",pady=3)
        ttk.Checkbutton(editor,text="Activo",variable=self._performance_editor_vars["activo"]).grid(row=12,column=1,sticky="w",pady=2)
        ttk.Label(editor,text="Observaciones").grid(row=13,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._performance_editor_vars["observaciones"]).grid(row=13,column=1,sticky="ew",pady=3)
        btns = ttk.Frame(parent); btns.grid(row=3,column=0,sticky="ew",pady=(12,0))
        ttk.Button(btns,text="Nuevo rendimiento",command=self._add_performance_row).pack(side="left",padx=4)
        ttk.Button(btns,text="Aplicar cambios a selección",command=self._apply_performance_row_changes).pack(side="left",padx=4)
        ttk.Button(btns,text="Guardar rendimientos",command=self._save_performance_settings).pack(side="left",padx=4)
        ttk.Button(btns,text="Restaurar valores por defecto",command=self._reset_performance_defaults).pack(side="left",padx=4)
        ttk.Button(btns,text="Eliminar rendimiento seleccionado",command=self._delete_performance_row).pack(side="left",padx=4)

    def _show_performance_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Rendimientos", heading="Rendimientos", intro="Esta pestaña define los rendimientos productivos de referencia que se usarán posteriormente para estimar horas necesarias, capacidad diaria y plantilla recomendada.", help_items=get_help_items(PRODUCTION_PERFORMANCE_HELP_KEYS, PRODUCTION_PERFORMANCE_HELP))


    def _load_performance_settings(self) -> None:
        self._refresh_performance_tree(self.service.get_performance_rules())

    def _refresh_performance_tree(self, rows: list[dict]) -> None:
        if not self._performance_tree: return
        self._performance_tree.delete(*self._performance_tree.get_children())
        for row in rows: self._performance_tree.insert("", "end", values=(row.get("id",""), row.get("codigo",""), row.get("familia",""), row.get("confeccion_formato",""), row.get("tipo_linea",""), row.get("condicion",""), row.get("oph_referencia",0), row.get("oph_minimo",0), row.get("oph_optimo",0), row.get("kg_h_referencia",0), row.get("factor_precalibrado",1), row.get("factor_destrio_alto",1), row.get("dificultad",""), row.get("activo",1), row.get("observaciones","")))

    def _on_performance_row_selected(self, _event=None) -> None:
        if not self._performance_tree: return
        selected=self._performance_tree.selection()
        if not selected: return
        vals=self._performance_tree.item(selected[0],"values")
        for i,k in enumerate(("id","codigo","familia","confeccion_formato","tipo_linea","condicion","oph_referencia","oph_minimo","oph_optimo","kg_h_referencia","factor_precalibrado","factor_destrio_alto","dificultad","activo","observaciones")): self._performance_editor_vars[k].set(vals[i])

    def _validate_performance_values(self, values: tuple) -> tuple:
        code,fam,form,tipo,cond = [str(values[i]).strip() for i in (1,2,3,4,5)]
        if not all([code,fam,form,tipo,cond]): raise ValueError("Código, familia, confección/formato, tipo línea y condición son obligatorios")
        oph_ref=float(str(values[6]).replace(',','.')); oph_min=float(str(values[7]).replace(',','.')); oph_opt=float(str(values[8]).replace(',','.')); kgh=float(str(values[9]).replace(',','.')); fpre=float(str(values[10]).replace(',','.')); fdes=float(str(values[11]).replace(',','.')); dif=str(values[12]).strip()
        if min(oph_ref,oph_min,oph_opt,kgh) < 0: raise ValueError("OPH y Kg/h deben ser >= 0")
        if fpre <= 0 or fdes <= 0: raise ValueError("Factores deben ser > 0")
        if oph_opt < oph_min: raise ValueError("OPH óptimo no debe ser menor que OPH mínimo")
        if not dif: raise ValueError("La dificultad es obligatoria")
        if oph_ref == 0 and kgh == 0: raise ValueError("OPH referencia y Kg/h referencia no pueden ser ambos 0")
        return (values[0],code,fam,form,tipo,cond,oph_ref,oph_min,oph_opt,kgh,fpre,fdes,dif,1 if str(values[13]).strip() in ("1","True","true") else 0,str(values[14]).strip())

    def _add_performance_row(self) -> None:
        if not self._performance_tree: return
        code=f"NUEVO_REND_{self._new_performance_counter}"; self._new_performance_counter += 1
        self._performance_tree.insert("", "end", values=("", code, "Otro", "Nuevo formato", "Otro", "Normal", 0, 0, 0, 0, 1.0, 1.0, "Media", 1, ""))

    def _apply_performance_row_changes(self) -> None:
        if not self._performance_tree: return
        selected=self._performance_tree.selection()
        if not selected: messagebox.showerror("Rendimientos", "Seleccione una fila para editar.", parent=self); return
        try:
            vals=self._validate_performance_values((self._performance_editor_vars["id"].get(), self._performance_editor_vars["codigo"].get(), self._performance_editor_vars["familia"].get(), self._performance_editor_vars["confeccion_formato"].get(), self._performance_editor_vars["tipo_linea"].get(), self._performance_editor_vars["condicion"].get(), self._performance_editor_vars["oph_referencia"].get(), self._performance_editor_vars["oph_minimo"].get(), self._performance_editor_vars["oph_optimo"].get(), self._performance_editor_vars["kg_h_referencia"].get(), self._performance_editor_vars["factor_precalibrado"].get(), self._performance_editor_vars["factor_destrio_alto"].get(), self._performance_editor_vars["dificultad"].get(), self._performance_editor_vars["activo"].get(), self._performance_editor_vars["observaciones"].get()))
            self._performance_tree.item(selected[0], values=vals)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _delete_performance_row(self) -> None:
        if not self._performance_tree: return
        selected=self._performance_tree.selection()
        if not selected: messagebox.showerror("Rendimientos", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar el rendimiento seleccionado?", parent=self): self._performance_tree.delete(selected[0])

    def _collect_performance_rows_payload(self) -> list[dict]:
        if not self._performance_tree: return []
        rows=[]; codes=set()
        for item_id in self._performance_tree.get_children():
            clean=self._validate_performance_values(self._performance_tree.item(item_id,"values")); key=clean[1].lower()
            if key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(key)
            rows.append({"codigo":clean[1],"familia":clean[2],"confeccion_formato":clean[3],"tipo_linea":clean[4],"condicion":clean[5],"oph_referencia":clean[6],"oph_minimo":clean[7],"oph_optimo":clean[8],"kg_h_referencia":clean[9],"factor_precalibrado":clean[10],"factor_destrio_alto":clean[11],"dificultad":clean[12],"activo":clean[13],"observaciones":clean[14]})
        return rows

    def _save_performance_settings(self) -> None:
        try: self.service.save_performance_rules(self._collect_performance_rows_payload()); self._load_performance_settings(); messagebox.showinfo("Configuración productiva", "Rendimientos guardados correctamente.", parent=self)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_performance_defaults(self) -> None:
        self.service.reset_performance_defaults(); self._load_performance_settings(); messagebox.showinfo("Configuración productiva", "Valores por defecto de rendimientos restaurados.", parent=self)

    def _build_penalties_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1); parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_penalties_help,
            export_command=lambda: self._export_master_excel("penalty_rules", self.service.get_penalty_rules),
            import_command=lambda: self._import_master_excel("penalty_rules", self.service.save_penalty_rules, self._load_penalty_settings))
        catalog = ttk.LabelFrame(parent, text="Catálogo de penalizaciones operativas", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew"); catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols = ("id","codigo","tipo_penalizacion","ambito","minutos_perdida","factor_rendimiento","aplica_por","umbral","activa","observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=8); self._penalties_tree = tree
        for c,h,w in [("id","ID",40),("codigo","Código",140),("tipo_penalizacion","Tipo penalización",160),("ambito","Ámbito",120),("minutos_perdida","Minutos pérdida",110),("factor_rendimiento","Factor rendimiento",120),("aplica_por","Aplica por",130),("umbral","Umbral",180),("activa","Activa",60),("observaciones","Observaciones",220)]:
            tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        yscroll = ttk.Scrollbar(catalog, orient="vertical", command=tree.yview); xscroll = ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        tree.grid(row=0,column=0,sticky="nsew"); yscroll.grid(row=0,column=1,sticky="ns"); xscroll.grid(row=1,column=0,sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_penalty_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar penalización seleccionada", padding=12)
        editor.grid(row=2,column=0,sticky="ew",pady=(10,0)); editor.grid_columnconfigure(1, weight=1)
        self._penalties_editor_vars = {"id": tk.StringVar(value=""), "codigo": tk.StringVar(value=""), "tipo_penalizacion": tk.StringVar(value=self.PENALTY_TYPES[0]), "ambito": tk.StringVar(value=self.PENALTY_SCOPES[0]), "minutos_perdida": tk.StringVar(value="0"), "factor_rendimiento": tk.StringVar(value="1.00"), "aplica_por": tk.StringVar(value=self.PENALTY_APPLIES[0]), "umbral": tk.StringVar(value=""), "activa": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor,text="Código").grid(row=0,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._penalties_editor_vars["codigo"]).grid(row=0,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Tipo penalización").grid(row=1,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._penalties_editor_vars["tipo_penalizacion"],values=self.PENALTY_TYPES).grid(row=1,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Ámbito").grid(row=2,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._penalties_editor_vars["ambito"],values=self.PENALTY_SCOPES).grid(row=2,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Minutos pérdida").grid(row=3,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._penalties_editor_vars["minutos_perdida"]).grid(row=3,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Factor rendimiento").grid(row=4,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._penalties_editor_vars["factor_rendimiento"]).grid(row=4,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Aplica por").grid(row=5,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._penalties_editor_vars["aplica_por"],values=self.PENALTY_APPLIES).grid(row=5,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Umbral").grid(row=6,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._penalties_editor_vars["umbral"]).grid(row=6,column=1,sticky="ew",pady=3)
        ttk.Checkbutton(editor,text="Activa",variable=self._penalties_editor_vars["activa"]).grid(row=7,column=1,sticky="w",pady=2)
        ttk.Label(editor,text="Observaciones").grid(row=8,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._penalties_editor_vars["observaciones"]).grid(row=8,column=1,sticky="ew",pady=3)

        ttk.Button(editor,text="Aplicar cambios a selección",command=self._apply_penalty_row_changes).grid(row=9,column=1,sticky="e",pady=(6,0))
        btns = ttk.Frame(parent); btns.grid(row=3,column=0,sticky="ew",pady=(12,0))
        ttk.Button(btns,text="Nueva penalización",command=self._add_penalty_row).pack(side="left",padx=4)
        ttk.Button(btns,text="Guardar penalizaciones",command=self._save_penalty_settings).pack(side="left",padx=4)
        ttk.Button(btns,text="Restaurar valores por defecto",command=self._reset_penalty_defaults).pack(side="left",padx=4)
        ttk.Button(btns,text="Eliminar penalización seleccionada",command=self._delete_penalty_row).pack(side="left",padx=4)

    def _show_penalties_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Penalizaciones", heading="Penalizaciones", intro="Esta pestaña define pérdidas de tiempo y ajustes de rendimiento asociados a cambios, paradas o fragmentación del trabajo. Estas reglas permitirán estimar una capacidad más realista que el simple cálculo por kilos u OPH.", help_items=get_help_items(PRODUCTION_PENALTIES_HELP_KEYS, PRODUCTION_PENALTIES_HELP))

    def _load_penalty_settings(self) -> None:
        self._refresh_penalties_tree(self.service.get_penalty_rules())

    def _refresh_penalties_tree(self, rows: list[dict]) -> None:
        if not self._penalties_tree: return
        self._penalties_tree.delete(*self._penalties_tree.get_children())
        for row in rows: self._penalties_tree.insert("", "end", values=(row.get("id",""), row.get("codigo",""), row.get("tipo_penalizacion",""), row.get("ambito",""), row.get("minutos_perdida",0), row.get("factor_rendimiento",1), row.get("aplica_por",""), row.get("umbral",""), row.get("activa",1), row.get("observaciones","")))

    def _on_penalty_row_selected(self, _event=None) -> None:
        if not self._penalties_tree: return
        selected=self._penalties_tree.selection()
        if not selected: return
        vals=self._penalties_tree.item(selected[0],"values")
        for i,k in enumerate(("id","codigo","tipo_penalizacion","ambito","minutos_perdida","factor_rendimiento","aplica_por","umbral","activa","observaciones")): self._penalties_editor_vars[k].set(vals[i])

    def _validate_penalty_values(self, values: tuple) -> tuple:
        codigo = str(values[1]).strip(); tipo = str(values[2]).strip(); ambito = str(values[3]).strip(); aplica = str(values[6]).strip()
        if not codigo: raise ValueError("El código es obligatorio")
        if not tipo: raise ValueError("El tipo de penalización es obligatorio")
        if not ambito: raise ValueError("El ámbito es obligatorio")
        minutos=float(str(values[4]).replace(',', '.')); factor=float(str(values[5]).replace(',', '.'))
        if minutos < 0: raise ValueError("Minutos pérdida debe ser >= 0")
        if factor <= 0 or factor > 1.5: raise ValueError("Factor rendimiento debe ser > 0 y <= 1.5")
        if not aplica: raise ValueError("Aplica por es obligatorio")
        if minutos == 0 and abs(factor - 1.0) < 1e-9: raise ValueError("La penalización no tendría efecto: ajuste minutos pérdida o factor rendimiento")
        return (values[0], codigo, tipo, ambito, minutos, factor, aplica, str(values[7]).strip(), 1 if str(values[8]).strip() in ("1","True","true") else 0, str(values[9]).strip())

    def _add_penalty_row(self) -> None:
        if not self._penalties_tree: return
        code=f"NUEVA_PENAL_{self._new_penalty_counter}"; self._new_penalty_counter += 1
        self._penalties_tree.insert("", "end", values=("", code, "Otro", "General", 0, 0.95, "Cada cambio", "", 1, ""))

    def _apply_penalty_row_changes(self) -> None:
        if not self._penalties_tree: return
        selected=self._penalties_tree.selection()
        if not selected: messagebox.showerror("Penalizaciones", "Seleccione una fila para editar.", parent=self); return
        try:
            vals=self._validate_penalty_values((self._penalties_editor_vars["id"].get(), self._penalties_editor_vars["codigo"].get(), self._penalties_editor_vars["tipo_penalizacion"].get(), self._penalties_editor_vars["ambito"].get(), self._penalties_editor_vars["minutos_perdida"].get(), self._penalties_editor_vars["factor_rendimiento"].get(), self._penalties_editor_vars["aplica_por"].get(), self._penalties_editor_vars["umbral"].get(), self._penalties_editor_vars["activa"].get(), self._penalties_editor_vars["observaciones"].get()))
            self._penalties_tree.item(selected[0], values=vals)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _delete_penalty_row(self) -> None:
        if not self._penalties_tree: return
        selected=self._penalties_tree.selection()
        if not selected: messagebox.showerror("Penalizaciones", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar la penalización seleccionada?", parent=self): self._penalties_tree.delete(selected[0])

    def _collect_penalty_rows_payload(self) -> list[dict]:
        if not self._penalties_tree: return []
        rows=[]; codes=set()
        for item_id in self._penalties_tree.get_children():
            clean=self._validate_penalty_values(self._penalties_tree.item(item_id,"values")); key=clean[1].lower()
            if key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(key)
            rows.append({"codigo": clean[1], "tipo_penalizacion": clean[2], "ambito": clean[3], "minutos_perdida": clean[4], "factor_rendimiento": clean[5], "aplica_por": clean[6], "umbral": clean[7], "activa": clean[8], "observaciones": clean[9]})
        return rows

    def _save_penalty_settings(self) -> None:
        try: self.service.save_penalty_rules(self._collect_penalty_rows_payload()); self._load_penalty_settings(); messagebox.showinfo("Configuración productiva", "Penalizaciones guardadas correctamente.", parent=self)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_penalty_defaults(self) -> None:
        self.service.reset_penalty_defaults(); self._load_penalty_settings(); messagebox.showinfo("Configuración productiva", "Valores por defecto de penalizaciones restaurados.", parent=self)


    def _build_semaphore_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1); parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_semaphore_help,
            export_command=lambda: self._export_master_excel("semaphore_rules", self.service.get_semaphore_rules),
            import_command=lambda: self._import_master_excel("semaphore_rules", self.service.save_semaphore_rules, self._load_semaphore_settings, self._validate_semaphore_import_rows))
        catalog = ttk.LabelFrame(parent, text="Catálogo de reglas operativas", padding=12)
        catalog.grid(row=1, column=0, sticky="nsew"); catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        cols=("id","codigo","tipo_regla","ambito","metrica","operador","umbral_amarillo","umbral_rojo","accion_sugerida","activa","observaciones")
        tree=ttk.Treeview(catalog, columns=cols, show="headings", height=8); self._semaphore_tree=tree
        for c,h,w in [("id","ID",40),("codigo","Código",170),("tipo_regla","Tipo regla",150),("ambito","Ámbito",120),("metrica","Métrica",140),("operador","Operador",70),("umbral_amarillo","Umbral amarillo",120),("umbral_rojo","Umbral rojo",100),("accion_sugerida","Acción sugerida",280),("activa","Activa",60),("observaciones","Observaciones",200)]:
            tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        y=ttk.Scrollbar(catalog, orient="vertical", command=tree.yview); x=ttk.Scrollbar(catalog, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y.set, xscrollcommand=x.set); tree.grid(row=0,column=0,sticky="nsew"); y.grid(row=0,column=1,sticky="ns"); x.grid(row=1,column=0,sticky="ew")
        tree.bind("<<TreeviewSelect>>", self._on_semaphore_row_selected)
        editor = ttk.LabelFrame(parent, text="Editar regla seleccionada", padding=12); editor.grid(row=2,column=0,sticky="ew",pady=(10,0)); editor.grid_columnconfigure(1, weight=1)
        self._semaphore_editor_vars={"id":tk.StringVar(value=""),"codigo":tk.StringVar(value=""),"tipo_regla":tk.StringVar(value=self.SEMAPHORE_RULE_TYPES[0]),"ambito":tk.StringVar(value=self.SEMAPHORE_SCOPES[0]),"metrica":tk.StringVar(value=self.SEMAPHORE_METRICS[0]),"operador":tk.StringVar(value=self.SEMAPHORE_OPERATORS[0]),"umbral_amarillo":tk.StringVar(value="0"),"umbral_rojo":tk.StringVar(value="0"),"accion_sugerida":tk.StringVar(value=""),"activa":tk.IntVar(value=1),"observaciones":tk.StringVar(value="")}
        rows=[("codigo","Código"),("tipo_regla","Tipo regla"),("ambito","Ámbito"),("metrica","Métrica"),("operador","Operador"),("umbral_amarillo","Umbral amarillo"),("umbral_rojo","Umbral rojo"),("accion_sugerida","Acción sugerida"),("observaciones","Observaciones")]
        for i,(k,l) in enumerate(rows):
            ttk.Label(editor,text=l).grid(row=i,column=0,sticky="w",padx=(0,8),pady=3)
            if k=="tipo_regla": ttk.Combobox(editor,textvariable=self._semaphore_editor_vars[k],values=self.SEMAPHORE_RULE_TYPES).grid(row=i,column=1,sticky="ew",pady=3)
            elif k=="ambito": ttk.Combobox(editor,textvariable=self._semaphore_editor_vars[k],values=self.SEMAPHORE_SCOPES).grid(row=i,column=1,sticky="ew",pady=3)
            elif k=="metrica": ttk.Combobox(editor,textvariable=self._semaphore_editor_vars[k],values=self.SEMAPHORE_METRICS).grid(row=i,column=1,sticky="ew",pady=3)
            elif k=="operador": ttk.Combobox(editor,textvariable=self._semaphore_editor_vars[k],values=self.SEMAPHORE_OPERATORS,state="readonly").grid(row=i,column=1,sticky="ew",pady=3)
            else: ttk.Entry(editor,textvariable=self._semaphore_editor_vars[k]).grid(row=i,column=1,sticky="ew",pady=3)
        ttk.Checkbutton(editor,text="Activa",variable=self._semaphore_editor_vars["activa"]).grid(row=9,column=1,sticky="w",pady=2)
        btns=ttk.Frame(parent); btns.grid(row=3,column=0,sticky="ew",pady=(12,0))
        ttk.Button(btns,text="Nueva regla",command=self._add_semaphore_row).pack(side="left",padx=4)
        ttk.Button(btns,text="Aplicar cambios a selección",command=self._apply_semaphore_row_changes).pack(side="left",padx=4)
        ttk.Button(btns,text="Guardar reglas",command=self._save_semaphore_settings).pack(side="left",padx=4)
        ttk.Button(btns,text="Restaurar valores por defecto",command=self._reset_semaphore_defaults).pack(side="left",padx=4)
        ttk.Button(btns,text="Eliminar regla seleccionada",command=self._delete_semaphore_row).pack(side="left",padx=4)

    def _show_semaphore_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Reglas / semáforo", heading="Reglas / semáforo", intro="Esta pestaña define los criterios que permitirán convertir los cálculos de producción en alertas claras: verde, amarillo o rojo. Las reglas no calculan por sí solas, pero indican cuándo una situación debe considerarse normal, en riesgo o crítica.", help_items=get_help_items(PRODUCTION_SEMAPHORE_HELP_KEYS, PRODUCTION_SEMAPHORE_HELP))

    def _load_semaphore_settings(self) -> None:
        self._refresh_semaphore_tree(self.service.get_semaphore_rules())

    def _refresh_semaphore_tree(self, rows: list[dict]) -> None:
        if not self._semaphore_tree: return
        self._semaphore_tree.delete(*self._semaphore_tree.get_children())
        for row in rows: self._semaphore_tree.insert("", "end", values=(row.get("id",""), row.get("codigo",""), row.get("tipo_regla",""), row.get("ambito",""), row.get("metrica",""), row.get("operador",""), row.get("umbral_amarillo",0), row.get("umbral_rojo",0), row.get("accion_sugerida",""), row.get("activa",1), row.get("observaciones","")))

    def _on_semaphore_row_selected(self, _event=None) -> None:
        if not self._semaphore_tree: return
        sel=self._semaphore_tree.selection()
        if not sel: return
        vals=self._semaphore_tree.item(sel[0],"values")
        for i,k in enumerate(("id","codigo","tipo_regla","ambito","metrica","operador","umbral_amarillo","umbral_rojo","accion_sugerida","activa","observaciones")): self._semaphore_editor_vars[k].set(vals[i])

    def _validate_semaphore_values(self, values: tuple) -> tuple:
        codigo,tipo,ambito,metrica,operador=[str(values[i]).strip() for i in (1,2,3,4,5)]
        if not codigo or not tipo or not ambito or not metrica or not operador: raise ValueError("Código, tipo regla, ámbito, métrica y operador son obligatorios")
        amarillo=float(str(values[6]).replace(',', '.')); rojo=float(str(values[7]).replace(',', '.'))
        if operador in (">=",">") and rojo < amarillo: raise ValueError("Para operadores >= o >, umbral rojo debe ser mayor o igual que umbral amarillo")
        if operador in ("<=","<") and rojo > amarillo: raise ValueError("Para operadores <= o <, umbral rojo debe ser menor o igual que umbral amarillo")
        return (values[0],codigo,tipo,ambito,metrica,operador,amarillo,rojo,str(values[8]).strip(),1 if str(values[9]).strip() in ("1","True","true") else 0,str(values[10]).strip())

    def _add_semaphore_row(self) -> None:
        if not self._semaphore_tree: return
        code=f"NUEVA_REGLA_{self._new_semaphore_counter}"; self._new_semaphore_counter += 1
        self._semaphore_tree.insert("", "end", values=("", code, "Otro", "General", "otro", ">=", 0, 0, "", 1, ""))

    def _apply_semaphore_row_changes(self) -> None:
        if not self._semaphore_tree: return
        sel=self._semaphore_tree.selection()
        if not sel: messagebox.showerror("Reglas / semáforo", "Seleccione una fila para editar.", parent=self); return
        try:
            vals=self._validate_semaphore_values((self._semaphore_editor_vars["id"].get(),self._semaphore_editor_vars["codigo"].get(),self._semaphore_editor_vars["tipo_regla"].get(),self._semaphore_editor_vars["ambito"].get(),self._semaphore_editor_vars["metrica"].get(),self._semaphore_editor_vars["operador"].get(),self._semaphore_editor_vars["umbral_amarillo"].get(),self._semaphore_editor_vars["umbral_rojo"].get(),self._semaphore_editor_vars["accion_sugerida"].get(),self._semaphore_editor_vars["activa"].get(),self._semaphore_editor_vars["observaciones"].get()))
            self._semaphore_tree.item(sel[0], values=vals)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _delete_semaphore_row(self) -> None:
        if not self._semaphore_tree: return
        sel=self._semaphore_tree.selection()
        if not sel: messagebox.showerror("Reglas / semáforo", "Seleccione una fila para eliminar.", parent=self); return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar la regla seleccionada?", parent=self): self._semaphore_tree.delete(sel[0])

    def _collect_semaphore_rows_payload(self) -> list[dict]:
        if not self._semaphore_tree: return []
        rows=[]; codes=set()
        for item_id in self._semaphore_tree.get_children():
            clean=self._validate_semaphore_values(self._semaphore_tree.item(item_id, "values")); key=clean[1].lower()
            if key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(key)
            rows.append({"codigo":clean[1],"tipo_regla":clean[2],"ambito":clean[3],"metrica":clean[4],"operador":clean[5],"umbral_amarillo":clean[6],"umbral_rojo":clean[7],"accion_sugerida":clean[8],"activa":clean[9],"observaciones":clean[10]})
        return rows

    def _save_semaphore_settings(self) -> None:
        try: self.service.save_semaphore_rules(self._collect_semaphore_rows_payload()); self._load_semaphore_settings(); messagebox.showinfo("Configuración productiva", "Reglas de semáforo guardadas correctamente.", parent=self)
        except Exception as exc: messagebox.showerror("Datos inválidos", str(exc), parent=self)

    def _reset_semaphore_defaults(self) -> None:
        self.service.reset_semaphore_defaults(); self._load_semaphore_settings(); messagebox.showinfo("Configuración productiva", "Valores por defecto de reglas de semáforo restaurados.", parent=self)

    def _build_caliber_factors_tab(self, parent: ttk.Frame) -> None:
        self._build_tab_toolbar(parent, help_command=self._show_caliber_factors_help,
            export_command=lambda: self._export_master_excel("caliber_factors", self.service.get_caliber_performance_factors),
            import_command=lambda: self._import_master_excel("caliber_factors", self.service.save_caliber_performance_factors, self._load_caliber_factors_settings, self._validate_caliber_import_rows))
        catalog = ttk.LabelFrame(parent, text="Catálogo de factores por calibre", padding=12); catalog.grid(row=1, column=0, sticky="nsew")
        cols = ("id","codigo","confeccion_familia","grupo_calibre","calibres_incluidos","factor_rendimiento","aplica_a","activo","observaciones")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=8); self._caliber_tree = tree
        for c,h,w in [("id","ID",40),("codigo","Código",160),("confeccion_familia","Confección / familia",150),("grupo_calibre","Grupo calibre",110),("calibres_incluidos","Calibres incluidos",130),("factor_rendimiento","Factor rendimiento",120),("aplica_a","Aplica a",90),("activo","Activo",60),("observaciones","Observaciones",250)]: tree.heading(c,text=h); tree.column(c,width=w,anchor="w")
        tree.grid(row=0, column=0, sticky="nsew"); catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1); tree.bind("<<TreeviewSelect>>", self._on_caliber_factor_selected)

        editor = ttk.LabelFrame(parent, text="Editar factor seleccionado", padding=12); editor.grid(row=2, column=0, sticky="ew", pady=(10, 0)); editor.grid_columnconfigure(1, weight=1)
        self._caliber_editor_vars = {"id": tk.StringVar(value=""), "codigo": tk.StringVar(value=""), "confeccion_familia": tk.StringVar(value=self.CALIBER_FAMILIES[0]), "grupo_calibre": tk.StringVar(value=self.CALIBER_GROUPS[0]), "calibres_incluidos": tk.StringVar(value=""), "factor_rendimiento": tk.StringVar(value="1.00"), "aplica_a": tk.StringVar(value=self.CALIBER_APPLIES[2]), "activo": tk.IntVar(value=1), "observaciones": tk.StringVar(value="")}
        ttk.Label(editor,text="Código").grid(row=0,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._caliber_editor_vars["codigo"]).grid(row=0,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Confección / familia").grid(row=1,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._caliber_editor_vars["confeccion_familia"],values=self.CALIBER_FAMILIES).grid(row=1,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Grupo calibre").grid(row=2,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._caliber_editor_vars["grupo_calibre"],values=self.CALIBER_GROUPS).grid(row=2,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Calibres incluidos").grid(row=3,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._caliber_editor_vars["calibres_incluidos"]).grid(row=3,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Factor rendimiento").grid(row=4,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._caliber_editor_vars["factor_rendimiento"]).grid(row=4,column=1,sticky="ew",pady=3)
        ttk.Label(editor,text="Aplica a").grid(row=5,column=0,sticky="w",padx=(0,8),pady=3); ttk.Combobox(editor,textvariable=self._caliber_editor_vars["aplica_a"],values=self.CALIBER_APPLIES).grid(row=5,column=1,sticky="ew",pady=3)
        ttk.Checkbutton(editor,text="Activo",variable=self._caliber_editor_vars["activo"]).grid(row=6,column=1,sticky="w",pady=2)
        ttk.Label(editor,text="Observaciones").grid(row=7,column=0,sticky="w",padx=(0,8),pady=3); ttk.Entry(editor,textvariable=self._caliber_editor_vars["observaciones"]).grid(row=7,column=1,sticky="ew",pady=3)
        ttk.Button(editor,text="Aplicar cambios a selección",command=self._apply_caliber_factor_changes).grid(row=8,column=1,sticky="e",pady=(6,0))
        btns = ttk.Frame(parent); btns.grid(row=3,column=0,sticky="ew",pady=(10,0))
        ttk.Button(btns,text="Nuevo factor",command=self._add_caliber_factor_row).pack(side="left",padx=4)
        ttk.Button(btns,text="Guardar factores",command=self._save_caliber_factors_settings).pack(side="left",padx=4)
        ttk.Button(btns,text="Restaurar valores por defecto",command=self._reset_caliber_factors_defaults).pack(side="left",padx=4)
        ttk.Button(btns,text="Eliminar factor seleccionado",command=self._delete_caliber_factor_row).pack(side="left",padx=4)

    def _show_caliber_factors_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Factores calibre", heading="Factores calibre", intro="Esta pestaña permite ajustar el rendimiento productivo según el calibre y la confección. El efecto del calibre no es igual en mallas, encajado o granel, por eso estas reglas son configurables.", help_items=get_help_items(PRODUCTION_CALIBER_FACTORS_HELP_KEYS, PRODUCTION_CALIBER_FACTORS_HELP))

    def _load_caliber_factors_settings(self) -> None:
        if not self._caliber_tree: return
        self._caliber_tree.delete(*self._caliber_tree.get_children())
        for row in self.service.get_caliber_performance_factors(): self._caliber_tree.insert("", "end", values=(row.get("id",""), row.get("codigo",""), row.get("confeccion_familia",""), row.get("grupo_calibre",""), row.get("calibres_incluidos",""), row.get("factor_rendimiento",1), row.get("aplica_a","Ambos"), row.get("activo",1), row.get("observaciones","")))

    def _on_caliber_factor_selected(self, _event=None) -> None:
        if not self._caliber_tree: return
        selected=self._caliber_tree.selection()
        if not selected: return
        vals=self._caliber_tree.item(selected[0],"values")
        for i,k in enumerate(("id","codigo","confeccion_familia","grupo_calibre","calibres_incluidos","factor_rendimiento","aplica_a","activo","observaciones")): self._caliber_editor_vars[k].set(vals[i])

    def _normalize_calibres(self, raw: str) -> str:
        parts=[p.strip() for p in raw.replace("/",", ").replace(" ",", ").split(",") if p.strip()]
        unique=[]
        for p in parts:
            if p not in unique: unique.append(p)
        return ",".join(unique)

    def _validate_caliber_factor_values(self, values: tuple) -> tuple:
        row_id,codigo,familia,grupo,calibres,factor,aplica,activo,obs=values
        codigo=(codigo or "").strip(); familia=(familia or "").strip(); grupo=(grupo or "").strip(); aplica=(aplica or "").strip()
        if not codigo: raise ValueError("El código es obligatorio")
        if not familia: raise ValueError("La confección / familia es obligatoria")
        if not grupo: raise ValueError("El grupo calibre es obligatorio")
        calibres_norm=self._normalize_calibres(calibres or "")
        if not calibres_norm: raise ValueError("Calibres incluidos es obligatorio")
        if len(calibres_norm.split(",")) != len(set(calibres_norm.split(","))): raise ValueError("No se permiten calibres duplicados en la misma regla")
        factor_val=float(factor)
        if factor_val <= 0 or factor_val > 2: raise ValueError("Factor rendimiento debe ser > 0 y <= 2")
        if not aplica: raise ValueError("Aplica a es obligatorio")
        return row_id,codigo,familia,grupo,calibres_norm,round(factor_val,4),aplica,1 if int(activo) else 0,(obs or "").strip()

    def _add_caliber_factor_row(self) -> None:
        if not self._caliber_tree: return
        code=f"NUEVO_FACTOR_CALIBRE_{self._new_caliber_counter}"; self._new_caliber_counter += 1
        self._caliber_tree.insert("", "end", values=("", code, "Encajado", "Medio", "4,5", 1.00, "Ambos", 1, ""))

    def _apply_caliber_factor_changes(self) -> None:
        if not self._caliber_tree: return
        selected=self._caliber_tree.selection()
        if not selected: return messagebox.showwarning("Configuración productiva", "Seleccione una fila para aplicar cambios.", parent=self)
        try:
            vals=self._validate_caliber_factor_values((self._caliber_editor_vars["id"].get(), self._caliber_editor_vars["codigo"].get(), self._caliber_editor_vars["confeccion_familia"].get(), self._caliber_editor_vars["grupo_calibre"].get(), self._caliber_editor_vars["calibres_incluidos"].get(), self._caliber_editor_vars["factor_rendimiento"].get(), self._caliber_editor_vars["aplica_a"].get(), self._caliber_editor_vars["activo"].get(), self._caliber_editor_vars["observaciones"].get()))
            self._caliber_tree.item(selected[0], values=vals)
        except ValueError as exc: messagebox.showerror("Validación", str(exc), parent=self)

    def _delete_caliber_factor_row(self) -> None:
        if not self._caliber_tree: return
        selected=self._caliber_tree.selection()
        if not selected: return
        if messagebox.askyesno("Confirmar eliminación", "¿Desea eliminar el factor seleccionado?", parent=self): self._caliber_tree.delete(selected[0])

    def _collect_caliber_rows_payload(self) -> list[dict]:
        if not self._caliber_tree: return []
        rows=[]; codes=set(); overlaps=[]
        conf_map={}
        for item_id in self._caliber_tree.get_children():
            clean=self._validate_caliber_factor_values(self._caliber_tree.item(item_id,"values")); key=clean[1].lower()
            if key in codes: raise ValueError(f"Código duplicado: {clean[1]}")
            codes.add(key)
            calibres=clean[4].split(",")
            if clean[7]:
                ckey=clean[2].lower(); conf_map.setdefault(ckey,{})
                for c in calibres:
                    if c in conf_map[ckey] and conf_map[ckey][c] != clean[3]: overlaps.append(f"{clean[2]}: calibre {c} en {conf_map[ckey][c]} y {clean[3]}")
                    conf_map[ckey][c]=clean[3]
            rows.append({"codigo": clean[1], "confeccion_familia": clean[2], "grupo_calibre": clean[3], "calibres_incluidos": clean[4], "factor_rendimiento": clean[5], "aplica_a": clean[6], "activo": clean[7], "observaciones": clean[8]})
        if overlaps and not messagebox.askyesno("Solapamiento de calibres", "Se detectaron calibres en más de un grupo activo para la misma confección/familia:\n\n- " + "\n- ".join(overlaps) + "\n\n¿Desea continuar y guardar igualmente?", parent=self): raise ValueError("Guardado cancelado por solapamiento de calibres.")
        return rows

    def _save_caliber_factors_settings(self) -> None:
        try: self.service.save_caliber_performance_factors(self._collect_caliber_rows_payload()); self._load_caliber_factors_settings(); messagebox.showinfo("Configuración productiva", "Factores de calibre guardados correctamente.", parent=self)
        except ValueError as exc: messagebox.showerror("Validación", str(exc), parent=self)

    def _reset_caliber_factors_defaults(self) -> None:
        self.service.reset_caliber_performance_factors_defaults(); self._load_caliber_factors_settings(); messagebox.showinfo("Configuración productiva", "Valores por defecto de factores calibre restaurados.", parent=self)

    def _build_packaging_mapping_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_packaging_mapping_help,
            export_command=lambda: self._export_master_excel("packaging_mapping", lambda: self.service.get_packaging_mapping(False)),
            import_command=lambda: self._import_master_excel("packaging_mapping", self.service.save_packaging_mapping, lambda: self._load_packaging_mapping_settings(False), self._validate_mapping_import_rows))
        catalog = ttk.LabelFrame(parent, text="Catálogo de mapeo productivo", padding=8)
        catalog.grid(row=1, column=0, sticky="nsew")
        cols = ("codigo","nombre","grupo","neto","npiezas","familia","subtipo","kg","tipo_malla","linea","activo","revisar","confianza","obs")
        tree = ttk.Treeview(catalog, columns=cols, show="headings", height=8)
        self._mapping_tree = tree
        headers = ["Código","Nombre","Grupo origen","Neto origen","N. piezas origen","Familia productiva","Subtipo productivo","Kg formato","Tipo malla","Línea productiva","Activo producción","Revisar","Confianza","Observaciones"]
        for c, h in zip(cols, headers):
            tree.heading(c, text=h); tree.column(c, width=110, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        catalog.grid_columnconfigure(0, weight=1); catalog.grid_rowconfigure(0, weight=1)
        tree.bind("<<TreeviewSelect>>", self._on_mapping_row_selected)

        editor = ttk.LabelFrame(parent, text="Editar mapeo seleccionado", padding=8)
        editor.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        editor.grid_columnconfigure(1, weight=1)
        self._mapping_editor_vars = {k: tk.StringVar(value="") for k in ("codigo_mconfeccion","nombre_mconfeccion","descripcion_corta","grupo_origen","neto_origen","npiezas_origen","familia_productiva","subtipo_productivo","kg_formato","tipo_malla","linea_productiva","observaciones","confianza_autodeteccion")}
        for k in ("requiere_precalibrado","compatible_box","activo_produccion","revisar"): self._mapping_editor_vars[k] = tk.IntVar(value=0)
        ro = {"state": "readonly"}
        ttk.Label(editor,text="Código MConfección").grid(row=0,column=0,sticky="w"); ttk.Entry(editor,textvariable=self._mapping_editor_vars["codigo_mconfeccion"], **ro).grid(row=0,column=1,sticky="ew")
        ttk.Label(editor,text="Nombre").grid(row=1,column=0,sticky="w"); ttk.Entry(editor,textvariable=self._mapping_editor_vars["nombre_mconfeccion"], **ro).grid(row=1,column=1,sticky="ew")
        ttk.Label(editor,text="Grupo origen").grid(row=2,column=0,sticky="w"); ttk.Entry(editor,textvariable=self._mapping_editor_vars["grupo_origen"], **ro).grid(row=2,column=1,sticky="ew")
        ttk.Label(editor,text="Familia productiva").grid(row=3,column=0,sticky="w"); ttk.Combobox(editor,textvariable=self._mapping_editor_vars["familia_productiva"],values=self.MAPPING_FAMILIES,state="readonly").grid(row=3,column=1,sticky="ew")
        ttk.Label(editor,text="Subtipo productivo").grid(row=4,column=0,sticky="w"); ttk.Combobox(editor,textvariable=self._mapping_editor_vars["subtipo_productivo"],values=self.MAPPING_SUBTYPES).grid(row=4,column=1,sticky="ew")
        ttk.Label(editor,text="Kg formato").grid(row=5,column=0,sticky="w"); ttk.Entry(editor,textvariable=self._mapping_editor_vars["kg_formato"]).grid(row=5,column=1,sticky="ew")
        ttk.Label(editor,text="Tipo malla").grid(row=6,column=0,sticky="w"); ttk.Combobox(editor,textvariable=self._mapping_editor_vars["tipo_malla"],values=self.MAPPING_MESH_TYPES,state="readonly").grid(row=6,column=1,sticky="ew")
        ttk.Label(editor,text="Línea productiva").grid(row=7,column=0,sticky="w"); ttk.Combobox(editor,textvariable=self._mapping_editor_vars["linea_productiva"],values=[r.get("codigo") for r in self.service.get_lines()] or ["MALLAS_TRADICIONAL","MALLAS_CLIP","MALLAS_GIRSAC","ENCAJADO","GRANEL_MANUAL","GRANELERA","OTRO"]).grid(row=7,column=1,sticky="ew")
        ttk.Checkbutton(editor,text="Requiere pre calibrado",variable=self._mapping_editor_vars["requiere_precalibrado"]).grid(row=8,column=1,sticky="w")
        ttk.Checkbutton(editor,text="Compatible BOX",variable=self._mapping_editor_vars["compatible_box"]).grid(row=9,column=1,sticky="w")
        ttk.Checkbutton(editor,text="Activo producción",variable=self._mapping_editor_vars["activo_produccion"]).grid(row=10,column=1,sticky="w")
        ttk.Checkbutton(editor,text="Revisar",variable=self._mapping_editor_vars["revisar"]).grid(row=11,column=1,sticky="w")
        ttk.Label(editor,text="Observaciones").grid(row=12,column=0,sticky="w"); ttk.Entry(editor,textvariable=self._mapping_editor_vars["observaciones"]).grid(row=12,column=1,sticky="ew")

        btns = ttk.Frame(parent); btns.grid(row=3,column=0,sticky="ew", pady=(8,0))
        ttk.Button(btns,text="Autorrellenar desde MConfecciones",command=self._autofill_packaging_mapping).pack(side="left",padx=4)
        ttk.Button(btns,text="Aplicar cambios a selección",command=self._apply_mapping_row_changes).pack(side="left",padx=4)
        ttk.Button(btns,text="Guardar mapeo",command=self._save_packaging_mapping).pack(side="left",padx=4)
        ttk.Button(btns,text="Restaurar autodetección",command=self._reset_packaging_mapping_autodetect).pack(side="left",padx=4)
        ttk.Button(btns,text="Filtrar pendientes de revisar",command=lambda: self._load_packaging_mapping_settings(True)).pack(side="left",padx=4)
        ttk.Button(btns,text="Ver incoherencias",command=self._show_packaging_mapping_incoherences).pack(side="left",padx=4)
        ttk.Button(btns,text="Ver todos",command=lambda: self._load_packaging_mapping_settings(False)).pack(side="left",padx=4)

    def _load_packaging_mapping_settings(self, only_review: bool = False) -> None:
        rows = self.service.get_packaging_mapping(only_review)
        if not rows:
            self.service.autofill_packaging_mapping_from_mconfecciones(False)
            rows = self.service.get_packaging_mapping(only_review)
        self._refresh_mapping_tree(rows)

    def _refresh_mapping_tree(self, rows: list[dict]) -> None:
        if not self._mapping_tree:
            return
        self._mapping_tree.delete(*self._mapping_tree.get_children())
        for r in rows:
            self._mapping_tree.insert("", "end", values=(r.get("codigo_mconfeccion",""),r.get("nombre_mconfeccion",""),r.get("grupo_origen",""),r.get("neto_origen",0),r.get("npiezas_origen",0),r.get("familia_productiva",""),r.get("subtipo_productivo",""),r.get("kg_formato",0),r.get("tipo_malla",""),r.get("linea_productiva",""),r.get("activo_produccion",1),r.get("revisar",0),r.get("confianza_autodeteccion",""),r.get("observaciones","")))

    def _show_packaging_mapping_incoherences(self) -> None:
        rows = self.service.get_packaging_mapping(False)
        if not rows:
            self.service.autofill_packaging_mapping_from_mconfecciones(False)
            rows = self.service.get_packaging_mapping(False)
        lines_by_code = {str(r.get("codigo", "") or "").strip(): r for r in self.service.get_lines()}
        filtered = []
        for row in rows:
            reasons = self._mapping_incoherence_reasons(row, lines_by_code)
            if not reasons:
                continue
            display_row = dict(row)
            obs = str(display_row.get("observaciones", "") or "")
            suffix = "INCOHERENCIAS: " + "; ".join(reasons)
            display_row["observaciones"] = suffix if not obs else f"{obs} | {suffix}"
            filtered.append(display_row)
        self._refresh_mapping_tree(filtered)
        if not filtered:
            messagebox.showinfo("Mapeo confecciones", "No se han detectado incoherencias en el mapeo productivo.", parent=self)

    def _mapping_incoherence_reasons(self, row: dict, lines_by_code: dict[str, dict]) -> list[str]:
        reasons = []
        line_code = str(row.get("linea_productiva", "") or "").strip()
        tipo_malla = str(row.get("tipo_malla", "") or "").strip()
        line = lines_by_code.get(line_code)
        if line_code and (not line or int(line.get("activa", 0) or 0) != 1):
            reasons.append("línea inactiva asignada")
        if line_code == "MALLAS_CLIP" and (not line or int(line.get("activa", 0) or 0) != 1):
            reasons.append("línea MALLAS_CLIP inactiva")
        expected_mesh_by_line = {
            "MALLAS_TRADICIONAL": "Tradicional",
            "MALLAS_GIRSAC": "Girsac",
            "MALLAS_CLIP": "Clip-to-clip",
        }
        expected = expected_mesh_by_line.get(line_code)
        if expected and tipo_malla != expected:
            reasons.append(f"tipo_malla incompatible con línea (esperado {expected})")
        if line_code == "MALLAS_TRADICIONAL" and tipo_malla == "Clip-to-clip":
            reasons.append("tipo_malla Clip-to-clip con línea Tradicional")
        return reasons

    def _on_mapping_row_selected(self, _event=None) -> None:
        if not self._mapping_tree or not self._mapping_tree.selection(): return
        vals = self._mapping_tree.item(self._mapping_tree.selection()[0], "values")
        keys = ["codigo_mconfeccion","nombre_mconfeccion","grupo_origen","neto_origen","npiezas_origen","familia_productiva","subtipo_productivo","kg_formato","tipo_malla","linea_productiva","activo_produccion","revisar","confianza_autodeteccion","observaciones"]
        for i,k in enumerate(keys): self._mapping_editor_vars[k].set(vals[i])

    def _apply_mapping_row_changes(self) -> None:
        if not self._mapping_tree or not self._mapping_tree.selection(): return
        kg = float(self._mapping_editor_vars["kg_formato"].get() or 0)
        if kg < 0: raise ValueError("Kg formato debe ser >= 0")
        vals = (self._mapping_editor_vars["codigo_mconfeccion"].get(), self._mapping_editor_vars["nombre_mconfeccion"].get(), self._mapping_editor_vars["grupo_origen"].get(), self._mapping_editor_vars["neto_origen"].get(), self._mapping_editor_vars["npiezas_origen"].get(), self._mapping_editor_vars["familia_productiva"].get(), self._mapping_editor_vars["subtipo_productivo"].get(), kg, self._mapping_editor_vars["tipo_malla"].get(), self._mapping_editor_vars["linea_productiva"].get(), int(self._mapping_editor_vars["activo_produccion"].get()), int(self._mapping_editor_vars["revisar"].get()), self._mapping_editor_vars["confianza_autodeteccion"].get(), self._mapping_editor_vars["observaciones"].get())
        self._mapping_tree.item(self._mapping_tree.selection()[0], values=vals)

    def _save_packaging_mapping(self) -> None:
        if not self._mapping_tree: return
        rows = []
        for item in self._mapping_tree.get_children():
            v = self._mapping_tree.item(item, "values")
            rows.append({"codigo_mconfeccion": v[0], "nombre_mconfeccion": v[1], "grupo_origen": v[2], "neto_origen": v[3], "npiezas_origen": v[4], "familia_productiva": v[5], "subtipo_productivo": v[6], "kg_formato": v[7], "tipo_malla": v[8], "linea_productiva": v[9], "activo_produccion": v[10], "revisar": v[11], "confianza_autodeteccion": v[12], "observaciones": v[13], "requiere_precalibrado": int(self._mapping_editor_vars["requiere_precalibrado"].get()), "compatible_box": int(self._mapping_editor_vars["compatible_box"].get())})
        self.service.save_packaging_mapping(rows)
        messagebox.showinfo("Configuración productiva", "Mapeo de confecciones guardado.", parent=self)

    def _autofill_packaging_mapping(self) -> None:
        overwrite = messagebox.askyesno("Configuración productiva", "¿Sobrescribir también registros ya existentes?")
        stats = self.service.autofill_packaging_mapping_from_mconfecciones(overwrite)
        self._load_packaging_mapping_settings(False)
        messagebox.showinfo("Configuración productiva", f"Autorrelleno completado. Nuevos: {stats['created']}, actualizados: {stats['updated']}", parent=self)

    def _reset_packaging_mapping_autodetect(self) -> None:
        self.service.reset_packaging_mapping_autodetect()
        self._load_packaging_mapping_settings(False)


    def _build_resources_flows_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(parent, help_command=self._show_resources_flows_help,
            export_command=self._export_resources_flows_excel,
            import_command=self._import_resources_flows_excel)

        inner = ttk.Notebook(parent)
        inner.grid(row=1, column=0, sticky="nsew")

        def _build_block(tab_title, columns, headers, widths, tree_attr, on_select):
            tab = ttk.Frame(inner, padding=8)
            tab.grid_columnconfigure(0, weight=1)
            tab.grid_rowconfigure(0, weight=1)
            inner.add(tab, text=tab_title)

            frame = ttk.LabelFrame(tab, text=tab_title, padding=8)
            frame.grid(row=0, column=0, sticky="nsew")
            frame.grid_columnconfigure(0, weight=1)
            frame.grid_rowconfigure(0, weight=1)

            tree = ttk.Treeview(frame, columns=columns, show="headings", height=12)
            for col in columns:
                tree.heading(col, text=headers[col]); tree.column(col, width=widths.get(col, 110), anchor="w")
            ys = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            xs = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
            tree.grid(row=0, column=0, sticky="nsew"); ys.grid(row=0, column=1, sticky="ns"); xs.grid(row=1, column=0, sticky="ew")
            tree.bind("<<TreeviewSelect>>", on_select)
            setattr(self, tree_attr, tree)
            return frame

        physical = _build_block("Recursos físicos", ("id","codigo","nombre","tipo_recurso","familia_operativa","capacidad_kg_h","capacidad_por","numero_unidades","personal_minimo","personal_optimo","activo","observaciones"),
            {"id":"ID","codigo":"Código","nombre":"Nombre","tipo_recurso":"Tipo recurso","familia_operativa":"Familia operativa","capacidad_kg_h":"Capacidad kg/h","capacidad_por":"Capacidad por","numero_unidades":"Nº unidades","personal_minimo":"Personal mínimo","personal_optimo":"Personal óptimo","activo":"Activo","observaciones":"Observaciones"},
            {"id":40,"codigo":140,"nombre":170,"tipo_recurso":120,"familia_operativa":140,"capacidad_kg_h":110,"capacidad_por":110,"numero_unidades":95,"personal_minimo":95,"personal_optimo":95,"activo":70,"observaciones":220},"_physical_resources_tree", self._on_physical_resource_selected)
        self._physical_resources_editor_vars = {k: tk.StringVar(value="") for k in ("codigo","nombre","tipo_recurso","familia_operativa","capacidad_kg_h","capacidad_por","numero_unidades","personal_minimo","personal_optimo","observaciones")}; self._physical_resources_editor_vars["activo"] = tk.IntVar(value=1)
        self._build_resources_editor(physical, self._physical_resources_editor_vars, [("codigo","Código"),("nombre","Nombre"),("tipo_recurso","Tipo recurso"),("familia_operativa","Familia operativa"),("capacidad_kg_h","Capacidad kg/h"),("capacidad_por","Capacidad por"),("numero_unidades","Nº unidades"),("personal_minimo","Personal mínimo"),("personal_optimo","Personal óptimo"),("observaciones","Observaciones")],
            [("Nuevo recurso", self._new_physical_resource_row),("Aplicar cambios", self._apply_physical_resource_changes),("Eliminar recurso", self._delete_physical_resource_row),("Guardar recursos", self._save_physical_resources_settings)])

        comp = _build_block("Compatibilidades", ("id","recurso_codigo","compatible_con","valor","activo","observaciones"),
            {"id":"ID","recurso_codigo":"Recurso código","compatible_con":"Compatible con","valor":"Valor","activo":"Activo","observaciones":"Observaciones"},
            {"id":40,"recurso_codigo":170,"compatible_con":160,"valor":140,"activo":70,"observaciones":280},"_resource_compatibilities_tree", self._on_resource_compatibility_selected)
        self._resource_compatibilities_editor_vars = {k: tk.StringVar(value="") for k in ("recurso_codigo","compatible_con","valor","observaciones")}; self._resource_compatibilities_editor_vars["activo"] = tk.IntVar(value=1)
        self._build_resources_editor(comp, self._resource_compatibilities_editor_vars, [("recurso_codigo","Recurso código"),("compatible_con","Compatible con"),("valor","Valor"),("observaciones","Observaciones")],
            [("Nueva compatibilidad", self._new_resource_compatibility_row),("Aplicar cambios", self._apply_resource_compatibility_changes),("Eliminar compatibilidad", self._delete_resource_compatibility_row),("Guardar compatibilidades", self._save_resource_compatibilities_settings)])

        feed = _build_block("Alimentación / conexiones", ("id","origen_codigo","destino_codigo","max_destinos_simultaneos","requiere_precalibrado","activo","observaciones"),
            {"id":"ID","origen_codigo":"Origen código","destino_codigo":"Destino código","max_destinos_simultaneos":"Máx destinos simultáneos","requiere_precalibrado":"Requiere precalibrado","activo":"Activo","observaciones":"Observaciones"},
            {"id":40,"origen_codigo":160,"destino_codigo":160,"max_destinos_simultaneos":140,"requiere_precalibrado":150,"activo":70,"observaciones":260},"_resource_feeds_tree", self._on_resource_feed_selected)
        self._resource_feeds_editor_vars = {k: tk.StringVar(value="") for k in ("origen_codigo","destino_codigo","max_destinos_simultaneos","observaciones")}; self._resource_feeds_editor_vars["requiere_precalibrado"] = tk.IntVar(value=0); self._resource_feeds_editor_vars["activo"] = tk.IntVar(value=1)
        self._build_resources_editor(feed, self._resource_feeds_editor_vars, [("origen_codigo","Origen código"),("destino_codigo","Destino código"),("max_destinos_simultaneos","Máx destinos simultáneos"),("observaciones","Observaciones")],
            [("Nueva conexión", self._new_resource_feed_row),("Aplicar cambios", self._apply_resource_feed_changes),("Eliminar conexión", self._delete_resource_feed_row),("Guardar alimentación", self._save_resource_feeds_settings)], include_precalibrado=True)

        avail = _build_block("Disponibilidad operativa", ("id","recurso_codigo","contexto","disponible","motivo","prioridad","observaciones"),
            {"id":"ID","recurso_codigo":"Recurso código","contexto":"Contexto","disponible":"Disponible","motivo":"Motivo","prioridad":"Prioridad","observaciones":"Observaciones"},
            {"id":40,"recurso_codigo":170,"contexto":180,"disponible":80,"motivo":170,"prioridad":80,"observaciones":230},"_resource_availability_tree", self._on_resource_availability_selected)
        self._resource_availability_editor_vars = {k: tk.StringVar(value="") for k in ("recurso_codigo","contexto","motivo","prioridad","observaciones")}; self._resource_availability_editor_vars["disponible"] = tk.IntVar(value=1)
        self._build_resources_editor(avail, self._resource_availability_editor_vars, [("recurso_codigo","Recurso código"),("contexto","Contexto"),("motivo","Motivo"),("prioridad","Prioridad"),("observaciones","Observaciones")],
            [("Nueva disponibilidad", self._new_resource_availability_row),("Aplicar cambios", self._apply_resource_availability_changes),("Eliminar disponibilidad", self._delete_resource_availability_row),("Guardar disponibilidad", self._save_resource_availability_settings)], include_disponible=True)

        ttk.Button(parent, text="Restaurar valores por defecto", command=self._reset_resources_flows_defaults).grid(row=2, column=0, sticky="w", pady=(10, 0))

    def _show_resources_flows_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Recursos y flujos", heading="Recursos y flujos", intro="Define recursos físicos (Compacta, calibrador principal, pesadoras, Awetta), compatibilidades entre recursos, conexiones de alimentación y disponibilidad operativa por contexto.", help_items=get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP))

    def _build_resources_editor(self, parent, vars_map, fields, buttons, include_precalibrado=False, include_disponible=False):
        e = ttk.LabelFrame(parent, text="Edición", padding=6); e.grid(row=2, column=0, sticky="ew", pady=(6, 0)); e.grid_columnconfigure(1, weight=1)
        for i, (k, t) in enumerate(fields):
            ttk.Label(e, text=t).grid(row=i, column=0, sticky="w"); ttk.Entry(e, textvariable=vars_map[k]).grid(row=i, column=1, sticky="ew")
        row = len(fields)
        ttk.Checkbutton(e, text="Activo", variable=vars_map.get("activo", tk.IntVar(value=1))).grid(row=row, column=1, sticky="w"); row += 1
        if include_precalibrado: ttk.Checkbutton(e, text="Requiere precalibrado", variable=vars_map["requiere_precalibrado"]).grid(row=row, column=1, sticky="w"); row += 1
        if include_disponible: ttk.Checkbutton(e, text="Disponible", variable=vars_map["disponible"]).grid(row=row, column=1, sticky="w"); row += 1
        b = ttk.Frame(e); b.grid(row=row, column=0, columnspan=2, sticky="w", pady=(6, 0))
        for text, cmd in buttons: ttk.Button(b, text=text, command=cmd).pack(side="left", padx=4)

    def _load_physical_resources_settings(self) -> None: self._refresh_physical_resources_tree(self.service.get_physical_resources())
    def _load_resource_compatibilities_settings(self) -> None: self._refresh_resource_compatibilities_tree(self.service.get_resource_compatibilities())
    def _load_resource_feeds_settings(self) -> None: self._refresh_resource_feeds_tree(self.service.get_resource_feeds())
    def _load_resource_availability_settings(self) -> None: self._refresh_resource_availability_tree(self.service.get_resource_availability())

    def _refresh_physical_resources_tree(self, rows: list[dict]) -> None:
        if not self._physical_resources_tree: return
        self._physical_resources_tree.delete(*self._physical_resources_tree.get_children())
        for r in rows: self._physical_resources_tree.insert("", "end", values=tuple(r.get(k, "") for k in ("id","codigo","nombre","tipo_recurso","familia_operativa","capacidad_kg_h","capacidad_por","numero_unidades","personal_minimo","personal_optimo","activo","observaciones")))
    def _refresh_resource_compatibilities_tree(self, rows: list[dict]) -> None:
        if not self._resource_compatibilities_tree: return
        self._resource_compatibilities_tree.delete(*self._resource_compatibilities_tree.get_children())
        for r in rows: self._resource_compatibilities_tree.insert("", "end", values=tuple(r.get(k, "") for k in ("id","recurso_codigo","compatible_con","valor","activo","observaciones")))
    def _refresh_resource_feeds_tree(self, rows: list[dict]) -> None:
        if not self._resource_feeds_tree: return
        self._resource_feeds_tree.delete(*self._resource_feeds_tree.get_children())
        for r in rows: self._resource_feeds_tree.insert("", "end", values=tuple(r.get(k, "") for k in ("id","origen_codigo","destino_codigo","max_destinos_simultaneos","requiere_precalibrado","activo","observaciones")))
    def _refresh_resource_availability_tree(self, rows: list[dict]) -> None:
        if not self._resource_availability_tree: return
        self._resource_availability_tree.delete(*self._resource_availability_tree.get_children())
        for r in rows: self._resource_availability_tree.insert("", "end", values=tuple(r.get(k, "") for k in ("id","recurso_codigo","contexto","disponible","motivo","prioridad","observaciones")))

    def _set_editor_from_values(self, vars_map, keys, values):
        for i, k in enumerate(keys):
            if k in vars_map: vars_map[k].set(values[i] if i < len(values) else "")

    def _new_row_in_tree(self, tree, values):
        tree.insert("", "end", values=values)

    def _delete_selected(self, tree, msg):
        sel = tree.selection()
        if not sel: messagebox.showerror("Recursos y flujos", msg, parent=self); return
        tree.delete(sel[0])

    def _on_physical_resource_selected(self, _=None):
        if self._physical_resources_tree and self._physical_resources_tree.selection(): self._set_editor_from_values(self._physical_resources_editor_vars, ["id","codigo","nombre","tipo_recurso","familia_operativa","capacidad_kg_h","capacidad_por","numero_unidades","personal_minimo","personal_optimo","activo","observaciones"], self._physical_resources_tree.item(self._physical_resources_tree.selection()[0], "values"))
    _on_resource_compatibility_selected = lambda self, _=None: self._resource_compatibilities_tree and self._resource_compatibilities_tree.selection() and self._set_editor_from_values(self._resource_compatibilities_editor_vars,["id","recurso_codigo","compatible_con","valor","activo","observaciones"], self._resource_compatibilities_tree.item(self._resource_compatibilities_tree.selection()[0], "values"))
    _on_resource_feed_selected = lambda self, _=None: self._resource_feeds_tree and self._resource_feeds_tree.selection() and self._set_editor_from_values(self._resource_feeds_editor_vars,["id","origen_codigo","destino_codigo","max_destinos_simultaneos","requiere_precalibrado","activo","observaciones"], self._resource_feeds_tree.item(self._resource_feeds_tree.selection()[0], "values"))
    _on_resource_availability_selected = lambda self, _=None: self._resource_availability_tree and self._resource_availability_tree.selection() and self._set_editor_from_values(self._resource_availability_editor_vars,["id","recurso_codigo","contexto","disponible","motivo","prioridad","observaciones"], self._resource_availability_tree.item(self._resource_availability_tree.selection()[0], "values"))

    def _new_physical_resource_row(self): self._new_row_in_tree(self._physical_resources_tree, ("", self._physical_resources_editor_vars["codigo"].get(), self._physical_resources_editor_vars["nombre"].get(), self._physical_resources_editor_vars["tipo_recurso"].get(), self._physical_resources_editor_vars["familia_operativa"].get(), self._physical_resources_editor_vars["capacidad_kg_h"].get() or 0, self._physical_resources_editor_vars["capacidad_por"].get() or "Recurso", self._physical_resources_editor_vars["numero_unidades"].get() or 1, self._physical_resources_editor_vars["personal_minimo"].get() or 0, self._physical_resources_editor_vars["personal_optimo"].get() or 0, int(self._physical_resources_editor_vars["activo"].get()), self._physical_resources_editor_vars["observaciones"].get()))
    def _new_resource_compatibility_row(self): self._new_row_in_tree(self._resource_compatibilities_tree, ("", self._resource_compatibilities_editor_vars["recurso_codigo"].get(), self._resource_compatibilities_editor_vars["compatible_con"].get(), self._resource_compatibilities_editor_vars["valor"].get(), int(self._resource_compatibilities_editor_vars["activo"].get()), self._resource_compatibilities_editor_vars["observaciones"].get()))
    def _new_resource_feed_row(self): self._new_row_in_tree(self._resource_feeds_tree, ("", self._resource_feeds_editor_vars["origen_codigo"].get(), self._resource_feeds_editor_vars["destino_codigo"].get(), self._resource_feeds_editor_vars["max_destinos_simultaneos"].get() or 1, int(self._resource_feeds_editor_vars["requiere_precalibrado"].get()), int(self._resource_feeds_editor_vars["activo"].get()), self._resource_feeds_editor_vars["observaciones"].get()))
    def _new_resource_availability_row(self): self._new_row_in_tree(self._resource_availability_tree, ("", self._resource_availability_editor_vars["recurso_codigo"].get(), self._resource_availability_editor_vars["contexto"].get(), int(self._resource_availability_editor_vars["disponible"].get()), self._resource_availability_editor_vars["motivo"].get(), self._resource_availability_editor_vars["prioridad"].get() or 0, self._resource_availability_editor_vars["observaciones"].get()))

    def _apply_physical_resource_changes(self):
        if not self._physical_resources_tree or not self._physical_resources_tree.selection(): return
        self._physical_resources_tree.item(self._physical_resources_tree.selection()[0], values=("", self._physical_resources_editor_vars["codigo"].get(), self._physical_resources_editor_vars["nombre"].get(), self._physical_resources_editor_vars["tipo_recurso"].get(), self._physical_resources_editor_vars["familia_operativa"].get(), self._physical_resources_editor_vars["capacidad_kg_h"].get() or 0, self._physical_resources_editor_vars["capacidad_por"].get(), self._physical_resources_editor_vars["numero_unidades"].get() or 0, self._physical_resources_editor_vars["personal_minimo"].get() or 0, self._physical_resources_editor_vars["personal_optimo"].get() or 0, int(self._physical_resources_editor_vars["activo"].get()), self._physical_resources_editor_vars["observaciones"].get()))
    def _apply_resource_compatibility_changes(self):
        if self._resource_compatibilities_tree and self._resource_compatibilities_tree.selection(): self._resource_compatibilities_tree.item(self._resource_compatibilities_tree.selection()[0], values=("", self._resource_compatibilities_editor_vars["recurso_codigo"].get(), self._resource_compatibilities_editor_vars["compatible_con"].get(), self._resource_compatibilities_editor_vars["valor"].get(), int(self._resource_compatibilities_editor_vars["activo"].get()), self._resource_compatibilities_editor_vars["observaciones"].get()))
    def _apply_resource_feed_changes(self):
        if self._resource_feeds_tree and self._resource_feeds_tree.selection(): self._resource_feeds_tree.item(self._resource_feeds_tree.selection()[0], values=("", self._resource_feeds_editor_vars["origen_codigo"].get(), self._resource_feeds_editor_vars["destino_codigo"].get(), self._resource_feeds_editor_vars["max_destinos_simultaneos"].get() or 1, int(self._resource_feeds_editor_vars["requiere_precalibrado"].get()), int(self._resource_feeds_editor_vars["activo"].get()), self._resource_feeds_editor_vars["observaciones"].get()))
    def _apply_resource_availability_changes(self):
        if self._resource_availability_tree and self._resource_availability_tree.selection(): self._resource_availability_tree.item(self._resource_availability_tree.selection()[0], values=("", self._resource_availability_editor_vars["recurso_codigo"].get(), self._resource_availability_editor_vars["contexto"].get(), int(self._resource_availability_editor_vars["disponible"].get()), self._resource_availability_editor_vars["motivo"].get(), self._resource_availability_editor_vars["prioridad"].get() or 0, self._resource_availability_editor_vars["observaciones"].get()))

    def _delete_physical_resource_row(self): self._delete_selected(self._physical_resources_tree, "Seleccione un recurso para eliminar.")
    def _delete_resource_compatibility_row(self): self._delete_selected(self._resource_compatibilities_tree, "Seleccione una compatibilidad para eliminar.")
    def _delete_resource_feed_row(self): self._delete_selected(self._resource_feeds_tree, "Seleccione una conexión para eliminar.")
    def _delete_resource_availability_row(self): self._delete_selected(self._resource_availability_tree, "Seleccione una disponibilidad para eliminar.")

    def _collect_physical_resources_rows_payload(self) -> list[dict]:
        rows=[]
        if not self._physical_resources_tree: return rows
        for it in self._physical_resources_tree.get_children():
            v=self._physical_resources_tree.item(it,"values"); rows.append({"codigo":v[1],"nombre":v[2],"tipo_recurso":v[3],"familia_operativa":v[4],"capacidad_kg_h":v[5],"capacidad_por":v[6],"numero_unidades":v[7],"personal_minimo":v[8],"personal_optimo":v[9],"activo":v[10],"observaciones":v[11]})
        return rows
    def _collect_resource_compatibilities_rows_payload(self) -> list[dict]:
        rows=[]
        if not self._resource_compatibilities_tree: return rows
        for it in self._resource_compatibilities_tree.get_children():
            v=self._resource_compatibilities_tree.item(it,"values"); rows.append({"recurso_codigo":v[1],"compatible_con":v[2],"valor":v[3],"activo":v[4],"observaciones":v[5]})
        return rows
    def _collect_resource_feeds_rows_payload(self) -> list[dict]:
        rows=[]
        if not self._resource_feeds_tree: return rows
        for it in self._resource_feeds_tree.get_children():
            v=self._resource_feeds_tree.item(it,"values"); rows.append({"origen_codigo":v[1],"destino_codigo":v[2],"max_destinos_simultaneos":v[3],"requiere_precalibrado":v[4],"activo":v[5],"observaciones":v[6]})
        return rows
    def _collect_resource_availability_rows_payload(self) -> list[dict]:
        rows=[]
        if not self._resource_availability_tree: return rows
        for it in self._resource_availability_tree.get_children():
            v=self._resource_availability_tree.item(it,"values"); rows.append({"recurso_codigo":v[1],"contexto":v[2],"disponible":v[3],"motivo":v[4],"prioridad":v[5],"observaciones":v[6]})
        return rows

    def _save_physical_resources_settings(self) -> None:
        rows = self._collect_physical_resources_rows_payload()
        codes = set()
        for i, r in enumerate(rows, start=1):
            if not str(r.get("codigo", "")).strip(): raise ValueError(f"Recursos fila {i}: codigo obligatorio.")
            if r["codigo"] in codes: raise ValueError(f"Recursos fila {i}: codigo duplicado.")
            codes.add(r["codigo"])
            if not str(r.get("nombre", "")).strip(): raise ValueError(f"Recursos fila {i}: nombre obligatorio.")
            if float(r.get("capacidad_kg_h", 0)) < 0 or int(float(r.get("numero_unidades", 0))) < 0 or int(float(r.get("personal_minimo", 0))) < 0: raise ValueError(f"Recursos fila {i}: valores no pueden ser negativos.")
            if int(float(r.get("personal_optimo", 0))) < int(float(r.get("personal_minimo", 0))): raise ValueError(f"Recursos fila {i}: personal_optimo debe ser >= personal_minimo.")
        self.service.save_physical_resources(rows); self._load_physical_resources_settings(); messagebox.showinfo("Configuración productiva", "Recursos guardados correctamente.", parent=self)
    def _save_resource_compatibilities_settings(self) -> None:
        rows = self._collect_resource_compatibilities_rows_payload()
        for i, r in enumerate(rows, start=1):
            if not str(r.get("recurso_codigo", "")).strip() or not str(r.get("compatible_con", "")).strip() or not str(r.get("valor", "")).strip(): raise ValueError(f"Compatibilidades fila {i}: recurso_codigo, compatible_con y valor son obligatorios.")
        self.service.save_resource_compatibilities(rows); self._load_resource_compatibilities_settings(); messagebox.showinfo("Configuración productiva", "Compatibilidades guardadas correctamente.", parent=self)
    def _save_resource_feeds_settings(self) -> None:
        rows = self._collect_resource_feeds_rows_payload()
        for i, r in enumerate(rows, start=1):
            if not str(r.get("origen_codigo", "")).strip() or not str(r.get("destino_codigo", "")).strip(): raise ValueError(f"Alimentación fila {i}: origen_codigo y destino_codigo obligatorios.")
            if int(float(r.get("max_destinos_simultaneos", 0))) < 1: raise ValueError(f"Alimentación fila {i}: max_destinos_simultaneos debe ser >= 1.")
        self.service.save_resource_feeds(rows); self._load_resource_feeds_settings(); messagebox.showinfo("Configuración productiva", "Alimentación guardada correctamente.", parent=self)
    def _save_resource_availability_settings(self) -> None:
        rows = self._collect_resource_availability_rows_payload()
        for i, r in enumerate(rows, start=1):
            if not str(r.get("recurso_codigo", "")).strip() or not str(r.get("contexto", "")).strip(): raise ValueError(f"Disponibilidad fila {i}: recurso_codigo y contexto obligatorios.")
            if int(float(r.get("prioridad", 0))) < 0: raise ValueError(f"Disponibilidad fila {i}: prioridad debe ser >= 0.")
        self.service.save_resource_availability(rows); self._load_resource_availability_settings(); messagebox.showinfo("Configuración productiva", "Disponibilidad guardada correctamente.", parent=self)

    def _reset_resources_flows_defaults(self) -> None:
        self.service.reset_resources_flows_defaults()
        self._load_physical_resources_settings(); self._load_resource_compatibilities_settings(); self._load_resource_feeds_settings(); self._load_resource_availability_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto de recursos y flujos restaurados.", parent=self)

    def _export_resources_flows_excel(self) -> None:
        help_rows = [{"campo": i.get("title", ""), "descripcion": i.get("description", "")} for i in get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP)]
        self._export_master_excel("resources_flows", lambda: {
            "Recursos físicos": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["physical_resources"]["columns"], "rows": self.service.get_physical_resources()},
            "Compatibilidades": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["resource_compatibilities"]["columns"], "rows": self.service.get_resource_compatibilities()},
            "Alimentación": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["resource_feeds"]["columns"], "rows": self.service.get_resource_feeds()},
            "Disponibilidad": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["resource_availability"]["columns"], "rows": self.service.get_resource_availability()},
            "Descripcion campos": {"columns": ["campo", "descripcion"], "rows": help_rows},
        })

    def _import_resources_flows_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path: return
        wb = load_workbook(path, data_only=True)
        cfg = PRODUCTION_MASTER_EXCEL_CONFIGS

        def read(sheet, config):
            if sheet not in wb.sheetnames: return []
            ws = wb[sheet]
            headers=[str(h or "").strip() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            rows=[]
            for vals in ws.iter_rows(min_row=2, values_only=True):
                row={headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
                rows.append({c: row.get(c, "") for c in config["columns"]})
            return rows

        def merge_rows(existing, imported, key_builder):
            merged = {key_builder(row): row for row in existing if key_builder(row)}
            for row in imported:
                key = key_builder(row)
                if key:
                    merged[key] = row
            return list(merged.values())

        physical = read("Recursos físicos", cfg["physical_resources"])
        comp = read("Compatibilidades", cfg["resource_compatibilities"])
        feeds = read("Alimentación", cfg["resource_feeds"])
        avail = read("Disponibilidad", cfg["resource_availability"])

        if not messagebox.askyesno("Confirmar importación", "Se importarán recursos y flujos actualizando por clave y sin borrar ausentes. ¿Continuar?", parent=self): return

        merged_physical = merge_rows(self.service.get_physical_resources(), physical, lambda r: str(r.get("codigo", "")).strip().lower())
        merged_comp = merge_rows(self.service.get_resource_compatibilities(), comp, lambda r: "|".join([str(r.get("recurso_codigo", "")).strip().lower(), str(r.get("compatible_con", "")).strip().lower(), str(r.get("valor", "")).strip().lower()]))
        merged_feeds = merge_rows(self.service.get_resource_feeds(), feeds, lambda r: "|".join([str(r.get("origen_codigo", "")).strip().lower(), str(r.get("destino_codigo", "")).strip().lower()]))
        merged_avail = merge_rows(self.service.get_resource_availability(), avail, lambda r: "|".join([str(r.get("recurso_codigo", "")).strip().lower(), str(r.get("contexto", "")).strip().lower()]))

        self.service.save_physical_resources(merged_physical)
        self.service.save_resource_compatibilities(merged_comp)
        self.service.save_resource_feeds(merged_feeds)
        self.service.save_resource_availability(merged_avail)
        self._load_physical_resources_settings(); self._load_resource_compatibilities_settings(); self._load_resource_feeds_settings(); self._load_resource_availability_settings()



    def _build_capacity_productive_tab(self, parent: ttk.Frame) -> None:
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        self._build_tab_toolbar(
            parent,
            help_command=self._show_capacity_productive_help,
            export_command=self._export_capacity_productive_excel,
            import_command=self._import_capacity_productive_excel,
        )
        inner = ttk.Notebook(parent)
        inner.grid(row=1, column=0, sticky="nsew")
        self._capacity_productive_notebook = inner
        self._capacity_master_actions = {}
        specs = [
            ("Familias productivas", "_productive_families_tree", ("id", "codigo", "descripcion", "orden", "activa", "observaciones", "updated_at"), {"id": 40, "codigo": 140, "descripcion": 220, "orden": 70, "activa": 70, "observaciones": 260, "updated_at": 180}, "productive_families", self.service.get_productive_families, self.service.save_productive_families, self.service.reset_productive_families_defaults),
            ("Configuración líneas capacidad", "_line_capacity_config_tree", ("id", "linea_productiva", "familia_productiva", "puesto_productivo_principal", "modo_uso_recursos", "usar_capacidad_agregada", "activa", "observaciones", "updated_at"), {"id": 40, "linea_productiva": 170, "familia_productiva": 140, "puesto_productivo_principal": 190, "modo_uso_recursos": 140, "usar_capacidad_agregada": 160, "activa": 70, "observaciones": 250, "updated_at": 180}, "line_capacity_config", self.service.get_line_capacity_config, self.service.save_line_capacity_config, self.service.reset_line_capacity_config_defaults),
            ("Recursos requeridos por línea", "_line_required_resources_tree", ("id", "linea_productiva", "recurso_codigo", "obligatorio", "modo_uso", "reparte_kg", "orden", "activo", "observaciones", "updated_at"), {"id": 40, "linea_productiva": 170, "recurso_codigo": 180, "obligatorio": 90, "modo_uso": 110, "reparte_kg": 90, "orden": 70, "activo": 70, "observaciones": 250, "updated_at": 180}, "line_required_resources", self.service.get_line_required_resources, self.service.save_line_required_resources, self.service.reset_line_required_resources_defaults),
            ("Equivalencias de puestos", "_staff_area_equivalences_tree", ("id", "area_requerida", "area_personal", "prioridad", "activa", "observaciones", "updated_at"), {"id": 40, "area_requerida": 190, "area_personal": 190, "prioridad": 80, "activa": 70, "observaciones": 260, "updated_at": 180}, "staff_area_equivalences", self.service.get_staff_area_equivalences, self.service.save_staff_area_equivalences, self.service.reset_staff_area_equivalences_defaults),
            ("Polivalencias de personal", "_staff_polyvalence_tree", ("id", "puesto_origen", "puesto_destino", "prioridad", "factor_productividad", "activa", "observaciones", "updated_at"), {"id": 40, "puesto_origen": 190, "puesto_destino": 190, "prioridad": 80, "factor_productividad": 140, "activa": 70, "observaciones": 260, "updated_at": 180}, "staff_polyvalence", self.service.get_staff_polyvalence, self.service.save_staff_polyvalence, self.service.reset_staff_polyvalence_defaults),
        ]
        for title, attr, columns, widths, key, loader, saver, resetter in specs:
            frame = ttk.Frame(inner, padding=8)
            frame.grid_columnconfigure(0, weight=1)
            frame.grid_rowconfigure(0, weight=1)
            inner.add(frame, text=title)
            self._capacity_master_actions[title] = (key, loader, saver)
            self._build_capacity_master_block(frame, title, attr, columns, widths, key, loader, saver, resetter)

    def _build_capacity_master_block(self, parent, title, tree_attr, columns, widths, master_key, loader, saver, resetter) -> None:
        tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=widths.get(col, 110), anchor="w")
        ys = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        xs = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        setattr(self, tree_attr, tree)
        editor = ttk.LabelFrame(parent, text=f"Editar {title}", padding=8)
        editor.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        editor.grid_columnconfigure(1, weight=1)
        editor.grid_columnconfigure(3, weight=1)
        editor_vars = {col: tk.StringVar(value="") for col in columns if col != "id"}
        for idx, col in enumerate([c for c in columns if c != "id"]):
            ttk.Label(editor, text=col).grid(row=idx // 2, column=(idx % 2) * 2, sticky="w", padx=(0, 8), pady=2)
            ttk.Entry(editor, textvariable=editor_vars[col]).grid(row=idx // 2, column=(idx % 2) * 2 + 1, sticky="ew", padx=(0, 12), pady=2)
        tree._editor_vars = editor_vars  # type: ignore[attr-defined]
        tree.bind("<<TreeviewSelect>>", lambda _e, t=tree, cols=columns: self._on_capacity_master_selected(t, cols))
        btns = ttk.Frame(editor)
        btns.grid(row=(len(columns) - 2) // 2 + 1, column=0, columnspan=4, sticky="w", pady=(6, 0))
        ttk.Button(btns, text="Nuevo", command=lambda t=tree, cols=columns: self._capacity_master_new_row(t, cols)).pack(side="left", padx=4)
        ttk.Button(btns, text="Aplicar cambios", command=lambda t=tree, cols=columns: self._capacity_master_apply_row(t, cols)).pack(side="left", padx=4)
        ttk.Button(btns, text="Eliminar", command=lambda t=tree: self._delete_selected(t, "Seleccione un registro para eliminar.")).pack(side="left", padx=4)
        ttk.Button(btns, text="Guardar cambios", command=lambda t=tree, cols=columns, sv=saver, ld=loader, k=master_key: self._save_capacity_master(t, cols, sv, ld, k)).pack(side="left", padx=4)
        ttk.Button(btns, text="Restaurar valores por defecto", command=lambda rs=resetter: self._reset_capacity_master(rs)).pack(side="left", padx=4)

    def _on_capacity_master_selected(self, tree, columns) -> None:
        if not tree.selection():
            return
        vals = tree.item(tree.selection()[0], "values")
        for idx, col in enumerate(columns):
            if col == "id":
                continue
            tree._editor_vars[col].set(vals[idx] if idx < len(vals) else "")  # type: ignore[attr-defined]

    def _capacity_master_new_row(self, tree, columns) -> None:
        values = [""] + [tree._editor_vars[col].get() for col in columns if col != "id"]  # type: ignore[attr-defined]
        tree.insert("", "end", values=values)

    def _capacity_master_apply_row(self, tree, columns) -> None:
        if not tree.selection():
            return
        current = list(tree.item(tree.selection()[0], "values"))
        values = []
        for idx, col in enumerate(columns):
            values.append(current[idx] if col == "id" and idx < len(current) else tree._editor_vars[col].get())  # type: ignore[attr-defined]
        tree.item(tree.selection()[0], values=values)

    def _save_capacity_master(self, tree, columns, saver, loader, master_key: str) -> None:
        rows = []
        for item in tree.get_children():
            values = tree.item(item, "values")
            rows.append({col: values[idx] if idx < len(values) else "" for idx, col in enumerate(columns) if col != "id"})
        saver(rows)
        self._load_capacity_productive_settings()
        messagebox.showinfo("Configuración productiva", f"Maestro {master_key} guardado correctamente.", parent=self)

    def _reset_capacity_master(self, resetter) -> None:
        resetter()
        self._load_capacity_productive_settings()
        messagebox.showinfo("Configuración productiva", "Valores por defecto restaurados.", parent=self)

    def _load_capacity_productive_settings(self) -> None:
        mapping = [
            (self._productive_families_tree, self.service.get_productive_families, ("id", "codigo", "descripcion", "orden", "activa", "observaciones", "updated_at")),
            (self._line_capacity_config_tree, self.service.get_line_capacity_config, ("id", "linea_productiva", "familia_productiva", "puesto_productivo_principal", "modo_uso_recursos", "usar_capacidad_agregada", "activa", "observaciones", "updated_at")),
            (self._line_required_resources_tree, self.service.get_line_required_resources, ("id", "linea_productiva", "recurso_codigo", "obligatorio", "modo_uso", "reparte_kg", "orden", "activo", "observaciones", "updated_at")),
            (self._staff_area_equivalences_tree, self.service.get_staff_area_equivalences, ("id", "area_requerida", "area_personal", "prioridad", "activa", "observaciones", "updated_at")),
            (self._staff_polyvalence_tree, self.service.get_staff_polyvalence, ("id", "puesto_origen", "puesto_destino", "prioridad", "factor_productividad", "activa", "observaciones", "updated_at")),
        ]
        for tree, loader, columns in mapping:
            if not tree:
                continue
            tree.delete(*tree.get_children())
            for row in loader():
                tree.insert("", "end", values=[row.get(col, "") for col in columns])


    def _get_active_capacity_master_actions(self):
        if not self._capacity_productive_notebook:
            return None
        selected = self._capacity_productive_notebook.select()
        if not selected:
            return None
        title = self._capacity_productive_notebook.tab(selected, "text")
        return self._capacity_master_actions.get(title)

    def _capacity_productive_sheet_specs(self) -> list[dict]:
        cfg = PRODUCTION_MASTER_EXCEL_CONFIGS
        return [
            {"sheet": "FamiliasProductivas", "key": "productive_families", "loader": self.service.get_productive_families, "saver": self.service.save_productive_families},
            {"sheet": "ConfigLineasCapacidad", "key": "line_capacity_config", "loader": self.service.get_line_capacity_config, "saver": self.service.save_line_capacity_config},
            {"sheet": "RecursosRequeridosLinea", "key": "line_required_resources", "loader": self.service.get_line_required_resources, "saver": self.service.save_line_required_resources},
            {"sheet": "EquivalenciasPuestos", "key": "staff_area_equivalences", "loader": self.service.get_staff_area_equivalences, "saver": self.service.save_staff_area_equivalences},
            {"sheet": "PolivalenciasPersonal", "key": "staff_polyvalence", "loader": self.service.get_staff_polyvalence, "saver": self.service.save_staff_polyvalence},
        ]

    def _export_capacity_productive_excel(self) -> None:
        help_rows = [{"campo": i.get("title", ""), "descripcion": i.get("description", "")} for i in get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP)]
        self._export_master_excel("capacity_productive", lambda: {
            spec["sheet"]: {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS[spec["key"]]["columns"], "rows": spec["loader"]()}
            for spec in self._capacity_productive_sheet_specs()
        } | {"Descripcion campos": {"columns": ["campo", "descripcion"], "rows": help_rows}})

    def _import_capacity_productive_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        wb = load_workbook(path, data_only=True)
        cfg = PRODUCTION_MASTER_EXCEL_CONFIGS
        errors = []

        def read(sheet_name: str, config: dict) -> tuple[list[dict], bool]:
            if sheet_name not in wb.sheetnames:
                return [], False
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return [], True
            raw_headers = [str(h or "").strip() for h in header_row]
            headers, _header_aliases = normalize_headers(raw_headers)
            missing = [col for col in config.get("required_columns", []) if col not in headers]
            if missing:
                errors.append(f"{sheet_name}: faltan columnas obligatorias: {', '.join(missing)}")
            rows = []
            seen = set()
            for row_idx, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row = {headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
                clean = {c: row.get(c, "") for c in config["columns"]}
                if all(is_blank(clean.get(c)) for c in config["columns"] if c != "id"):
                    continue
                if sheet_name == "RecursosRequeridosLinea":
                    print(
                        "[IMPORT DEBUG RecursosRequeridosLinea] "
                        f"fila={row_idx} headers={headers} row_dict={row} "
                        f"obligatorio_raw={row.get('obligatorio')} reparte_kg_raw={row.get('reparte_kg')} "
                        f"obligatorio_normalized={normalize_bool(row.get('obligatorio'))} "
                        f"reparte_kg_normalized={normalize_bool(row.get('reparte_kg'))}"
                    )
                for column in config.get("required_columns", []):
                    if is_blank(clean.get(column)):
                        errors.append(f"{sheet_name} fila {row_idx}: {column} obligatorio.")
                for column in config.get("numeric_columns", []):
                    raw = clean.get(column)
                    if raw in (None, ""):
                        clean[column] = 0.0
                    else:
                        try:
                            clean[column] = float(str(raw).replace(",", "."))
                        except ValueError:
                            errors.append(f"{sheet_name} fila {row_idx}: {column} no es numérico válido ({raw}).")
                for column in config.get("boolean_columns", []):
                    clean[column] = normalize_bool(clean.get(column))
                unique_key = config.get("unique_key", "")
                if unique_key:
                    key = "|".join(str(clean.get(part, "")).strip().lower() for part in unique_key.split("|"))
                    if key in seen:
                        errors.append(f"{sheet_name} fila {row_idx}: clave duplicada {unique_key}")
                    seen.add(key)
                rows.append(clean)
            return rows, True

        imported_by_key = {}
        specs = self._capacity_productive_sheet_specs()
        for spec in specs:
            rows, present = read(spec["sheet"], cfg[spec["key"]])
            if present:
                imported_by_key[spec["key"]] = rows

        if not imported_by_key and "Datos" in wb.sheetnames:
            datos_headers, _datos_aliases = normalize_headers([str(h or "").strip() for h in next(wb["Datos"].iter_rows(min_row=1, max_row=1, values_only=True))])
            for spec in specs:
                config = cfg[spec["key"]]
                if all(column in datos_headers for column in config.get("required_columns", [])):
                    rows, _present = read("Datos", config)
                    imported_by_key[spec["key"]] = rows
                    break

        if not imported_by_key:
            messagebox.showerror("Capacidad productiva", "El Excel no contiene hojas de Capacidad productiva reconocidas.", parent=self)
            return
        if errors:
            messagebox.showerror("Errores de validación", "\n".join(errors[:20]), parent=self)
            return

        def merge_rows(existing, imported, unique_key: str):
            def key_for(row):
                return "|".join(str(row.get(part, "")).strip().lower() for part in unique_key.split("|"))
            merged = {key_for(row): row for row in existing if key_for(row)}
            for row in imported:
                key = key_for(row)
                if key:
                    merged[key] = row
            return list(merged.values())

        total = sum(len(rows) for rows in imported_by_key.values())
        if not messagebox.askyesno("Confirmar importación", f"Se importarán {total} registros de Capacidad productiva actualizando por clave y sin borrar ausentes. ¿Continuar?", parent=self):
            return
        spec_by_key = {spec["key"]: spec for spec in specs}
        for key, imported in imported_by_key.items():
            spec = spec_by_key[key]
            merged = merge_rows(spec["loader"](), imported, cfg[key]["unique_key"])
            spec["saver"](merged)
        self._load_capacity_productive_settings()
        messagebox.showinfo("Configuración productiva", f"Importación completada. Registros procesados: {total}.", parent=self)

    def _show_capacity_productive_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Capacidad productiva", heading="Capacidad productiva", intro="Estos maestros describen la realidad operativa usada por el motor de capacidad sin reglas fijas en código.", help_items=get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP))

    def _build_tab_toolbar(self, parent, help_command, export_command=None, import_command=None):
        bar = ttk.Frame(parent)
        bar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        bar.grid_columnconfigure(1, weight=1)
        left = ttk.Frame(bar)
        left.grid(row=0, column=0, sticky="w")
        if export_command:
            ttk.Button(left, text="Exportar Excel", command=export_command).pack(side="left", padx=(0, 4))
        if import_command:
            ttk.Button(left, text="Importar Excel", command=import_command).pack(side="left", padx=(0, 8))
        ttk.Button(bar, text="ⓘ Descripción de campos", command=help_command).grid(row=0, column=2, sticky="e")

    def _export_master_excel(self, master_key: str, rows_loader) -> None:
        config = PRODUCTION_MASTER_EXCEL_CONFIGS[master_key]
        rows_data = rows_loader()
        help_items = self._get_master_help_items(master_key)
        if not isinstance(rows_data, dict):
            out_path = export_master_to_excel(rows=rows_data, config=config, help_items=help_items)
            if out_path:
                messagebox.showinfo("Configuración productiva", f"Excel exportado en:\n{out_path}", parent=self)
            return
        target = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=config.get("default_filename", "maestro.xlsx"), filetypes=[("Excel", "*.xlsx")])
        if not target:
            return
        wb = Workbook()
        if isinstance(rows_data, dict):
            first=True
            for sheet_name, payload in rows_data.items():
                ws = wb.active if first else wb.create_sheet()
                first=False
                ws.title = sheet_name
                columns = payload["columns"]
                ws.append(columns)
                for row in payload["rows"]: ws.append([row.get(c, "") for c in columns])
                ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
                for i, col in enumerate(columns, start=1):
                    width = max(len(str(col)), max((len(str(v or "")) for v in [r.get(col, "") for r in payload["rows"]]), default=0))
                    ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)
        else:
            ws = wb.active; ws.title = config["sheet_name"]; cols = config["columns"]; ws.append(cols)
            for row in rows_data: ws.append([row.get(c, "") for c in cols])
            ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
            for i, col in enumerate(cols, start=1): ws.column_dimensions[get_column_letter(i)].width = min(max(len(col), 14) + 2, 60)
        wb.save(target)
        messagebox.showinfo("Configuración productiva", f"Excel exportado en:\n{target}", parent=self)

    def _import_master_excel(self, master_key: str, saver, reload_callback, extra_validator=None) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path: return
        config = PRODUCTION_MASTER_EXCEL_CONFIGS[master_key]
        wb = load_workbook(path, data_only=True)
        errors=[]
        def _sheet_to_rows(ws, cfg):
            headers=[str(h or "").strip() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            data=[]; seen=set()
            for row_idx, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row={headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
                clean={c: row.get(c, "") for c in cfg["columns"]}
                for h in headers:
                    clean.setdefault(h, row.get(h, ""))
                for c in cfg.get("numeric_columns", []):
                    if clean.get(c) not in (None, ""): clean[c]=float(str(clean[c]).replace(",","."))
                for c in cfg.get("boolean_columns", []): clean[c]=1 if str(clean.get(c,"0")).strip().upper() in {"1","SI","S","TRUE","X","Y","YES"} else 0
                unique_key = cfg["unique_key"]
                if "|" in unique_key:
                    uk = "|".join(str(clean.get(part, "")).strip().lower() for part in unique_key.split("|"))
                else:
                    uk=str(clean.get(unique_key, "")).strip().lower()
                if uk in seen: errors.append(f"{ws.title} fila {row_idx}: clave duplicada {unique_key}")
                seen.add(uk); data.append(clean)
            return data
        personal_summary = None
        if master_key == "personal":
            areas = _sheet_to_rows(wb[config["sheet_name"]], config)
            for idx, row in enumerate(areas, start=2):
                try:
                    row["tipo_personal"] = self._staff_type_from_excel_row(row)
                    row.update(staff_type_flags(row["tipo_personal"]))
                except Exception as exc:
                    errors.append(f"{config['sheet_name']} fila {idx}: {exc}")
            s_cfg=config["extra_sheets"]["Resumen personal"]
            if "Resumen personal" in wb.sheetnames:
                summary_rows = _sheet_to_rows(wb["Resumen personal"], {**s_cfg, "columns": s_cfg["columns"]})
                if summary_rows:
                    personal_summary = summary_rows[0]
            rows=areas
        else:
            rows = _sheet_to_rows(wb.active if config["sheet_name"] not in wb.sheetnames else wb[config["sheet_name"]], config)
        if extra_validator: errors.extend(extra_validator(rows))
        if errors: return messagebox.showerror("Errores de validación", "\n".join(errors[:20]), parent=self)
        if not messagebox.askyesno("Confirmar importación", f"Se importarán {len(rows)} registros. ¿Continuar?", parent=self): return
        saver(rows)
        if master_key == "personal" and personal_summary:
            self.service.save_staff_summary(personal_summary)
        reload_callback(); messagebox.showinfo("Configuración productiva", f"Importación completada. Registros procesados: {len(rows)}.", parent=self)

    def _load_personal_rows_for_excel(self):
        area_rows = []
        for row in self.service.get_staff_areas():
            export_row = dict(row)
            tipo = self._normalize_staff_type(export_row.get("tipo_personal"))
            export_row.update(staff_type_flags(tipo))
            area_rows.append(export_row)
        return {
            "Resumen personal": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["personal"]["extra_sheets"]["Resumen personal"]["columns"], "rows": [self.service.get_staff_summary()]},
            "Areas personal": {"columns": PRODUCTION_MASTER_EXCEL_CONFIGS["personal"]["columns"], "rows": area_rows},
        }

    def _staff_type_from_excel_row(self, row: dict) -> str:
        tipo = str(row.get("tipo_personal") or "").strip()
        if tipo:
            return self._normalize_staff_type(tipo)
        inferred = infer_default_staff_type(row.get("area"))
        if inferred:
            return inferred
        active_flags = []
        for key, label in (("Directo", "Directo"), ("Soporte", "Soporte"), ("Indirecto", "Indirecto"), ("directo", "Directo"), ("soporte", "Soporte"), ("indirecto", "Indirecto")):
            raw = row.get(key)
            if str(raw).strip().upper() in {"1", "SI", "S", "SÍ", "TRUE", "X", "Y", "YES"}:
                active_flags.append(label)
        if len(set(active_flags)) == 1:
            return active_flags[0]
        raise ValueError(f"Tipo personal obligatorio para el área {row.get('area', '')}")

    def _save_personal_rows_from_excel(self, rows):
        self.service.save_staff_areas(rows)

    def _validate_semaphore_import_rows(self, rows):
        errors=[]
        for i, row in enumerate(rows, start=2):
            op=row.get("operador"); y=row.get("umbral_amarillo",0); r=row.get("umbral_rojo",0)
            if op in (">", ">=") and y > r: errors.append(f"Fila {i}: umbral amarillo no puede ser mayor que rojo para {op}.")
            if op in ("<", "<=") and y < r: errors.append(f"Fila {i}: umbral amarillo no puede ser menor que rojo para {op}.")
        return errors

    def _validate_caliber_import_rows(self, rows):
        errors=[]; seen={}
        for i,row in enumerate(rows, start=2):
            row["calibres_incluidos"] = self._normalize_calibres(str(row.get("calibres_incluidos", "")))
            key=str(row.get("confeccion_familia","")).strip().lower()
            for cal in row["calibres_incluidos"].split(","):
                if (key, cal) in seen: errors.append(f"Fila {i}: calibre duplicado '{cal}' en la confección {row.get('confeccion_familia')}")
                seen[(key,cal)] = i
        return errors

    def _validate_mapping_import_rows(self, rows):
        errors=[]
        for i,row in enumerate(rows, start=2):
            if not str(row.get("codigo_mconfeccion","")).strip(): errors.append(f"Fila {i}: codigo_mconfeccion obligatorio")
            if str(row.get("kg_formato",0)) and float(row.get("kg_formato",0)) < 0: errors.append(f"Fila {i}: kg_formato debe ser >= 0")
        return errors
    def _show_packaging_mapping_help(self) -> None:
        show_tab_help(self, title="Descripción de campos - Mapeo confecciones", heading="Mapeo confecciones", intro="Esta pestaña relaciona las confecciones comerciales de MConfecciones con la interpretación productiva que usará el cálculo de capacidad. Permite corregir manualmente casos dudosos, especialmente mallas, subtipo, kg formato y línea productiva.", help_items=get_help_items(PRODUCTION_PACKAGING_MAPPING_HELP_KEYS, PRODUCTION_PACKAGING_MAPPING_HELP))

    def _get_master_help_items(self, master_key: str) -> list[dict]:
        by_key = {
            "personal": get_help_items(PRODUCTION_PERSONAL_HELP_KEYS, PRODUCTION_PERSONAL_HELP),
            "flow_staffing": get_help_items(PRODUCTION_FLOW_STAFFING_HELP_KEYS, PRODUCTION_FLOW_STAFFING_HELP),
            "packaging_types": get_help_items(PRODUCTION_PACKAGING_HELP_KEYS, PRODUCTION_PACKAGING_HELP),
            "packaging_mapping": get_help_items(PRODUCTION_PACKAGING_MAPPING_HELP_KEYS, PRODUCTION_PACKAGING_MAPPING_HELP),
            "lines": get_help_items(PRODUCTION_LINES_HELP_KEYS, PRODUCTION_LINES_HELP),
            "performance_rules": get_help_items(PRODUCTION_PERFORMANCE_HELP_KEYS, PRODUCTION_PERFORMANCE_HELP),
            "penalty_rules": get_help_items(PRODUCTION_PENALTIES_HELP_KEYS, PRODUCTION_PENALTIES_HELP),
            "semaphore_rules": get_help_items(PRODUCTION_SEMAPHORE_HELP_KEYS, PRODUCTION_SEMAPHORE_HELP),
            "caliber_factors": get_help_items(PRODUCTION_CALIBER_FACTORS_HELP_KEYS, PRODUCTION_CALIBER_FACTORS_HELP),
            "physical_resources": get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP),
            "resource_compatibilities": get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP),
            "resource_feeds": get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP),
            "resource_availability": get_help_items(PRODUCTION_RESOURCES_HELP_KEYS, PRODUCTION_RESOURCES_HELP),
            "productive_families": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
            "line_capacity_config": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
            "line_required_resources": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
            "staff_area_equivalences": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
            "staff_polyvalence": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
            "capacity_productive": get_help_items(PRODUCTION_CAPACITY_MASTERS_HELP_KEYS, PRODUCTION_CAPACITY_MASTERS_HELP),
        }
        return by_key.get(master_key, [])

    def _export_packaging_mapping_to_excel(self) -> None:
        try:
            rows = self.service.get_packaging_mapping(False)
            out_path = export_master_to_excel(rows=rows, config=PRODUCTION_MASTER_EXCEL_CONFIGS["packaging_mapping"], help_items=self._get_master_help_items("packaging_mapping"))
            if out_path:
                messagebox.showinfo("Configuración productiva", f"Excel exportado en:\n{out_path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Error exportando Excel", str(exc), parent=self)

    def _import_packaging_mapping_from_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            normalized_rows, errors = import_master_from_excel(path=path, config=PRODUCTION_MASTER_EXCEL_CONFIGS["packaging_mapping"])
            for idx, row in enumerate(normalized_rows, start=2):
                if row.get("kg_formato", 0) < 0:
                    errors.append(f"Fila {idx}: kg_formato debe ser >= 0.")
            if errors:
                error_preview = "\n".join(errors[:20])
                more = "\n..." if len(errors) > 20 else ""
                messagebox.showerror("Errores de validación", f"No se pudo importar el Excel:\n{error_preview}{more}", parent=self)
                return
            if not normalized_rows:
                messagebox.showerror("Importación", "El archivo no contiene registros para importar.", parent=self)
                return
            if not messagebox.askyesno("Confirmar importación", f"Se importarán {len(normalized_rows)} registros y se actualizará el mapeo de confecciones. ¿Continuar?", parent=self):
                return
            self.service.save_packaging_mapping(normalized_rows)
            self._load_packaging_mapping_settings(False)
            messagebox.showinfo("Configuración productiva", f"Importación completada. Registros procesados: {len(normalized_rows)}.", parent=self)
        except Exception as exc:
            messagebox.showerror("Error importando Excel", str(exc), parent=self)
