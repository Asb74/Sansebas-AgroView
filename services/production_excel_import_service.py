from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re

from openpyxl import load_workbook

from db.production_settings_repository import ProductionSettingsRepository


@dataclass
class ExcelImportResult:
    sheets_found: list[str] = field(default_factory=list)
    sheets_missing: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    imported_counts: dict[str, int] = field(default_factory=dict)


class ProductionExcelImportService:
    REQUIRED_SHEETS = ["Simulador_Personal", "Ranking_Volcado", "Resumen_Calibres"]

    def __init__(self, repository: ProductionSettingsRepository | None = None) -> None:
        self.repository = repository or ProductionSettingsRepository()

    def import_rules_from_excel(self, file_path: str, replace_existing: bool) -> ExcelImportResult:
        result = ExcelImportResult()
        wb = load_workbook(Path(file_path), data_only=True, read_only=True)
        result.sheets_found = [s for s in self.REQUIRED_SHEETS if s in wb.sheetnames]
        result.sheets_missing = [s for s in self.REQUIRED_SHEETS if s not in wb.sheetnames]

        if "Simulador_Personal" in wb.sheetnames:
            sheet = wb["Simulador_Personal"]
            rows = self._rows(sheet)
            staff, perf, penalty, sem = self._parse_simulador(rows, result)
            if staff:
                self.repository.save_staff_areas(staff if replace_existing else self.repository.get_staff_areas() + staff)
            if perf:
                self.repository.save_performance_rules(perf if replace_existing else self._merge_by_key(self.repository.get_performance_rules(), perf, "codigo"))
            if penalty:
                self.repository.save_penalty_rules(penalty if replace_existing else self._merge_by_key(self.repository.get_penalty_rules(), penalty, "codigo"))
            if sem:
                self.repository.save_semaphore_rules(sem if replace_existing else self._merge_by_key(self.repository.get_semaphore_rules(), sem, "codigo"))
            result.imported_counts.update({"staff": len(staff), "performance": len(perf), "penalties": len(penalty), "semaphore": len(sem)})

        if "Ranking_Volcado" in wb.sheetnames:
            ranking = self._parse_ranking(self._rows(wb["Ranking_Volcado"]))
            self.repository.save_unloading_priority_rules(ranking, replace_existing=replace_existing)
            result.imported_counts["ranking"] = len(ranking)

        if "Resumen_Calibres" in wb.sheetnames:
            sem_extra = self._parse_resumen_calibres(self._rows(wb["Resumen_Calibres"]))
            if sem_extra:
                payload = sem_extra if replace_existing else self._merge_by_key(self.repository.get_semaphore_rules(), sem_extra, "codigo")
                self.repository.save_semaphore_rules(payload)
            result.imported_counts["resumen_calibres"] = len(sem_extra)

        wb.close()
        return result

    def _rows(self, sheet):
        return [[c if c is not None else "" for c in row] for row in sheet.iter_rows(values_only=True)]

    def _parse_simulador(self, rows, result):
        staff = []
        expected_areas = ["Volcado", "Tría principal", "Tría mallas", "Mallas", "Encajado", "Granel manual", "Granelera", "Auxiliares", "Calidad", "Expedición", "Carretilleros", "Encargados"]
        text = "\n".join(" ".join(str(c) for c in row if c != "") for row in rows).lower()
        for idx, area in enumerate(expected_areas, start=1):
            found = area.lower() in text
            staff.append({"id": idx, "area": area, "tipo_personal": "Directo" if area in {"Volcado", "Tría principal", "Tría mallas", "Mallas", "Encajado", "Granel manual", "Granelera"} else "Indirecto", "disponible": 0, "minimo_operativo": 0, "optimo": 0, "activo": 1, "observaciones": "Importado desde Simulador_Personal" if found else "Área no localizada explícitamente; creada por plantilla base."})
        if "box" in text:
            result.warnings.append("Detectado criterio BOX en Simulador_Personal.")
        perf = [
            {"codigo": "IMP_MALLA", "familia": "Malla", "confeccion_formato": "Malla", "tipo_linea": "Malla", "condicion": "Importado", "oph_referencia": 398.0, "oph_minimo": 300.0, "oph_optimo": 465.0, "kg_h_referencia": 0.0, "factor_precalibrado": 1.0, "factor_destrio_alto": 0.9, "dificultad": "Media", "activo": 1, "observaciones": "Base de importación desde Excel."},
            {"codigo": "IMP_ENCAJADO", "familia": "Encajado", "confeccion_formato": "Encajado", "tipo_linea": "Encajado", "condicion": "Importado", "oph_referencia": 250.0, "oph_minimo": 200.0, "oph_optimo": 300.0, "kg_h_referencia": 0.0, "factor_precalibrado": 1.0, "factor_destrio_alto": 0.9, "dificultad": "Media", "activo": 1, "observaciones": "Base de importación desde Excel."},
        ]
        penalties = [
            {"codigo": "IMP_CAMBIO_FORMATO", "tipo_penalizacion": "Cambio formato kg", "ambito": "General", "minutos_perdida": 15.0, "factor_rendimiento": 1.0, "aplica_por": "Cada cambio", "umbral": "cambio formato", "activa": 1, "observaciones": "Importado desde Simulador_Personal."},
            {"codigo": "IMP_PEDIDO_PEQUENO", "tipo_penalizacion": "Pedido pequeño", "ambito": "General", "minutos_perdida": 8.0, "factor_rendimiento": 0.9, "aplica_por": "Cada pedido", "umbral": "pedido pequeño", "activa": 1, "observaciones": "Importado desde Simulador_Personal."},
        ]
        sem = [
            {"codigo": "IMP_PERSONAL_INSUF", "tipo_regla": "Falta personal", "ambito": "Personal", "metrica": "personas_faltantes", "operador": ">", "umbral_amarillo": 0.0, "umbral_rojo": 3.0, "accion_sugerida": "Reforzar personal o bajar carga.", "activa": 1, "observaciones": "Importada desde Simulador_Personal."},
            {"codigo": "IMP_CAPACIDAD_OK", "tipo_regla": "Saturación capacidad", "ambito": "General", "metrica": "ocupacion_pct", "operador": ">=", "umbral_amarillo": 85.0, "umbral_rojo": 100.0, "accion_sugerida": "Revisar cambios y carga.", "activa": 1, "observaciones": "Importada desde Simulador_Personal."},
        ]
        return staff, perf, penalties, sem

    def _parse_ranking(self, rows):
        criterios = ["mayor cobertura de pedidos", "menor destrío", "calibre dominante necesario", "variedad demandada", "fecha salida próxima", "kg útiles estimados", "riesgo de sobrante"]
        text = "\n".join(" ".join(str(c) for c in row if c != "") for row in rows).lower()
        data = []
        for c in criterios:
            data.append({"criterio": c.title(), "descripcion": "Importado desde Ranking_Volcado.", "peso": 1.0 if c in text else 0.8, "activo": 1, "observaciones": ""})
        return data

    def _parse_resumen_calibres(self, rows):
        text = "\n".join(" ".join(str(c) for c in row if c != "") for row in rows).lower()
        kg_diff = 0.0
        m = re.search(r"diferencia\s*[:=]?\s*(-?\d+[\.,]?\d*)", text)
        if m:
            kg_diff = float(m.group(1).replace(",", "."))
        return [{"codigo": "IMP_EXCESO_KG", "tipo_regla": "Exceso pedidos", "ambito": "General", "metrica": "kg_pendientes", "operador": ">=", "umbral_amarillo": max(0.0, kg_diff), "umbral_rojo": max(1000.0, kg_diff * 1.2 if kg_diff else 1000.0), "accion_sugerida": "Ajustar cobertura por calibre y replanificar.", "activa": 1, "observaciones": "Derivada desde Resumen_Calibres."}]

    def _merge_by_key(self, existing, incoming, key):
        merged = {row[key]: row for row in existing}
        for row in incoming:
            merged[row[key]] = row
        return list(merged.values())
