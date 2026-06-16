from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import Font

from db.planning_repository import PlanningRepository
from services.legacy_sync_service import LegacySyncService


class PlanningService:
    def __init__(self) -> None:
        self.repo = PlanningRepository()
        self.legacy_sync = LegacySyncService()

    def load_stock_campo(self, filters: dict) -> tuple[list[dict], str | None, bool]:
        return self.repo.get_stock_campo(filters)

    def load_stock_almacen(self, filters: dict) -> tuple[list[dict], str | None]:
        return self.repo.get_stock_almacen(filters)

    def load_stock_almacen_detalle_palets(self, filters: dict) -> list[dict]:
        return self.repo.get_stock_almacen_detalle_palets(filters)

    def load_pedidos_pendientes(self, filters: dict, modo_pedidos: str = "10_dias") -> tuple[list[dict], dict]:
        return self.repo.get_pedidos_pendientes(filters, modo_pedidos=modo_pedidos)

    def load_balance_planificacion(self, filters: dict, policy: dict | None = None) -> list[dict]:
        return self.repo.get_balance_planificacion(filters, policy=policy)

    def get_balance_cobertura_detalle(self, filters: dict, balance_row: dict, policy: dict | None = None) -> list[dict]:
        return self.repo.get_balance_cobertura_detalle(filters, balance_row, policy=policy)

    def get_inventario_operativo_global(self, filters: dict, policy: dict | None = None) -> list[dict]:
        return self.repo.get_inventario_operativo_global(filters, policy=policy)

    def get_candidatos_compatibles_para_pedido(self, filters: dict, pedido: dict, policy_cfg: dict | None = None) -> list[dict]:
        return self.repo.get_candidatos_compatibles_para_pedido(filters, pedido, policy_cfg=policy_cfg)


    def get_pedidos_pendientes(self, filters: dict, modo: str = "10_dias") -> tuple[list[dict], dict]:
        return self.load_pedidos_pendientes(filters, modo_pedidos=modo)

    def get_aprovechamientos_reales(self, filters: dict) -> list[dict]:
        return self.repo.get_aprovechamientos_reales(filters)

    def get_aprovechamiento_stock_campo(self, stock_campo_rows: list[dict], filters: dict) -> tuple[dict[tuple[str, str, str, str, str, float], dict], dict[str, list[dict]]]:
        return self.repo.build_aprovechamiento_stock_campo(stock_campo_rows, filters)

    def get_aprovechamientos_estimados_por_boleta(self, boleta: str) -> list[dict]:
        return self.repo.get_aprovechamientos_estimados_por_boleta(boleta)

    def get_aprovechamientos_estimados_filtrados(self, filters: dict | None = None, boletas: list | None = None) -> list[dict]:
        return self.repo.get_aprovechamientos_estimados_filtrados(filters, boletas=boletas)

    def upsert_aprovechamiento_estimado(self, row: dict) -> list[int]:
        return self.repo.upsert_aprovechamiento_estimado(row)

    def delete_aprovechamiento_estimado(self, id: int) -> None:
        self.repo.delete_aprovechamiento_estimado(id)

    def normalizar_calibre_a_set(self, calibre: str) -> set[str]:
        return self.repo.normalizar_calibre_a_set(calibre)

    def diagnose_loteado_tables(self) -> dict:
        return self.repo.diagnose_loteado_tables()

    def get_filter_options(self, key: str) -> list[str]:
        return self.repo.get_filter_options(key)

    def get_filter_options_contextual(self, key: str, filters: dict) -> list[str]:
        return self.repo.get_filter_options_contextual(key, filters)

    def get_correspondencias_calibres(self, cultivo: str) -> list[dict]:
        return self.repo.get_correspondencias_calibres(cultivo)

    def aggregate_stock_campo(self, rows: list[dict]) -> list[dict]:
        agg: dict[tuple, float] = defaultdict(float)
        for r in rows:
            key = (r.get("Cultivo", ""), r.get("Variedad", ""), r.get("Grupo varietal", ""), r.get("Semana", ""))
            agg[key] += float(r.get("Kg campo", 0) or 0)
        return [{"Cultivo": k[0], "Variedad": k[1], "Grupo varietal": k[2], "Semana": k[3], "Kg campo": round(v, 2)} for k, v in agg.items()]

    def aggregate_stock_almacen(self, rows: list[dict]) -> list[dict]:
        agg: dict[tuple, float] = defaultdict(float)
        for r in rows:
            key = (r.get("Cultivo", ""), r.get("Variedad", ""), r.get("Calibre", ""), r.get("Categoría", ""), r.get("IdConfeccion", ""))
            agg[key] += float(r.get("Kg stock", 0) or 0)
        return [{"Cultivo": k[0], "Variedad": k[1], "Calibre": k[2], "Categoría": k[3], "IdConfeccion": k[4], "Kg stock": round(v, 2)} for k, v in agg.items()]

    def export_rows_to_excel(self, rows: list[dict], tab_name: str, cultivos: list[str], campanas: list[str]) -> str | None:
        if not rows:
            return None
        if tab_name == "Stock campo":
            suffix = "Stock_campo"
        elif tab_name == "Stock almacén":
            suffix = "Stock_almacen"
        elif tab_name == "Balance":
            suffix = "Balance"
        else:
            suffix = "Pedidos_pendientes"
        fecha = datetime.now().strftime("%Y%m%d")
        cultivo_txt = "_".join(cultivos) if cultivos else "TODOS"
        campana_txt = "_".join(campanas) if campanas else "TODOS"
        if tab_name == "Balance":
            base = f"{fecha} Balance_{cultivo_txt}_{campana_txt} disponibilidad"
        else:
            base = f"{fecha} {suffix}_{cultivo_txt}_{campana_txt} hechos disponibles"
        safe_base = re.sub(r'[\\/:*?"<>|]', "_", base)
        default_name = f"{safe_base}.xlsx"
        target = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel", "*.xlsx")])
        if not target:
            return None
        wb = Workbook()
        ws = wb.active
        ws.title = tab_name
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 30)
        for i, h in enumerate(headers, start=1):
            if h in ("Kg campo", "Kg stock", "Kg pedido teórico", "Kg hecho real", "Kg pendiente", "Merma kg", "Kg stock comercial", "Kg pedidos pendientes", "Diferencia comercial", "Kg stock industrial almacén", "Kg entrada estimada", "Kg base total estimada", "Kg cobertura exacta", "Kg cobertura agrupada", "Kg cobertura potencial total"):
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, i).number_format = "#,##0.00"
            elif h in ("% hecho", "% merma"):
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, i).number_format = "0.00"
        wb.save(Path(target))
        return target

    def actualizar_planificacion_hoy_en_adelante(self) -> tuple[bool, str]:
        return self.legacy_sync.sync_planificacion_hoy_en_adelante()
