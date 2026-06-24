from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from tkinter import filedialog

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def normalize_header(value) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")

def normalize_headers(headers) -> tuple[list[str], dict[str, str]]:
    normalized_headers = [normalize_header(header) for header in headers]
    aliases = {normalized: original for original, normalized in zip(headers, normalized_headers) if normalized}
    return normalized_headers, aliases


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def normalize_bool(value) -> int:
    if value is None:
        return 0
    normalized = str(value).strip().upper()
    if normalized in {"1", "S", "SI", "SÍ", "TRUE", "X", "Y", "YES", "ACTIVO"}:
        return 1
    if normalized in {"0", "N", "NO", "FALSE", "", "NONE", "INACTIVO"}:
        return 0
    return 1 if normalized else 0


def normalize_number(value) -> float:
    if value in (None, ""):
        raise ValueError("empty numeric")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if text == "":
        raise ValueError("empty numeric")
    return float(text)


def validate_required_columns(headers: list[str], required_columns: list[str]) -> None:
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")


def _adjust_column_widths(ws, max_width: int = 60) -> None:
    for col_idx, _ in enumerate(ws[1], start=1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            max_len = max(max_len, len(str(ws.cell(row=row_idx, column=col_idx).value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _write_help_sheet(wb: Workbook, master_title: str, help_items: list[dict] | None) -> None:
    ws = wb.create_sheet("Descripcion campos")
    ws["A1"] = master_title
    ws["A1"].font = Font(bold=True)
    if not help_items:
        ws["A3"] = "No hay descripción de campos configurada para este maestro."
        return
    headers = ["Campo", "Descripción", "Ejemplo", "Impacto operativo"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for item in help_items:
        ws.append([item.get("title", ""), item.get("description", ""), item.get("example", ""), item.get("impact", "")])
    ws.auto_filter.ref = f"A3:D{ws.max_row}"
    ws.freeze_panes = "A4"
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = wrap
    _adjust_column_widths(ws)


def export_master_to_excel(rows: list[dict], config: dict, help_items: list[dict] | None = None) -> str | None:
    target = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=config.get("default_filename", "maestro_productivo.xlsx"),
        filetypes=[("Excel", "*.xlsx")],
    )
    if not target:
        return None

    columns = config["columns"]
    wb = Workbook()
    ws = wb.active
    ws.title = config.get("sheet_name", "Maestro")
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _adjust_column_widths(ws)
    _write_help_sheet(wb, config.get("sheet_name", "Maestro"), help_items)

    output_path = Path(target)
    wb.save(output_path)
    return output_path.as_posix()


def import_master_from_excel(path: str, config: dict) -> tuple[list[dict], list[str]]:
    wb = load_workbook(path, data_only=True)
    sheet_name = config.get("sheet_name")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("El Excel no contiene cabeceras.")

    raw_headers = [str(h or "").strip() for h in header_row]
    headers, _aliases = normalize_headers(raw_headers)
    validate_required_columns(headers, config.get("required_columns", []))

    columns = config.get("columns", headers)
    numeric_columns = set(config.get("numeric_columns", []))
    boolean_columns = set(config.get("boolean_columns", []))
    required_columns = set(config.get("required_columns", []))

    normalized_rows: list[dict] = []
    errors: list[str] = []

    for excel_row_index, values in enumerate(rows_iter, start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        clean_row: dict = {column: row.get(column, "") for column in columns}

        for column in required_columns:
            if is_blank(row.get(column)):
                errors.append(f"Fila {excel_row_index}: {column} obligatorio.")

        for column in numeric_columns:
            raw = row.get(column)
            if raw in (None, ""):
                clean_row[column] = 0.0
                continue
            try:
                clean_row[column] = normalize_number(raw)
            except ValueError:
                errors.append(f"Fila {excel_row_index}: {column} no es numérico válido ({raw}).")

        for column in boolean_columns:
            clean_row[column] = normalize_bool(row.get(column))

        normalized_rows.append(clean_row)

    return normalized_rows, errors
