from __future__ import annotations

from pathlib import Path
from tkinter import filedialog

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def export_dict_rows_to_excel(
    rows: list[dict],
    headers: list[str],
    default_filename: str,
    sheet_name: str,
) -> str | None:
    target = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=default_filename,
        filetypes=[("Excel", "*.xlsx")],
    )
    if not target:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            value = cell[0].value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    output_path = Path(target)
    wb.save(output_path)
    return output_path.as_posix()


def import_dict_rows_from_excel(path: str, required_columns: list[str]) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError("El Excel no contiene cabeceras.")

    headers = [str(h or "").strip() for h in header_row]
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    data: list[dict] = []
    for row in rows:
        row_dict = {}
        for idx, header in enumerate(headers):
            row_dict[header] = row[idx] if idx < len(row) else None
        data.append(row_dict)
    return data
